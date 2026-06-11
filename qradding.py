"""
qradding.py — Generate QR codes from CSV and embed into Excel.

Bug fixes:
1. csv_path and final_excel were empty strings (syntax error) — now use argparse / env vars.
2. Column-letter formula for cols > 26 was wrong (gave wrong cell addresses) — replaced with
   openpyxl's get_column_letter() which handles all column indices correctly.
3. QR data format is now consistent with what app.py expects for matching (uses "QR" column value).
4. Random unique_id used random.choices without seeding — fine for uniqueness but now checks for
   duplicates within the same batch.
5. Row height not set — images would overlap rows; now sets row height proportionally.
6. No progress indication — added simple print progress.
7. Excel was saved mid-loop if images were large (OOM risk on big csvs) — now saves once at end.
"""

from __future__ import annotations

import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

import argparse
import os
import random
import string

import pandas as pd
import qrcode
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter  # FIX: replaces broken chr() formula


# ── CLI args ──────────────────────────────────────────────────────────────────
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate QR codes and embed into Excel.")
    parser.add_argument("--csv",    required=False, default=os.environ.get("CSV_PATH"),
                        help="Path to input CSV file")
    parser.add_argument("--output", required=False, default=os.environ.get("OUTPUT_EXCEL"),
                        help="Path for output Excel file")
    parser.add_argument("--qr-dir", default="qrcodes_samples",
                        help="Folder to save QR code images (default: qrcodes_samples)")
    parser.add_argument("--id-length", type=int, default=8,
                        help="Length of generated Unique ID (default: 8)")
    return parser.parse_args()


# ── Unique ID generator ───────────────────────────────────────────────────────
_CHARSET = string.ascii_uppercase + string.digits

def generate_unique_ids(n: int, length: int = 8) -> list[str]:
    """Generate n collision-free unique IDs."""
    ids: set[str] = set()
    while len(ids) < n:
        ids.add("".join(random.choices(_CHARSET, k=length)))
    return list(ids)


# ── Main ───────────────────────────────────────────────────────────────────────
def main() -> None:
    args = parse_args()

    if not args.csv:
        print("❌  No CSV path provided. Use --csv <path> or set CSV_PATH env var.")
        sys.exit(1)
    if not args.output:
        print("❌  No output path provided. Use --output <path> or set OUTPUT_EXCEL env var.")
        sys.exit(1)
    if not os.path.exists(args.csv):
        print(f"❌  CSV not found: {args.csv}")
        sys.exit(1)

    # Column names
    EMAIL_COL  = "Email Address"
    NAME_COL   = "Name"
    REG_COL    = "Registration Number"
    UID_COL    = "Unique ID"
    QR_COL     = "QR"          # raw string stored in Excel for scanning

    # 1) Load CSV ─────────────────────────────────────────────────────────────
    df = pd.read_csv(args.csv)
    df.columns = df.columns.str.strip()

    for col in [EMAIL_COL, NAME_COL, REG_COL]:
        if col not in df.columns:
            print(f"❌  Missing column in CSV: '{col}'")
            sys.exit(1)

    # 2) Generate Unique IDs ───────────────────────────────────────────────────
    df[UID_COL] = generate_unique_ids(len(df), args.id_length)

    # 3) Build QR data strings and store ──────────────────────────────────────
    def make_qr_string(row: pd.Series) -> str:
        return (
            f"Name: {str(row[NAME_COL]).strip()}\n"
            f"Email: {str(row[EMAIL_COL]).strip()}\n"
            f"Reg No: {str(row[REG_COL]).strip()}\n"
            f"ID: {str(row[UID_COL]).strip()}"
        )

    df[QR_COL] = df.apply(make_qr_string, axis=1)

    # 4) Save to Excel (initial, without images) ───────────────────────────────
    df.to_excel(args.output, index=False)

    # 5) Create QR images ─────────────────────────────────────────────────────
    os.makedirs(args.qr_dir, exist_ok=True)
    print(f"Generating {len(df)} QR codes…")
    for i, row in df.iterrows():
        reg     = str(row[REG_COL]).strip()
        qr_str  = row[QR_COL]
        qr_img  = qrcode.make(qr_str)
        qr_path = os.path.join(args.qr_dir, f"{reg}.png")
        qr_img.save(qr_path)
        if (i + 1) % 50 == 0 or (i + 1) == len(df):
            print(f"  {i + 1}/{len(df)}")

    # 6) Embed QR images into Excel ────────────────────────────────────────────
    wb = load_workbook(args.output)
    ws = wb.active

    # Find or create QR image column (separate from the QR text column)
    qr_img_col_idx = ws.max_column + 1
    ws.cell(row=1, column=qr_img_col_idx).value = "QR Code Image"

    img_size_px = 100
    # Approximate row height: 1 pixel ≈ 0.75pt
    ws.row_dimensions[1].height = 20
    for idx, row in df.iterrows():
        excel_row = idx + 2
        reg       = str(row[REG_COL]).strip()
        img_path  = os.path.join(args.qr_dir, f"{reg}.png")

        if not os.path.exists(img_path):
            print(f"  ⚠️  QR image missing for {reg}, skipping embed.")
            continue

        xl_img          = XLImage(img_path)
        xl_img.width    = img_size_px
        xl_img.height   = img_size_px

        # FIX: use get_column_letter instead of broken chr() formula
        cell_addr = f"{get_column_letter(qr_img_col_idx)}{excel_row}"
        ws.add_image(xl_img, cell_addr)

        # Set row height so images don't overlap
        ws.row_dimensions[excel_row].height = img_size_px * 0.75

    # 7) Save final Excel ──────────────────────────────────────────────────────
    wb.save(args.output)
    print(f"\n✅  Done — saved to '{args.output}'")
    print(f"   QR images: {args.qr_dir}/")


if __name__ == "__main__":
    main()
