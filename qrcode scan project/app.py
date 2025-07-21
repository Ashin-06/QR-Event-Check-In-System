from flask import Flask, request, jsonify, render_template
from flask_socketio import SocketIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import threading, os, shutil, re

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")
lock = threading.Lock()

EXCEL_FILE       = "data_with_qrdemo.xlsx"     #Excel File Name
HIGHLIGHTED_FILE = "data_with_qrdemo_highlighted.xlsx"
LOG_FILE         = "scanned_log.csv"
SCAN_COL_NAME    = "Scanned Status"

# Load or initialize the CSV log
if os.path.exists(LOG_FILE):
    scanned_log = pd.read_csv(LOG_FILE)
else:
    scanned_log = pd.DataFrame(columns=["QR Data","Scan Count","Timestamps"])
# Ensure Timestamps column exists
if "Timestamps" not in scanned_log.columns:
    scanned_log["Timestamps"] = ""

def sort_log(df):
    # derive the last timestamp, sort desc, drop helper
    df["__last"] = df["Timestamps"].apply(
        lambda s: s.split(";")[-1] if isinstance(s, str) and s else ""
    )
    df.sort_values("__last", ascending=False, inplace=True)
    df.drop(columns="__last", inplace=True)

# initial sort
sort_log(scanned_log)

# make sure there's a highlighted copy to start
if not os.path.exists(HIGHLIGHTED_FILE):
    shutil.copy(EXCEL_FILE, HIGHLIGHTED_FILE)

@app.route('/')
def index():
    return render_template("index.html")

@app.route('/data_csv')
def data_csv():
    with lock:
        df = scanned_log.copy()
        df["Last Timestamp"] = df["Timestamps"].apply(
            lambda s: s.split(";")[-1] if isinstance(s, str) and s else ""
        )
        return jsonify(df.to_dict(orient='records'))

@app.route('/scan', methods=['POST'])
def scan():
    qr_data = request.json.get('qr_data')
    if not qr_data:
        return jsonify(message="No QR data provided."), 400

    try:
        with lock:
            now = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')

            # 1) Update CSV log
            if qr_data in scanned_log["QR Data"].values:
                idx = scanned_log.index[scanned_log["QR Data"]==qr_data][0]
                scanned_log.at[idx, "Scan Count"] += 1
                prev = scanned_log.at[idx, "Timestamps"]
                scanned_log.at[idx, "Timestamps"] = f"{prev};{now}" if prev else now
            else:
                scanned_log.loc[len(scanned_log)] = [qr_data, 1, now]

            sort_log(scanned_log)
            scanned_log.to_csv(LOG_FILE, index=False)

            # 2) Update master Excel statuses & collect details
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            hdrs = {c.value.strip().lower(): i+1 for i,c in enumerate(ws[1]) if c.value}
            scan_key = SCAN_COL_NAME.lower()
            if scan_key not in hdrs:
                col = ws.max_column+1
                ws.cell(row=1, column=col, value=SCAN_COL_NAME)
                hdrs[scan_key] = col

            details = {}
            for row in ws.iter_rows(min_row=2):
                name  = str(row[hdrs["name"]-1].value or "").strip()
                email = str(row[hdrs["email address"]-1].value or "").strip()
                reg   = str(row[hdrs["registration number"]-1].value or "").strip()
                qrval = str(row[hdrs["qr"]-1].value or "").strip()

                if qr_data == qrval or (name and email and reg and name in qr_data and email in qr_data and reg in qr_data):
                    r = row[0].row
                    curr = ws.cell(row=r, column=hdrs[scan_key]).value or ""
                    m = re.search(r'(\d+)', str(curr))
                    cnt = int(m.group(1)) if m else 0
                    cnt += 1
                    new_status = "Scanned" if cnt == 1 else f"Scanned {cnt} Times"
                    ws.cell(row=r, column=hdrs[scan_key], value=new_status)
                    details = {
                        "Name": name,
                        "Email": email,
                        "Registration Number": reg,
                        "Status": new_status
                    }
                    break
            wb.save(EXCEL_FILE)

            # 3) Rebuild highlighted file & highlight **all** scanned rows
            shutil.copy(EXCEL_FILE, HIGHLIGHTED_FILE)
            wb2 = load_workbook(HIGHLIGHTED_FILE)
            ws2 = wb2.active
            hdrs2 = {c.value.strip().lower(): i+1 for i,c in enumerate(ws2[1]) if c.value}
            yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            scan_col = hdrs2.get(scan_key)
            if scan_col:
                for row in ws2.iter_rows(min_row=2):
                    status = row[scan_col-1].value
                    if status and str(status).strip():
                        for col in range(1, ws2.max_column+1):
                            ws2.cell(row=row[0].row, column=col).fill = yellow
            wb2.save(HIGHLIGHTED_FILE)

        # 4) Broadcast to front end
        last_ts = scanned_log.loc[scanned_log["QR Data"]==qr_data, "Timestamps"].iloc[0].split(";")[-1]
        socketio.emit('row_updated', {
            "qr_data":      qr_data,
            "scan_count":   int(scanned_log.loc[scanned_log["QR Data"]==qr_data, "Scan Count"]),
            "timestamps":   scanned_log.loc[scanned_log["QR Data"]==qr_data, "Timestamps"].iloc[0],
            "last_timestamp": last_ts
        })

        return jsonify(message=f"Scanned: {details.get('Name','')}", details=details)

    except Exception as e:
        print("Scan error:", e)
        return jsonify(message="Internal server error", error=str(e)), 500

if __name__ == '__main__':
    socketio.run(app, debug=True, port=5001)
