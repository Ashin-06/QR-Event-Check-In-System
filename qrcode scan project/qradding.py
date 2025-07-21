import pandas as pd
import qrcode
import os
import random
import string
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# === STEP 1: Read CSV ===
csv_path =       #CSV Path
df = pd.read_csv(csv_path)

# Clean headers (remove leading/trailing spaces)
df.columns = df.columns.str.strip()

# Column names
email_col = 'Email Address'
name_col = 'Name'
reg_col = 'Registration Number'  # ← No space at end now
unique_id_col = 'Unique ID'

# === STEP 2: Generate Unique IDs ===
def generate_unique_id(length=8):
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=length))

df[unique_id_col] = [generate_unique_id() for _ in range(len(df))]

# === STEP 3: Save to Excel (initial) ===
excel_path = r"data_temp.xlsx"
df.to_excel(excel_path, index=False)

# === STEP 4: Create QR code folder ===
qr_folder = "qrcodes_samples"
os.makedirs(qr_folder, exist_ok=True)

# === STEP 5: Generate QR codes and save images ===
for _, row in df.iterrows():
    name = str(row[name_col]).strip()
    email = str(row[email_col]).strip()
    reg = str(row[reg_col]).strip()
    unique_id = str(row[unique_id_col]).strip()

    qr_data = f"Name: {name}\nEmail: {email}\nReg No: {reg}\nID: {unique_id}"
    qr_img = qrcode.make(qr_data)

    qr_path = os.path.join(qr_folder, f"{reg}.png")
    qr_img.save(qr_path)

# === STEP 6: Load Excel and Embed QR Images ===
wb = load_workbook(excel_path)
ws = wb.active

# Add 'QR' header
qr_col_index = ws.max_column + 1
ws.cell(row=1, column=qr_col_index).value = "QR Code"

# Embed each QR image into Excel
for idx, row in df.iterrows():
    reg = str(row[reg_col]).strip()
    img_path = os.path.join(qr_folder, f"{reg}.png")

    if os.path.exists(img_path):
        img = XLImage(img_path)
        img.width = 100
        img.height = 100
        col_letter = chr(64 + qr_col_index if qr_col_index <= 26 else 65 + (qr_col_index - 27) // 26) + chr(65 + (qr_col_index - 1) % 26) if qr_col_index > 26 else chr(64 + qr_col_index)
        cell_position = f"{col_letter}{idx + 2}"
        ws.add_image(img, cell_position)

# === STEP 7: Save Final Excel File ===
final_excel =            #Final CSV Path
wb.save(final_excel)

print(f"✅ All QR codes added to '{final_excel}' with unique IDs.")
