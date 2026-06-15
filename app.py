"""
QR Event Check-In System — app.py
Fixed & upgraded production version.

Bug fixes applied:
1. Race condition: global mutable DataFrame protected by lock everywhere (original missed /data_csv read).
2. scanned_log.loc[len(scanned_log)] index bug — replaced with pd.concat() to avoid index collisions.
3. QR match logic was broken: original used "name in qr_data" (substring match) which gives false positives.
   Now stores and matches the encoded QR string exactly.
4. Column-letter formula in qradding.py was wrong for columns > 26 — fixed in utility.
5. No error handling if EXCEL_FILE missing at startup — now raises clear message.
6. sort_log() mutated the shared df in-place without lock — moved inside lock.
7. processing = False in socket callback only — if socket event never arrives (network drop) scanner locks.
   Added timeout fallback in frontend.
8. HIGHLIGHTED_FILE rebuild on every scan is slow (full copy+rewrite) — moved to background thread.
9. No CORS restriction / secret key on Flask app — added.
10. Timestamps column dtype inconsistency (NaN vs str) — normalised on load.
11. /data_csv leaked full Timestamps column to frontend unnecessarily — still included but sanitised.
12. Missing Content-Type validation on /scan — added.
13. No rate-limiting guard — basic per-IP cooldown added.
14. qrsendemail.py: REG_COL had trailing space ' ' — fixed.
15. qradding.py: csv_path and final_excel were empty — now use argparse / env vars.
"""

from __future__ import annotations

import eventlet
eventlet.monkey_patch()

import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

import os
import re
import shutil
import socket as _socket
import threading
import time
from collections import defaultdict
from datetime import datetime
import csv
import json
import queue
import io
import random
import string

import pandas as pd
import qrcode
from flask import Flask, jsonify, render_template, request, session, send_file, send_from_directory, redirect
from flask_socketio import SocketIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── User Agent Parser ────────────────────────────────────────────────────────
def _parse_ua(ua: str) -> str:
    ua = ua.lower()
    os_name = "Device"
    if "windows" in ua:
        os_name = "Windows"
    elif "android" in ua:
        os_name = "Android"
    elif "ipad" in ua or "ipod" in ua:
        os_name = "iPad"
    elif "iphone" in ua:
        os_name = "iPhone"
    elif "macintosh" in ua or "mac os" in ua:
        os_name = "macOS"
    elif "linux" in ua:
        os_name = "Linux"
        
    browser = "Browser"
    if "chrome" in ua or "crios" in ua:
        browser = "Chrome"
    elif "firefox" in ua or "fxios" in ua:
        browser = "Firefox"
    elif "safari" in ua and "chrome" not in ua:
        browser = "Safari"
    elif "edge" in ua or "edg" in ua:
        browser = "Edge"
    elif "opera" in ua or "opr" in ua:
        browser = "Opera"
        
    return f"{os_name} {browser}"

# ── LAN IP helper ────────────────────────────────────────────────────────────
def get_lan_ip() -> str:
    """Return the LAN IPv4 address of this machine (best-effort)."""
    try:
        s = _socket.socket(_socket.AF_INET, _socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def is_localhost_request() -> bool:
    if request.headers.get("X-Forwarded-For"):
        return False
    host = request.host.lower()
    if not ("localhost" in host or "127.0.0.1" in host):
        return False
    if request.remote_addr not in ["127.0.0.1", "::1"]:
        return False
    return True


def is_dashboard_authorized() -> bool:
    global dashboard_sharing, dashboard_passcode
    try:
        cfg = get_event_config()
        enforce_local = cfg.get("enforce_localhost_auth", False)
    except Exception:
        enforce_local = False
        
    if is_localhost_request() and not enforce_local:
        return True
    if dashboard_sharing == "public":
        return True
    elif dashboard_sharing == "disabled":
        return False
    
    if session.get("dashboard_authorized") == True:
        return True
        
    passcode_param = request.args.get("passcode")
    if passcode_param == dashboard_passcode:
        session["dashboard_authorized"] = True
        return True
        
    return False


# ── App setup ────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", os.urandom(24).hex())
socketio = SocketIO(app, cors_allowed_origins=os.environ.get("CORS_ORIGIN", "*"))

@app.after_request
def add_header(r):
    r.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    r.headers["Pragma"] = "no-cache"
    r.headers["Expires"] = "0"
    return r

# ── Connected-client tracker ──────────────────────────────────────────────────
_clients_lock = threading.Lock()
_connected_clients: int = 0
_sockets_lock = threading.Lock()
connected_sockets: dict[str, dict] = {}


# ── Connected-device tracker ──────────────────────────────────────────────────
_devices_lock = threading.Lock()
connected_devices: dict[str, dict] = {}

# ── Html2Image singleton (Chrome reuse — avoids cold-start on every render) ──
_hti_instance = None
_hti_lock = threading.Lock()

def _get_hti(output_path: str = None):
    """Return a shared Html2Image instance, creating it once and reusing it."""
    global _hti_instance
    with _hti_lock:
        if _hti_instance is None:
            try:
                from html2image import Html2Image
                profile_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chrome_profile")
                os.makedirs(profile_dir, exist_ok=True)
                _hti_instance = Html2Image(
                    output_path=output_path or os.path.join(BASE_DIR, "events"),
                    custom_flags=["--no-sandbox", "--disable-dev-shm-usage",
                                  "--disable-gpu", "--headless=new",
                                  "--hide-scrollbars", "--disable-extensions",
                                  f"--user-data-dir={profile_dir}"],
                    disable_logging=True,
                )
                print("[*] Html2Image Chrome instance initialized with persistent profile (will be reused)")
            except Exception as e:
                print(f"[!] Html2Image init failed: {e}")
                _hti_instance = None
        if _hti_instance is not None and output_path:
            _hti_instance.output_path = output_path
        return _hti_instance


# ── File paths & Events Structure ─────────────────────────────────────────────
BASE_DIR         = os.path.dirname(os.path.abspath(__file__))
EVENTS_DIR       = os.path.join(BASE_DIR, "events")
os.makedirs(EVENTS_DIR, exist_ok=True)
SCAN_COL_NAME    = "Scanned Status"

active_event = "Default Event"

def get_active_event_path() -> str:
    global active_event
    p = os.path.join(EVENTS_DIR, active_event)
    os.makedirs(p, exist_ok=True)
    return p

def get_excel_file() -> str:
    return os.path.join(get_active_event_path(), "registrations.xlsx")

def get_highlighted_file() -> str:
    return os.path.join(get_active_event_path(), "registrations_highlighted.xlsx")

def get_log_file() -> str:
    return os.path.join(get_active_event_path(), "scanned_log.csv")

def get_qr_dir() -> str:
    p = os.path.join(get_active_event_path(), "qrcodes")
    os.makedirs(p, exist_ok=True)
    return p

def get_barcode_dir() -> str:
    p = os.path.join(get_active_event_path(), "barcodes")
    os.makedirs(p, exist_ok=True)
    return p

def get_config_file() -> str:
    return os.path.join(get_active_event_path(), "config.json")


# ── Global state ──────────────────────────────────────────────────────────────
lock            = threading.RLock()
highlight_lock  = threading.Lock()

dashboard_sharing = "passcode"
# Generate a random 4-digit passcode on server startup
dashboard_passcode = "".join(random.choices(string.digits, k=4))

# ── Rate limiting (simple in-memory) ─────────────────────────────────────────
_rate_store: dict[str, float] = defaultdict(float)
_processed_idempotency_keys: dict[str, tuple[dict, int]] = {}
RATE_LIMIT_SECONDS = 2          # minimum seconds between scans from same IP


def _is_rate_limited(ip: str, device_id: str = "unknown") -> bool:
    key = f"{ip}:{device_id}"
    now = time.time()
    if now - _rate_store[key] < RATE_LIMIT_SECONDS:
        return True
    _rate_store[key] = now
    return False


# ── Load / initialise log ─────────────────────────────────────────────────────
# ── Event Config Loader / Saver ───────────────────────────────────────────────
def get_event_config() -> dict:
    config_file = get_config_file()
    defaults = {
        "email_template": "Hi {Name},\n\nPlease find your QR Code attached. Present it at the entrance of {Event}.\n\nRegistration: {Registration Number}\n\nRegards,\nEvent Team",
        "whatsapp_template": "Hi {Name},\n\nYour registration QR Code for {Event} is ready! Present it at the entrance.\n\nRegistration: {Registration Number}\n\nDownload QR here: {QR_URL}",
        "auto_notify_on_scan": False,
        "scan_notify_template": "Hi {Name},\n\nYou checked in successfully at {Location} at {Time}.\n\nRegards,\nEvent Team",
        "scan_notify_channels": "none", # "email", "whatsapp", "both", "none"
        "email_subject": "Your Event QR Code",
        "email_sender": "",
        "email_password": "",
        "twilio_sid": "",
        "twilio_token": "",
        "twilio_sender": "",
        "event_name_template": "the Event",
        "checkin_start_date": "",
        "checkin_end_date": "",
        "checkin_start_time": "",
        "checkin_end_time": "",
        "id_card_theme": "auto",
        "id_card_header": "",
        "id_card_footer": "",
        "id_card_show_reg": True,
        "id_card_show_email": True,
        "id_card_show_phone": True,
        "id_card_show_uid": True,
        "id_card_show_pass": True,
        "id_card_label_reg": "REGISTRATION:",
        "id_card_label_email": "EMAIL:",
        "id_card_label_phone": "PHONE:",
        "id_card_label_uid": "UNIQUE ID:",
        "id_card_label_pass": "PASS TYPE:",
        "group_column": "",
        "subgroup_templates": {},
        "sms_template": "Hi {Name},\n\nYour registration QR Code for {Event} is ready! Present it at the entrance.\n\nRegistration: {Registration Number}\n\nDownload QR here: {QR_URL}",
        "sms_provider": "android",
        "android_gateway_url": "",
        "twilio_sms_sid": "",
        "twilio_sms_token": "",
        "twilio_sms_sender": "",
        "meta_access_token": "",
        "meta_phone_number_id": "",
        "meta_template_name": "",
        "meta_lang_code": "en_US",
        "wa_provider": "manual",
        "reg_notify_email": False,
        "reg_notify_whatsapp": False,
        "reg_notify_sms": False,
        "scan_notify_email": False,
        "scan_notify_whatsapp": False,
        "scan_notify_sms": False,
        "enforce_localhost_auth": False,
        "cryptographic_qr_verification": False,
        "event_signing_key": "",
        "allowed_country_codes": "91"
    }
    cfg = defaults
    if os.path.exists(config_file):
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                loaded_cfg = json.load(f)
                for k, v in defaults.items():
                    if k not in loaded_cfg:
                        loaded_cfg[k] = v
                cfg = loaded_cfg
        except Exception:
            pass
            
    # Auto-generate signing key if empty
    if not cfg.get("event_signing_key"):
        import secrets
        cfg["event_signing_key"] = secrets.token_hex(16)
        save_event_config(cfg)
        
    return cfg

def save_event_config(cfg: dict):
    config_file = get_config_file()
    try:
        with open(config_file, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=4)
    except Exception as e:
        print(f"[config] Error saving config: {e}")


def _log_audit(action: str, details: str, device_id: str = "System"):
    """Appends a tamper-evident audit record to audit_log.csv in the active event directory.
    Each entry contains a hash that chains it to the previous entry.
    """
    import hashlib
    import csv
    audit_file = os.path.join(get_active_event_path(), "audit_log.csv")
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ip = "localhost"
    ua = "System"
    try:
        if request:
            ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip()
            ua = request.headers.get("User-Agent", "Unknown")
    except RuntimeError:
        pass
        
    prev_hash = "0" * 64
    file_exists = os.path.exists(audit_file)
    if file_exists:
        try:
            with open(audit_file, "r", newline="", encoding="utf-8") as f:
                reader = list(csv.reader(f))
                if len(reader) > 1:
                    prev_hash = reader[-1][-1]
        except Exception:
            pass
            
    hash_src = f"{prev_hash}|{timestamp}|{action}|{details}|{ip}|{device_id}|{ua}"
    curr_hash = hashlib.sha256(hash_src.encode("utf-8")).hexdigest()
    
    try:
        with open(audit_file, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not file_exists or os.path.getsize(audit_file) == 0:
                writer.writerow(["Timestamp", "Action", "Details", "IP", "Device ID", "User Agent", "Hash"])
            writer.writerow([timestamp, action, details, ip, device_id, ua, curr_hash])
    except Exception as ex:
        print(f"[audit] Error writing audit log: {ex}")


def _verify_audit_log_integrity(file_path: str) -> tuple[bool, int]:
    """Verifies that the hashes in the audit log form a valid chain.
    Returns (is_valid, failing_row_index).
    """
    import hashlib
    import csv
    if not os.path.exists(file_path):
        return True, 0
        
    try:
        with open(file_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if not header:
                return True, 0
                
            prev_hash = "0" * 64
            for idx, row in enumerate(reader, 2):
                if len(row) < 7:
                    return False, idx
                timestamp, action, details, ip, device_id, ua, curr_hash = row
                hash_src = f"{prev_hash}|{timestamp}|{action}|{details}|{ip}|{device_id}|{ua}"
                expected_hash = hashlib.sha256(hash_src.encode("utf-8")).hexdigest()
                if curr_hash != expected_hash:
                    return False, idx
                prev_hash = curr_hash
        return True, 0
    except Exception:
        return False, -1


def _quarantine_scan(qr_data: str, device_name: str, reason: str):
    """Adds a failed/invalid scan to the quarantine queue (quarantine.json)."""
    q_file = os.path.join(get_active_event_path(), "quarantine.json")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    q_scans = []
    if os.path.exists(q_file):
        try:
            with open(q_file, "r", encoding="utf-8") as f:
                q_scans = json.load(f)
        except Exception:
            pass
            
    q_id = f"Q{random.randint(100000, 999999)}"
    new_scan = {
        "id": q_id,
        "qr_data": qr_data,
        "device_name": device_name,
        "reason": reason,
        "timestamp": timestamp,
        "status": "quarantined"
    }
    q_scans.append(new_scan)
    
    try:
        with open(q_file, "w", encoding="utf-8") as f:
            json.dump(q_scans, f, indent=4)
        socketio.emit("quarantine_updated", {})
    except Exception as ex:
        print(f"[quarantine] Error saving quarantine list: {ex}")

# ── Asynchronous Registration Notifications Queue & Worker ─────────────────────
registration_notifications_queue = queue.Queue()

def _registration_notifications_worker():
    while True:
        try:
            reg_no, cfg, host_url = registration_notifications_queue.get()
            
            # Send Email if configured
            if cfg.get("reg_notify_email") == True:
                try:
                    _send_single_email_helper(reg_no, cfg)
                except Exception as ex:
                    print(f"[worker] Error auto-sending registration email for {reg_no}: {ex}")
                time.sleep(0.5)
                
            # Send WhatsApp if configured
            if cfg.get("reg_notify_whatsapp") == True:
                try:
                    _send_single_whatsapp_helper(reg_no, cfg, host_url)
                except Exception as ex:
                    print(f"[worker] Error auto-sending registration WhatsApp for {reg_no}: {ex}")
                time.sleep(0.5)
                
            # Send SMS if configured
            if cfg.get("reg_notify_sms") == True:
                try:
                    _send_single_sms_helper(reg_no, cfg, host_url)
                except Exception as ex:
                    print(f"[worker] Error auto-sending registration SMS for {reg_no}: {ex}")
                time.sleep(0.5)
                
        except Exception as e:
            print(f"[worker] Error in automatic registration notification worker: {e}")
        finally:
            registration_notifications_queue.task_done()

# Start background worker thread
threading.Thread(target=_registration_notifications_worker, daemon=True).start()

# ── Load / initialise log ─────────────────────────────────────────────────────
def _load_log() -> pd.DataFrame:
    log_file = get_log_file()
    if os.path.exists(log_file):
        df = pd.read_csv(log_file)
    else:
        df = pd.DataFrame(columns=["QR Data", "Scan Count", "Timestamps", "Devices"])

    # Normalise columns
    for col in ["QR Data", "Timestamps", "Devices"]:
        if col not in df.columns:
            df[col] = ""
    if "Scan Count" not in df.columns:
        df["Scan Count"] = 0

    # Ensure dtypes
    df["Scan Count"] = pd.to_numeric(df["Scan Count"], errors="coerce").fillna(0).astype(int)
    df["Timestamps"] = df["Timestamps"].fillna("").astype(str)
    df["Devices"]    = df["Devices"].fillna("").astype(str)
    df["QR Data"]    = df["QR Data"].fillna("").astype(str)
    return df


def sort_log(df: pd.DataFrame) -> None:
    """Sort by most-recent timestamp descending, in-place."""
    df["__last"] = df["Timestamps"].apply(
        lambda s: s.split(";")[-1].strip() if s else ""
    )
    df.sort_values("__last", ascending=False, inplace=True)
    df.drop(columns="__last", inplace=True)
    df.reset_index(drop=True, inplace=True)


scanned_log = None

def load_active_event() -> None:
    global scanned_log
    excel_path = get_excel_file()
    log_path = get_log_file()
    
    # Clean up any leftover temporary files in the event folder
    try:
        event_path = get_active_event_path()
        if os.path.exists(event_path):
            for f in os.listdir(event_path):
                if f.startswith("tmp") and f.endswith(".xlsx"):
                    try:
                        os.unlink(os.path.join(event_path, f))
                    except OSError:
                        pass
    except Exception as e:
        print(f"[load_active_event] Error cleaning leftover temp files: {e}")

    # Look for registrations.csv first and convert to registrations.xlsx (always take precedence if CSV exists)
    csv_path = os.path.join(get_active_event_path(), "registrations.csv")
    if os.path.exists(csv_path):
        try:
            print(f"[load_active_event] Found registrations.csv. Converting to registrations.xlsx...")
            df = pd.read_csv(csv_path)
            df.to_excel(excel_path, index=False, engine="openpyxl")
            try:
                os.remove(csv_path)
            except OSError:
                pass
        except Exception as e:
            print(f"[load_active_event] Error converting registrations.csv to Excel: {e}")

    # Ensure registrations file exists
    if not os.path.exists(excel_path):
        # Create a default blank registrations sheet
        master_src = os.path.join(BASE_DIR, "data_with_qrdemo.xlsx")
        if os.path.exists(master_src):
            shutil.copy(master_src, excel_path)
        else:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = [
                "Name", "Email Address", "Registration Number", "Phone Number",
                "Unique ID", "QR", "Barcode", SCAN_COL_NAME, "Email Sent Status", "WhatsApp Sent Status",
                "Scan Timestamps", "Scan Devices", "QR Code Image", "Barcode Image"
            ]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            wb.save(excel_path)
            wb.close()
            
    # Auto-initialize and self-heal registrations metadata if it exists
    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            if _initialize_missing_metadata(ws):
                _atomic_save(wb, excel_path)
            else:
                wb.close()
        except Exception as e:
            print(f"[load_active_event] Error self-healing Excel metadata: {e}")

    # Ensure scanned log exists
    if not os.path.exists(log_path):
        df = pd.DataFrame(columns=["QR Data", "Scan Count", "Timestamps", "Devices"])
        df.to_csv(log_path, index=False)
        
    scanned_log = _load_log()
    sort_log(scanned_log)
    
    # Highlighted file
    high_path = get_highlighted_file()
    if not os.path.exists(high_path):
        shutil.copy(excel_path, high_path)
        
    # Rebuild highlighted in background
    threading.Thread(target=_rebuild_highlighted, daemon=True).start()



# ── Background highlight rebuild ───────────────────────────────────────────────
def _rebuild_highlighted() -> None:
    """Copy master Excel and re-highlight all scanned rows (runs in daemon thread).
    Uses a staging temp-copy so the source file is never locked during processing.
    """
    import tempfile
    wb = None
    staging = None
    excel_file = get_excel_file()
    highlighted_file = get_highlighted_file()
    try:
        with highlight_lock:
            # Copy source to a temp file in system temp (very brief lock)
            with lock:
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    staging = tmp.name
                shutil.copy(excel_file, staging)

            # Process the staging copy (no lock held on excel_file)
            wb = load_workbook(staging)
            ws = wb.active
            hdrs = {
                str(c.value).strip().lower(): i + 1
                for i, c in enumerate(ws[1])
                if c.value
            }
            scan_key = SCAN_COL_NAME.lower()
            scan_col = hdrs.get(scan_key)
            yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            if scan_col:
                for row in ws.iter_rows(min_row=2):
                    status = row[scan_col - 1].value
                    if status and str(status).strip():
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row[0].row, column=col).fill = yellow

            # Atomic save → highlighted_file
            _atomic_save(wb, highlighted_file)
            wb = None

    except Exception as e:
        print(f"[highlight] Error: {e}")
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        if staging and os.path.exists(staging):
            try:
                os.unlink(staging)
            except OSError:
                pass


def _clean_val(val, default="") -> str:
    if pd.isna(val) or val is None:
        return default
    # Excel reads integer-like numbers as float (e.g. phone 919895955673 → 919895955673.0).
    # Convert whole-number floats to int first so the .0 suffix never corrupts the string.
    if isinstance(val, float) and val == int(val):
        val = int(val)
    s = str(val).strip()
    if s.lower() == "nan":
        return default
    return s


# ── Smart phone number normaliser ──────────────────────────────────────────
# Converts any phone string to E.164 (+XXXXXXXXXX) format.
# Auto-detects country code from length/prefix so callers never need to
# hard-code +91 themselves.
def _normalize_phone(phone, default_cc: str = None) -> str:
    """Return phone in E.164 format (+CC + subscriber number).
    If it represents an impossible number (all zeros, or less than 7 digits, or more than 15 digits),
    returns an empty string.
    Checks config for allowed_country_codes (comma separated list of country codes).
    """
    if phone is None or pd.isna(phone):
        return ""

    try:
        import numpy as np
        is_num = isinstance(phone, (int, float, np.integer, np.floating))
    except ImportError:
        is_num = isinstance(phone, (int, float))

    if is_num:
        try:
            if phone == int(phone):
                phone = int(phone)
        except Exception:
            pass

    phone_str = str(phone).strip()
    if phone_str.endswith(".0"):
        phone_str = phone_str[:-2]

    if not phone_str:
        return ""

    # Strip formatting chars that are never part of a number
    raw = re.sub(r"[\s\-\.\(\)/]", "", phone_str.strip())
    
    # Check if all zeros
    digits_only = re.sub(r"\D", "", raw)
    if not digits_only or all(d == '0' for d in digits_only):
        return ""

    # Already has a '+' prefix
    if raw.startswith("+"):
        clean_digits = re.sub(r"\D", "", raw[1:])
        if len(clean_digits) < 7 or len(clean_digits) > 15:
            return ""
        return "+" + clean_digits

    # Strip leading zeros
    digits = digits_only.lstrip("0") or digits_only
    n = len(digits)

    if n < 7 or n > 15:
        return ""

    # Load country codes from config if default_cc not provided or is default
    if not default_cc or default_cc == "91":
        try:
            cfg = get_event_config()
            cc_str = cfg.get("allowed_country_codes", "91")
            cc_list = [c.strip() for c in cc_str.split(",") if c.strip()]
            if cc_list:
                default_cc = cc_list[0]
            else:
                default_cc = "91"
        except Exception:
            default_cc = "91"
            cc_list = ["91"]
    else:
        cc_list = [default_cc]

    # --- Check if starts with one of the configured country codes ---
    # Sort country codes by length descending to match longest first
    sorted_cc_list = sorted(cc_list, key=len, reverse=True)
    matched_cc = None
    for cc in sorted_cc_list:
        if digits.startswith(cc) and len(digits) > len(cc) + 4: # CC prefix plus at least 5 digits subscriber
            matched_cc = cc
            break

    if matched_cc:
        return "+" + digits

    # --- 3-digit country codes fallback ---
    _3CC = {
        "971": 12, "972": 12, "966": 12, "974": 11, "973": 11,
        "968": 11, "962": 12, "961": 11, "964": 13, "965": 11,
        "880": 13, "855": 11, "856": 11, "853": 11, "852": 11,
        "354": 11, "353": 11, "358": 11, "370": 11, "371": 11,
        "372": 11, "373": 11, "375": 12, "376": 11, "377": 11,
    }
    # --- 2-digit country codes fallback ---
    _2CC = {
        "91": 12, "44": 12, "61": 11, "49": 12, "33": 11, "34": 11,
        "39": 12, "81": 11, "82": 12, "86": 13, "55": 13, "52": 12,
        "27": 11, "20": 12, "60": 11, "62": 13, "63": 12, "65": 10,
        "66": 11, "92": 12, "94": 11, "95": 11, "98": 12, "90": 12,
        "48": 11, "31": 11, "32": 11, "41": 11, "43": 11, "46": 11,
        "47": 10, "45": 10, "30": 12, "36": 11, "40": 11, "38": 12,
    }

    if n >= 10:
        p3 = digits[:3]
        if p3 in _3CC and n == _3CC[p3]:
            return "+" + digits

        p2 = digits[:2]
        if p2 in _2CC and n == _2CC[p2]:
            return "+" + digits

        # US / Canada: 1XXXXXXXXXX (11 digits starting with 1)
        if digits[0] == "1" and n == 11:
            return "+" + digits

        # Heuristic: if >= 12 digits it almost certainly has a CC already
        if n >= 12:
            return "+" + digits

    # 10 digits: assume default country code
    if n == 10:
        return f"+{default_cc}{digits}"

    # Fallback: prepend + and hope for the best
    return "+" + digits


def _validate_attendee_integrity(row_dict: dict, existing_regs: set[str] = None) -> dict[str, str]:
    """Validate a single attendee row's data.
    Returns a dict mapping field names to error descriptions. Empty dict means valid.
    """
    errors = {}
    name = row_dict.get("name", "").strip()
    email = row_dict.get("email", "").strip()
    reg_no = row_dict.get("reg_no", "").strip()
    phone = row_dict.get("phone", "").strip()
    
    if not name:
        errors["name"] = "Name cannot be empty."
    
    if not reg_no:
        errors["reg_no"] = "Registration Number cannot be empty."
    elif existing_regs and reg_no.lower() in existing_regs:
        errors["reg_no"] = f"Registration Number '{reg_no}' is already taken."
        
    if email:
        if "@" not in email:
            errors["email"] = "Email must contain '@' symbol."
            
    if phone:
        norm = _normalize_phone(phone)
        if not norm:
            errors["phone"] = "Phone number is invalid."
        else:
            digits = re.sub(r"\D", "", norm)
            if all(d == "0" for d in digits):
                errors["phone"] = "Phone number cannot be all zeros."
            else:
                cc_found = None
                sub_len = len(digits)
                for cc in ["971", "972", "966", "974", "973", "968", "962", "961", "964", "965", "880", "855", "856", "853", "852", "354", "353", "358", "370", "371", "372", "373", "375", "376", "377"]:
                    if digits.startswith(cc):
                        cc_found = cc
                        sub_len = len(digits) - len(cc)
                        break
                if not cc_found:
                    for cc in ["91", "44", "61", "49", "33", "34", "39", "81", "82", "86", "55", "52", "27", "20", "60", "62", "63", "65", "66", "92", "94", "95", "98", "90", "48", "31", "32", "41", "43", "46", "47", "45", "30", "36", "40", "38"]:
                        if digits.startswith(cc):
                            cc_found = cc
                            sub_len = len(digits) - len(cc)
                            break
                if not cc_found and digits.startswith("1"):
                    cc_found = "1"
                    sub_len = len(digits) - 1
                    
                if cc_found and sub_len < 10:
                    errors["phone"] = "Phone number must have at least 10 digits after the country code."
                
    return errors


# ── Helper: normalise column header lookup ─────────────────────────────────────
def _get_headers(ws) -> dict[str, int]:
    return {
        str(c.value).strip().lower(): i + 1
        for i, c in enumerate(ws[1])
        if c.value
    }


def _broadcast_devices():
    with _devices_lock:
        devs = list(connected_devices.values())
    socketio.emit("devices_updated", devs)


def _emit_clients_count():
    with _sockets_lock:
        scanners_count = sum(1 for s in connected_sockets.values() if s.get("type") == "scanner" and not s.get("is_local"))
        local_scanners_count = sum(1 for s in connected_sockets.values() if s.get("type") == "scanner" and s.get("is_local"))
        dashboards_count = sum(1 for s in connected_sockets.values() if s.get("type") == "dashboard")
        external_dashboards_count = sum(1 for s in connected_sockets.values() if s.get("type") == "dashboard" and not s.get("is_local"))
        
    socketio.emit("clients_count", {
        "scanners_count": scanners_count,
        "local_scanners_count": local_scanners_count,
        "dashboards_count": dashboards_count,
        "external_dashboards_count": external_dashboards_count
    })


# ── Socket.IO connect / disconnect tracking ───────────────────────────────────
@socketio.on("connect")
def _on_connect():
    global _connected_clients
    with _clients_lock:
        _connected_clients += 1
        count = _connected_clients
        
    with _sockets_lock:
        ip = request.headers.get("X-Forwarded-For", request.remote_addr or "127.0.0.1").split(",")[0].strip()
        is_local = ip in ["127.0.0.1", "::1", "localhost"]
        connected_sockets[request.sid] = {
            "type": "unknown",
            "ip": ip,
            "is_local": is_local
        }
        
    _emit_clients_count()
    
    with _devices_lock:
        devs = list(connected_devices.values())
    socketio.emit("devices_updated", devs, to=request.sid)


@socketio.on("disconnect")
def _on_disconnect():
    global _connected_clients
    with _clients_lock:
        _connected_clients = max(0, _connected_clients - 1)
        count = _connected_clients
        
    with _sockets_lock:
        if request.sid in connected_sockets:
            del connected_sockets[request.sid]
            
    _emit_clients_count()
    
    device_id = session.get("device_id")
    if device_id:
        with _devices_lock:
            if device_id in connected_devices:
                connected_devices[device_id]["online"] = False
                connected_devices[device_id]["last_active_time"] = datetime.now().strftime("%H:%M:%S")
        _broadcast_devices()


@socketio.on("register_device")
def _on_register_device(data):
    device_id = data.get("device_id")
    if not device_id:
        return
    
    device_name = data.get("device_name", "Unknown Device").strip()
    user_agent = request.headers.get("User-Agent", "Unknown")
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip()
    short_ua = _parse_ua(user_agent)
    
    with _sockets_lock:
        if request.sid in connected_sockets:
            connected_sockets[request.sid]["type"] = "scanner"
            connected_sockets[request.sid]["device_id"] = device_id
    _emit_clients_count()
    
    with _devices_lock:
        session["device_id"] = device_id
        if device_id not in connected_devices:
            connected_devices[device_id] = {
                "id": device_id,
                "name": device_name,
                "manager_label": "",
                "ip": ip,
                "user_agent": short_ua,
                "online": True,
                "scans": 0,
                "last_activity": "Connected",
                "last_active_time": datetime.now().strftime("%H:%M:%S")
            }
        else:
            connected_devices[device_id]["name"] = device_name
            connected_devices[device_id]["ip"] = ip
            connected_devices[device_id]["user_agent"] = short_ua
            connected_devices[device_id]["online"] = True
            connected_devices[device_id]["last_active_time"] = datetime.now().strftime("%H:%M:%S")
            if "manager_label" not in connected_devices[device_id]:
                connected_devices[device_id]["manager_label"] = ""
            if connected_devices[device_id]["last_activity"] == "Offline":
                connected_devices[device_id]["last_activity"] = "Reconnected"
                
    _broadcast_devices()


@socketio.on("register_dashboard")
def _on_register_dashboard(data):
    with _sockets_lock:
        if request.sid in connected_sockets:
            connected_sockets[request.sid]["type"] = "dashboard"
            connected_sockets[request.sid]["dashboard_id"] = data.get("dashboard_id")
    _emit_clients_count()


@socketio.on("rename_device")
def _on_rename_device(data):
    device_id = data.get("device_id")
    new_name = data.get("name", "").strip()
    if not device_id or not new_name:
        return
    with _devices_lock:
        if device_id in connected_devices:
            connected_devices[device_id]["name"] = new_name
            connected_devices[device_id]["last_active_time"] = datetime.now().strftime("%H:%M:%S")
    _broadcast_devices()


@socketio.on("rename_device_manager")
def _on_rename_device_manager(data):
    device_id = data.get("device_id")
    new_label = data.get("label", "").strip()
    if not device_id:
        return
    with _devices_lock:
        if device_id in connected_devices:
            connected_devices[device_id]["manager_label"] = new_label
            connected_devices[device_id]["last_active_time"] = datetime.now().strftime("%H:%M:%S")
    _broadcast_devices()



# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    is_admin = is_localhost_request() or session.get("dashboard_authorized") == True
    return render_template("index.html", is_admin=is_admin)


@app.route("/dashboard")
def dashboard_view():
    if not is_dashboard_authorized():
        return render_template("dashboard_login.html", error=request.args.get("error"))
    return render_template("dashboard.html")


@app.route("/dashboard_login", methods=["GET", "POST"])
def dashboard_login():
    global dashboard_passcode
    if request.method == "POST":
        code_entered = request.form.get("passcode", "").strip()
        if code_entered == dashboard_passcode:
            session["dashboard_authorized"] = True
            return redirect("/dashboard")
        return render_template("dashboard_login.html", error="Invalid passcode.")
    return render_template("dashboard_login.html")


@app.route("/dashboard_sharing_info", methods=["GET"])
def dashboard_sharing_info():
    port = int(os.environ.get("PORT", 5001))
    lan_ip = get_lan_ip()
    lan_url = f"http://{lan_ip}:{port}/dashboard?passcode={dashboard_passcode}"
    
    tunnel_url_dash = ""
    if _tunnel_url:
        tunnel_url_dash = f"{_tunnel_url}/dashboard?passcode={dashboard_passcode}"
        
    return jsonify(
        sharing=dashboard_sharing,
        passcode=dashboard_passcode,
        lan_url=lan_url,
        tunnel_url=tunnel_url_dash
    )


@app.route("/save_sharing_settings", methods=["POST"])
def save_sharing_settings():
    global dashboard_sharing, dashboard_passcode
    data = request.json or {}
    sharing = data.get("sharing", "passcode")
    passcode = data.get("passcode", "").strip()
    
    if sharing not in ["passcode", "public", "disabled"]:
        return jsonify(message="Invalid sharing mode."), 400
        
    dashboard_sharing = sharing
    if passcode:
        dashboard_passcode = passcode
        
    socketio.emit("sharing_updated", {})
    return jsonify(message="Dashboard sharing settings saved successfully.")


@app.route("/network_info")
def network_info():
    """Return the LAN URL so the UI can display it for other devices."""
    port = int(os.environ.get("PORT", 5001))
    ip   = get_lan_ip()
    return jsonify(url=f"http://{ip}:{port}", ip=ip, port=port)


def _get_qr_to_attendee_map() -> dict[str, dict]:
    excel_file = get_excel_file()
    mapping = {}
    if not os.path.exists(excel_file):
        return mapping
    try:
        with lock:
            wb = load_workbook(excel_file, read_only=True)
            ws = wb.active
            hdrs = _get_headers(ws)
            name_idx = hdrs.get("name")
            reg_idx = hdrs.get("registration number")
            qr_idx = hdrs.get("qr")
            if name_idx and reg_idx and qr_idx:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) >= max(name_idx, reg_idx, qr_idx):
                        qr_val = str(row[qr_idx - 1] or "").strip()
                        if qr_val:
                            mapping[qr_val] = {
                                "name": str(row[name_idx - 1] or "").strip(),
                                "reg_no": str(row[reg_idx - 1] or "").strip()
                            }
            wb.close()
    except Exception as e:
        print(f"Error building QR mapping: {e}")
    return mapping


@app.route("/data_csv")
def data_csv():
    with lock:
        df = scanned_log.copy()

    if "Devices" not in df.columns:
        df["Devices"] = ""

    df["Last Timestamp"] = df["Timestamps"].apply(
        lambda s: s.split(";")[-1].strip() if s else ""
    )
    df["Last Device"] = df["Devices"].apply(
        lambda s: s.split(";")[-1].strip() if s else ""
    )
    
    # Map QR Data to Attendee Details
    qr_map = _get_qr_to_attendee_map()
    
    records = []
    for idx, row in df.iterrows():
        qr_val = str(row["QR Data"]).strip()
        att = qr_map.get(qr_val)
        if not att:
            # Fallback parse for old verbose formats
            name_match = re.search(r"Name:\s*([^\n\r]+)", qr_val)
            reg_match = re.search(r"(Reg No|Reg_No|Registration Number|ID):\s*([^\n\r]+)", qr_val)
            if name_match:
                att = {
                    "name": name_match.group(1).strip(),
                    "reg_no": reg_match.group(2).strip() if reg_match else "Unknown"
                }
            else:
                att = {"name": qr_val, "reg_no": "Unknown"}
                
        records.append({
            "QR Data": qr_val,
            "attendee_name": att["name"],
            "reg_no": att["reg_no"],
            "Scan Count": int(row["Scan Count"]),
            "Last Timestamp": row["Last Timestamp"],
            "Timestamps": row["Timestamps"],
            "Devices": row["Devices"],
            "Last Device": row["Last Device"]
        })
    return jsonify(records)


@app.route("/stats")
def stats():
    """Summary statistics for dashboard cards."""
    with lock:
        total     = len(scanned_log)
        unique    = int((scanned_log["Scan Count"] > 0).sum())
        duplicate = int((scanned_log["Scan Count"] > 1).sum())
    return jsonify(total=total, unique=unique, duplicate=duplicate)


@app.route("/stats/timeline")
def stats_timeline():
    """Hourly check-in timeline stats."""
    with lock:
        all_timestamps = []
        for ts_str in scanned_log["Timestamps"].dropna():
            for part in str(ts_str).split(";"):
                part = part.strip()
                if part:
                    all_timestamps.append(part)
                    
        hourly_counts = defaultdict(int)
        for ts in all_timestamps:
            try:
                dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
                hour_str = f"Hour {dt.strftime('%H:00')}"
                hourly_counts[hour_str] += 1
            except Exception:
                pass
                
        sorted_hours = sorted(hourly_counts.keys())
        timeline = [{"hour": h, "count": hourly_counts[h]} for h in sorted_hours]
        
    return jsonify(timeline=timeline)


def _emit_stats() -> None:
    """Broadcast current stats to all connected clients."""
    with lock:
        total     = len(scanned_log)
        unique    = int((scanned_log["Scan Count"] > 0).sum())
        duplicate = int((scanned_log["Scan Count"] > 1).sum())
    socketio.emit("stats_updated", {"total": total, "unique": unique, "duplicate": duplicate})


def _atomic_save(wb, target_path: str) -> None:
    """Save workbook to a temp file then atomically replace target.
    Avoids Windows file-lock conflicts between concurrent readers and writers.
    Also rotates and keeps the last 3 versions of registrations.xlsx.
    """
    import tempfile
    
    # Backup rotation (only for registrations.xlsx)
    if os.path.exists(target_path) and "registrations.xlsx" in os.path.basename(target_path):
        try:
            bak1 = target_path + ".bak1"
            bak2 = target_path + ".bak2"
            bak3 = target_path + ".bak3"
            
            # Rotate backups: bak2 -> bak3, bak1 -> bak2, current -> bak1
            if os.path.exists(bak2):
                os.replace(bak2, bak3)
            if os.path.exists(bak1):
                os.replace(bak1, bak2)
            if os.path.exists(target_path):
                shutil.copy2(target_path, bak1)
        except Exception as ex:
            print(f"[atomic_save] Backup rotation error: {ex}")
            
    dir_ = os.path.dirname(target_path)
    with tempfile.NamedTemporaryFile(dir=dir_, suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        wb.save(tmp_path)
        wb.close()
        os.replace(tmp_path, target_path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


# ── Scan Event Notification Helpers ───────────────────────────────────────────
def _send_single_email(sender, password, recipient, subject, body):
    import yagmail
    try:
        yag = yagmail.SMTP(user=sender, password=password)
        yag.send(to=recipient, subject=subject, contents=body)
        yag.close()
        print(f"[notify] Email sent to {recipient}")
    except Exception as e:
        print(f"[notify] Failed to send email to {recipient}: {e}")


def _send_single_whatsapp(sid, token, sender_phone, recipient_phone, body):
    from twilio.rest import Client
    import re
    try:
        client = Client(sid, token)
        phone_clean = _normalize_phone(recipient_phone)
        if phone_clean:
            from_number = f"whatsapp:{sender_phone}"
            to_number = f"whatsapp:{phone_clean}"
            client.messages.create(body=body, from_=from_number, to=to_number)
            print(f"[notify] WhatsApp sent to {phone_clean}")
    except Exception as e:
        print(f"[notify] Failed to send WhatsApp to {recipient_phone}: {e}")


def send_scan_notification_async(details, location, timestamp):
    cfg = get_event_config()
    
    # Checkbox configurations
    notify_email = cfg.get("scan_notify_email") == True
    notify_whatsapp = cfg.get("scan_notify_whatsapp") == True
    notify_sms = cfg.get("scan_notify_sms") == True
    
    # Legacy fallback
    channels = cfg.get("scan_notify_channels", "none")
    if not (notify_email or notify_whatsapp or notify_sms) and channels != "none":
        if channels in ["email", "both"]:
            notify_email = True
        if channels in ["whatsapp", "both"]:
            notify_whatsapp = True
            
    if not (notify_email or notify_whatsapp or notify_sms):
        return
        
    name = details.get("Name", "")
    email = details.get("Email", "")
    phone = details.get("Phone", "")
    reg_no = details.get("Registration Number", "")
    
    # Prepare body using template
    template = cfg.get("scan_notify_template", "")
    body = template.replace("{Name}", name).replace("{Location}", location).replace("{Time}", timestamp).replace("{Registration Number}", reg_no)
    
    # Send email if configured
    if notify_email and email:
        sender = cfg.get("email_sender")
        password = cfg.get("email_password")
        subject = "Check-in Confirmation: " + cfg.get("event_name_template", "Event")
        if sender and password:
            threading.Thread(target=_send_single_email, args=(sender, password, email, subject, body), daemon=True).start()
            
    # Send WhatsApp if configured
    if notify_whatsapp and phone:
        sid = cfg.get("twilio_sid")
        token = cfg.get("twilio_token")
        twilio_phone = cfg.get("twilio_sender")
        if sid and token and twilio_phone:
            threading.Thread(target=_send_single_whatsapp, args=(sid, token, twilio_phone, phone, body), daemon=True).start()

    # Send SMS if configured
    if notify_sms and phone:
        provider = cfg.get("sms_provider", "android")
        threading.Thread(target=_send_single_sms, args=(phone, body, provider, cfg), daemon=True).start()


def _log_to_excel_sheet(wb, now_str, name, reg_no, device, status, qr_val):
    try:
        if "Scan Log" not in wb.sheetnames:
            ws_log = wb.create_sheet("Scan Log")
            ws_log.append(["Timestamp", "Name", "Registration Number", "Scanner Device", "Scan Status", "QR Data"])
        else:
            ws_log = wb["Scan Log"]
        ws_log.append([now_str, name, reg_no, device, status, qr_val])
    except Exception as e:
        print(f"[Scan Log] Error appending to sheet: {e}")


def _perform_checkin(qr_data: str, device_id: str = "unknown") -> tuple[dict, int]:
    """Helper to perform check-in operations. Returns (response_dict, status_code)."""
    # Normalize incoming QR data string
    qr_data_norm = qr_data.replace("\r\n", "\n").strip()

    excel_file = get_excel_file()
    log_file = get_log_file()
    wb = None
    
    # Resolve combined device name
    device_name = device_id
    with _devices_lock:
        if device_id in connected_devices:
            d_info = connected_devices[device_id]
            op_name = d_info.get("name", "").strip()
            m_label = d_info.get("manager_label", "").strip()
            if m_label:
                device_name = f"{m_label} ({op_name})" if op_name else m_label
            else:
                device_name = op_name if op_name else device_id

    # Verify event-scoped cryptographic QR payload signature if enabled
    cfg = get_event_config()
    is_valid_qr, verified_reg = _verify_qr_payload(qr_data_norm, cfg)
    if not is_valid_qr and device_id != "Manual Check-In":
        # Log to audit trail
        _log_audit("Invalid QR Signature", f"Scanned QR: {qr_data_norm}", device_id)
        # Quarantine failed scan
        _quarantine_scan(qr_data_norm, device_name, "Invalid Cryptographic Signature")
        
        # Save to unregistered/invalid log in Excel
        try:
            wb = load_workbook(excel_file)
            _log_to_excel_sheet(wb, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Invalid Signature", "Unknown", device_name, "Invalid QR Signature", qr_data)
            _atomic_save(wb, excel_file)
        except Exception as ex:
            print(f"[checkin] Error saving unregistered scan to Excel: {ex}")
            if wb:
                wb.close()
        wb = None
            
        with _devices_lock:
            if device_id in connected_devices:
                connected_devices[device_id]["last_activity"] = "Invalid Signature Scan"
                connected_devices[device_id]["last_active_time"] = datetime.now().strftime("%H:%M:%S")
        _broadcast_devices()
        return {
            "message": "❌ Invalid Cryptographic Signature (Event Mismatch)",
            "error": "invalid_signature",
            "is_duplicate": False
        }, 400

    try:
        with lock:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 1) Open Excel and find matching attendee
            wb   = load_workbook(excel_file)
            ws   = wb.active
            hdrs, _ = _get_or_create_headers(ws)
            scan_key = SCAN_COL_NAME.lower()

            required = {"name", "registration number", "qr"}
            missing  = required - hdrs.keys()
            if missing:
                return {
                    "message": f"Excel is missing columns: {missing}",
                    "error": "schema_error"
                }, 500

            details: dict = {}
            found_row = None
            for row in ws.iter_rows(min_row=2):
                qrval = str(row[hdrs["qr"] - 1].value or "").strip()
                qrval_norm = qrval.replace("\r\n", "\n").strip()
                regval = str(row[hdrs["registration number"] - 1].value or "").strip()
                regval_norm = regval.replace("\r\n", "\n").strip()
                uidval = str(row[hdrs["unique id"] - 1].value or "").strip()
                uidval_norm = uidval.replace("\r\n", "\n").strip()
                
                # Check for match
                if cfg.get("cryptographic_qr_verification", False) and device_id != "Manual Check-In":
                    # strictly match on registration number
                    if verified_reg.lower() == regval_norm.lower():
                        found_row = row
                        break
                else:
                    if qr_data_norm in [qrval_norm, regval_norm, uidval_norm]:
                        found_row = row
                        break

            # If not found in Excel, reject check-in, quarantine, and log it in the Excel sheet
            if found_row is None:
                # Add to quarantine queue
                _quarantine_scan(qr_data_norm, device_name, "Unregistered QR Code")
                _log_audit("Unregistered Scan Quarantined", f"QR Data: {qr_data_norm}", device_id)
                try:
                    _log_to_excel_sheet(wb, now, "Unregistered", "Unknown", device_name, "Unregistered QR Code", qr_data)
                    _atomic_save(wb, excel_file)
                except Exception as ex:
                    print(f"[checkin] Error saving unregistered scan to Excel: {ex}")
                    if wb:
                        wb.close()
                wb = None
                
                with _devices_lock:
                    if device_id in connected_devices:
                        connected_devices[device_id]["last_activity"] = "Unregistered Scan"
                        connected_devices[device_id]["last_active_time"] = datetime.now().strftime("%H:%M:%S")
                _broadcast_devices()
                return {
                    "message": "❌ Unregistered QR Code",
                    "error": "unregistered_attendee",
                    "details": {},
                    "is_duplicate": False
                }, 400

            # 2) Attendee matched! Extract details
            r   = found_row[0].row
            name_val = str(found_row[hdrs["name"] - 1].value or "").strip()
            reg_val  = str(found_row[hdrs["registration number"] - 1].value or "").strip()
            
            # Check date and time restriction
            cfg = get_event_config()
            start_d = cfg.get("checkin_start_date", "").strip()
            end_d = cfg.get("checkin_end_date", "").strip()
            start_t = cfg.get("checkin_start_time", "").strip()
            end_t = cfg.get("checkin_end_time", "").strip()
            
            is_time_restricted = False
            time_error_msg = ""
            now_dt = datetime.now()
            
            # Start restriction check
            if start_d:
                start_dt_str = f"{start_d} {start_t if start_t else '00:00'}"
                try:
                    start_dt = datetime.strptime(start_dt_str, "%Y-%m-%d %H:%M")
                    if now_dt < start_dt:
                        is_time_restricted = True
                        time_error_msg = f"❌ Check-In Not Started (Allowed: after {start_dt_str})"
                except ValueError:
                    pass
            elif start_t:
                now_time_str = now_dt.strftime("%H:%M")
                if now_time_str < start_t:
                    is_time_restricted = True
                    time_error_msg = f"❌ Check-In Not Started (Allowed: after {start_t})"
                    
            # End restriction check
            if not is_time_restricted:
                if end_d:
                    end_dt_str = f"{end_d} {end_t if end_t else '23:59'}"
                    try:
                        end_dt = datetime.strptime(end_dt_str, "%Y-%m-%d %H:%M")
                        if now_dt > end_dt:
                            is_time_restricted = True
                            time_error_msg = f"❌ Check-In Closed (Allowed: before {end_dt_str})"
                    except ValueError:
                        pass
                elif end_t:
                    now_time_str = now_dt.strftime("%H:%M")
                    if now_time_str > end_t:
                        is_time_restricted = True
                        time_error_msg = f"❌ Check-In Closed (Allowed: before {end_t})"
                
            if is_time_restricted:
                # Still log the scan timestamps and devices in Excel main row under special status
                col = hdrs[scan_key]
                ws.cell(row=r, column=col, value="Scanned Outside Time")
                
                ts_col_idx = hdrs.get("scan timestamps")
                dev_col_idx = hdrs.get("scan devices")
                if ts_col_idx:
                    prev_ts = str(ws.cell(row=r, column=ts_col_idx).value or "").strip()
                    ws.cell(row=r, column=ts_col_idx, value=f"{prev_ts};{now}" if prev_ts else now)
                if dev_col_idx:
                    prev_dev = str(ws.cell(row=r, column=dev_col_idx).value or "").strip()
                    ws.cell(row=r, column=dev_col_idx, value=f"{prev_dev};{device_name}" if prev_dev else device_name)
                    
                # Append to Scan Log sheet
                _log_to_excel_sheet(wb, now, name_val, reg_val, device_name, "Before/After Allowed Time", qr_data)
                
                # Quarantine the scan
                _quarantine_scan(qr_data_norm, device_name, f"Outside Allowed Window ({time_error_msg})")
                _log_audit("Scan Quarantined (Outside Window)", f"Attendee: {name_val} ({reg_val})", device_id)
                
                # Atomic save Excel
                _atomic_save(wb, excel_file)
                wb = None
                
                # Send update to device monitor
                with _devices_lock:
                    if device_id in connected_devices:
                        connected_devices[device_id]["last_activity"] = f"Scanned {name_val} (Outside Time)"
                        connected_devices[device_id]["last_active_time"] = datetime.now().strftime("%H:%M:%S")
                _broadcast_devices()
                
                # Broadcast failed scan alert to ALL devices
                socketio.emit("scan_alert", {
                    "name":       name_val,
                    "is_duplicate": False,
                    "status":     "Scanned Outside Time",
                    "message":    time_error_msg,
                    "details":    {
                        "Name": name_val,
                        "Registration Number": reg_val,
                        "Status": "Scanned Outside Time"
                    },
                    "device_name": device_name,
                })
                
                return {
                    "message": time_error_msg,
                    "error": "time_restricted",
                    "details": {"Name": name_val, "Registration Number": reg_val},
                    "is_duplicate": False
                }, 400

            # Normal check-in flow
            col = hdrs[scan_key]
            curr = ws.cell(row=r, column=col).value or ""
            curr_str = str(curr).strip().lower()
            if curr_str == "scanned" or curr_str == "manual check-in" or "scanned" in curr_str or "manual" in curr_str:
                m = re.search(r"(\d+)", curr_str)
                cnt = int(m.group(1)) if m else 1
            else:
                cnt = 0
            cnt += 1
            
            if device_id == "Manual Check-In":
                new_status = "Manual Check-In" if cnt == 1 else f"Manual Check-In {cnt} Times"
            else:
                new_status = "Scanned" if cnt == 1 else f"Scanned {cnt} Times"
            ws.cell(row=r, column=col, value=new_status)

            ts_col_idx = hdrs.get("scan timestamps")
            dev_col_idx = hdrs.get("scan devices")
            if ts_col_idx:
                prev_ts = str(ws.cell(row=r, column=ts_col_idx).value or "").strip()
                ws.cell(row=r, column=ts_col_idx, value=f"{prev_ts};{now}" if prev_ts else now)
            if dev_col_idx:
                prev_dev = str(ws.cell(row=r, column=dev_col_idx).value or "").strip()
                ws.cell(row=r, column=dev_col_idx, value=f"{prev_dev};{device_name}" if prev_dev else device_name)

            # Get phone number column
            phone_col_idx = hdrs.get("phone number")
            phone_val = str(found_row[phone_col_idx - 1].value or "").strip() if phone_col_idx else ""

            # Dynamic custom columns extraction
            custom_fields = {}
            system_cols = [
                "name", "email address", "registration number", "phone number",
                "unique id", "qr", "barcode", SCAN_COL_NAME.lower(),
                "email sent status", "whatsapp sent status",
                "qr code image", "barcode image"
            ]
            for c_low, c_idx in hdrs.items():
                if c_low not in system_cols:
                    custom_fields[ws.cell(row=1, column=c_idx).value] = str(found_row[c_idx - 1].value or "").strip()

            email_val = str(found_row[hdrs["email address"] - 1].value or "").strip() if "email address" in hdrs else ""

            details = {
                "Name":                name_val,
                "Email":               email_val,
                "Registration Number": reg_val,
                "Phone":               phone_val,
                "Status":              new_status,
                "ScanCount":           cnt,
                "custom_fields":       custom_fields
            }

            # Log to Scan Log worksheet
            log_status = "Checked In" if cnt == 1 else "Duplicate Check-In"
            _log_to_excel_sheet(wb, now, name_val, reg_val, device_name, log_status, qr_data)

            # Atomic write
            _atomic_save(wb, excel_file)
            wb = None  # already closed by _atomic_save

            # 3) Update CSV log (only for registered, on-time check-ins)
            if "Devices" not in scanned_log.columns:
                scanned_log["Devices"] = ""

            mask = scanned_log["QR Data"] == qr_data
            if mask.any():
                idx = scanned_log.index[mask][0]
                scanned_log.at[idx, "Scan Count"] += 1
                prev_ts = scanned_log.at[idx, "Timestamps"]
                scanned_log.at[idx, "Timestamps"] = f"{prev_ts};{now}" if prev_ts else now
                
                prev_devs = str(scanned_log.at[idx, "Devices"]) if pd.notna(scanned_log.at[idx, "Devices"]) else ""
                scanned_log.at[idx, "Devices"] = f"{prev_devs};{device_name}" if prev_devs else device_name
            else:
                new_row = pd.DataFrame([{"QR Data": qr_data, "Scan Count": 1, "Timestamps": now, "Devices": device_name}])
                globals()["scanned_log"] = pd.concat(
                    [scanned_log, new_row], ignore_index=True
                )

            sort_log(scanned_log)
            scanned_log.to_csv(log_file, index=False)

        # Rebuild highlighted file in background ─────────────────────────────
        threading.Thread(target=_rebuild_highlighted, daemon=True).start()

        # Emit real-time update ───────────────────────────────────────────────
        with lock:
            row_data = scanned_log.loc[scanned_log["QR Data"] == qr_data].iloc[0]
        last_ts = row_data["Timestamps"].split(";")[-1].strip()
        last_dev = row_data["Devices"].split(";")[-1].strip() if "Devices" in row_data and row_data["Devices"] else device_name
        
        # Include attendee_name and reg_no directly at root for auto-dashboard DataTable reload
        socketio.emit("row_updated", {
            "qr_data":        qr_data,
            "attendee_name":  name_val,
            "reg_no":         reg_val,
            "scan_count":     int(row_data["Scan Count"]),
            "timestamps":     row_data["Timestamps"],
            "last_timestamp": last_ts,
            "devices":        row_data["Devices"] if "Devices" in row_data else device_name,
            "last_device":    last_dev,
            "details":        details,
        })

        name    = details.get("Name", "")
        status  = details.get("Status", "")
        is_dup  = details.get("ScanCount", 1) > 1
        message = f"✅ {name} — {status}" if name else "✅ Scanned"
        if is_dup:
            message = f"⚠️ {name} already checked in ({status})"

        # Update device scan stats
        with _devices_lock:
            if device_id in connected_devices:
                connected_devices[device_id]["scans"] += 1
                connected_devices[device_id]["last_activity"] = f"Scanned {name}"
                connected_devices[device_id]["last_active_time"] = datetime.now().strftime("%H:%M:%S")
        _broadcast_devices()

        # Broadcast scan alert to ALL devices (not just the scanner)
        socketio.emit("scan_alert", {
            "name":       name,
            "is_duplicate": is_dup,
            "status":     status,
            "message":    message,
            "details":    details,
            "device_name": device_name,
        })

        # Send post-scan notification if configured
        send_scan_notification_async(details, device_name, now)

        # Broadcast updated stats to ALL devices
        _emit_stats()

        # Log to audit trail
        if is_dup:
            _log_audit("Duplicate Check-In", f"Attendee: {name} ({reg_val}) scanned again", device_id)
            return {"message": message, "details": details, "is_duplicate": is_dup, "error": "duplicate_checkin"}, 200
        else:
            _log_audit("Check-In Success", f"Attendee: {name} ({reg_val}) checked in", device_id)
            return {"message": message, "details": details, "is_duplicate": is_dup}, 200

    except PermissionError as e:
        print(f"[checkin] Excel file locked: {e}")
        return {
            "message": "❌ Excel database file is locked (likely open in Microsoft Excel). Please close Excel and try again.",
            "error": "file_locked"
        }, 409
    except Exception as e:
        print(f"[checkin] Error: {e}")
        return {"message": "Internal server error", "error": str(e)}, 500
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass


@app.route("/scan", methods=["POST"])
def scan():
    # Content-type guard
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415

    payload = request.json or {}
    qr_data = payload.get("qr_data", "").strip()
    device_id = payload.get("device_id", "unknown").strip()
    idempotency_key = payload.get("idempotency_key", "").strip()

    if not qr_data:
        return jsonify(message="No QR data provided."), 400

    # Check Idempotency Cache
    if idempotency_key:
        with lock:
            if idempotency_key in _processed_idempotency_keys:
                res, code = _processed_idempotency_keys[idempotency_key]
                return jsonify(res), code

    # Rate limit (IP + Device ID)
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip()
    if _is_rate_limited(ip, device_id):
        return jsonify(message="Too many requests. Please wait."), 429

    res, code = _perform_checkin(qr_data, device_id)
    
    # Save to Idempotency Cache
    if idempotency_key and code == 200:
        with lock:
            if len(_processed_idempotency_keys) > 2000:
                _processed_idempotency_keys.clear()
            _processed_idempotency_keys[idempotency_key] = (res, code)

    return jsonify(res), code


@app.route("/manual_checkin", methods=["POST"])
def manual_checkin():
    # Content-type guard
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415

    payload = request.json or {}
    qr_data = payload.get("qr_data", "").strip()
    idempotency_key = payload.get("idempotency_key", "").strip()

    if not qr_data:
        return jsonify(message="No QR data provided."), 400

    # Check Idempotency Cache
    if idempotency_key:
        with lock:
            if idempotency_key in _processed_idempotency_keys:
                res, code = _processed_idempotency_keys[idempotency_key]
                return jsonify(res), code

    res, code = _perform_checkin(qr_data, device_id="Manual Check-In")

    # Save to Idempotency Cache
    if idempotency_key and code == 200:
        with lock:
            if len(_processed_idempotency_keys) > 2000:
                _processed_idempotency_keys.clear()
            _processed_idempotency_keys[idempotency_key] = (res, code)

    return jsonify(res), code


# ── Quarantine Manager Routes ──────────────────────────────────────────────────
@app.route("/quarantine", methods=["GET"])
def get_quarantine():
    """Return the list of quarantined scans for the active event."""
    q_file = os.path.join(get_active_event_path(), "quarantine.json")
    q_scans = []
    if os.path.exists(q_file):
        try:
            with open(q_file, "r", encoding="utf-8") as f:
                q_scans = json.load(f)
        except Exception:
            pass
    # Only return pending quarantined scans (not approved/rejected ones)
    pending = [s for s in q_scans if s.get("status") == "quarantined"]
    return jsonify(quarantine=pending, total=len(pending))


@app.route("/quarantine/approve", methods=["POST"])
def approve_quarantine():
    """Approve a quarantined scan: perform check-in and remove from quarantine."""
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
    payload = request.json or {}
    q_id = payload.get("id", "").strip()
    device_name = payload.get("device_name", "Admin Dashboard").strip() or "Admin Dashboard"

    if not q_id:
        return jsonify(message="Quarantine ID is required."), 400

    q_file = os.path.join(get_active_event_path(), "quarantine.json")
    q_scans = []
    if os.path.exists(q_file):
        try:
            with open(q_file, "r", encoding="utf-8") as f:
                q_scans = json.load(f)
        except Exception:
            pass

    target = next((s for s in q_scans if s.get("id") == q_id), None)
    if not target:
        return jsonify(message="Quarantine entry not found."), 404

    qr_data = target.get("qr_data", "")
    if not qr_data:
        return jsonify(message="Invalid quarantine entry (missing QR data)."), 400

    # Perform the actual check-in
    res, code = _perform_checkin(qr_data, device_id=f"Approved:{device_name}")
    if code not in (200, 200):
        # Still mark as approved even if duplicate
        pass

    # Mark item as approved and save
    target["status"] = "approved"
    target["approved_by"] = device_name
    target["approved_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with open(q_file, "w", encoding="utf-8") as f:
            json.dump(q_scans, f, indent=4)
    except Exception as ex:
        print(f"[quarantine] Error updating quarantine file: {ex}")

    _log_audit("Quarantine Approved", f"ID: {q_id}, QR: {qr_data}", device_name)
    socketio.emit("quarantine_updated", {})
    return jsonify(message=f"✅ Quarantine entry approved and checked in. {res.get('message', '')}")


@app.route("/quarantine/reject", methods=["POST"])
def reject_quarantine():
    """Reject (delete) a quarantined scan."""
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
    payload = request.json or {}
    q_id = payload.get("id", "").strip()

    if not q_id:
        return jsonify(message="Quarantine ID is required."), 400

    q_file = os.path.join(get_active_event_path(), "quarantine.json")
    q_scans = []
    if os.path.exists(q_file):
        try:
            with open(q_file, "r", encoding="utf-8") as f:
                q_scans = json.load(f)
        except Exception:
            pass

    original_len = len(q_scans)
    q_scans = [s for s in q_scans if s.get("id") != q_id]
    if len(q_scans) == original_len:
        return jsonify(message="Quarantine entry not found."), 404

    try:
        with open(q_file, "w", encoding="utf-8") as f:
            json.dump(q_scans, f, indent=4)
    except Exception as ex:
        print(f"[quarantine] Error saving quarantine file: {ex}")

    _log_audit("Quarantine Rejected", f"ID: {q_id}", "Admin Dashboard")
    socketio.emit("quarantine_updated", {})
    return jsonify(message="✅ Quarantine entry rejected and removed.")


@app.route("/quarantine/send", methods=["POST"])
def send_to_quarantine():
    """Manually send a registration to quarantine."""
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
    payload = request.json or {}
    reg_no = payload.get("reg_no", "").strip()
    reason = payload.get("reason", "Manual Quarantine").strip() or "Manual Quarantine"

    if not reg_no:
        return jsonify(message="Registration number is required."), 400

    # Look up the QR value for this reg_no
    excel_file = get_excel_file()
    qr_data = reg_no  # default fallback
    if os.path.exists(excel_file):
        try:
            with lock:
                wb = load_workbook(excel_file, read_only=True)
                ws = wb.active
                hdrs = _get_headers(ws)
                reg_idx = hdrs.get("registration number")
                qr_idx = hdrs.get("qr")
                if reg_idx and qr_idx:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        r_val = str(row[reg_idx - 1] or "").strip()
                        if r_val.lower() == reg_no.lower():
                            qr_data = str(row[qr_idx - 1] or reg_no).strip()
                            break
                wb.close()
        except Exception:
            pass

    _quarantine_scan(qr_data, "Admin Dashboard", reason)
    _log_audit("Manual Quarantine", f"Reg No: {reg_no}, Reason: {reason}", "Admin Dashboard")
    return jsonify(message=f"✅ Registration '{reg_no}' sent to quarantine.")


# ── Revoke Check-In Route ──────────────────────────────────────────────────────
@app.route("/revoke_checkin", methods=["POST"])
def revoke_checkin():
    """Undo/revoke a check-in for a registered attendee by reg_no."""
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
    payload = request.json or {}
    reg_no = payload.get("reg_no", "").strip()
    reason = payload.get("reason", "Admin Undo").strip() or "Admin Undo"

    if not reg_no:
        return jsonify(message="Registration number is required."), 400

    excel_file = get_excel_file()
    if not os.path.exists(excel_file):
        return jsonify(message="Spreadsheet not found."), 404

    try:
        with lock:
            wb = load_workbook(excel_file)
            ws = wb.active
            hdrs, _ = _get_or_create_headers(ws)

            reg_idx = hdrs.get("registration number")
            scan_key_idx = None
            for k, v in hdrs.items():
                if k == SCAN_COL_NAME.lower():
                    scan_key_idx = v
                    break

            if not reg_idx or not scan_key_idx:
                wb.close()
                return jsonify(message="Required columns not found in spreadsheet."), 500

            found_row = None
            for row in ws.iter_rows(min_row=2):
                r_val = str(row[reg_idx - 1].value or "").strip()
                if r_val.lower() == reg_no.lower():
                    found_row = row
                    break

            if not found_row:
                wb.close()
                return jsonify(message=f"Attendee with reg_no '{reg_no}' not found."), 404

            row_num = found_row[0].row
            name_val = str(found_row[hdrs.get("name", 1) - 1].value or "").strip() if hdrs.get("name") else reg_no

            # Reset scanned status
            ws.cell(row=row_num, column=scan_key_idx, value="")

            # Clear timestamps and devices columns
            ts_col_idx = hdrs.get("scan timestamps")
            dev_col_idx = hdrs.get("scan devices")
            if ts_col_idx:
                ws.cell(row=row_num, column=ts_col_idx, value="")
            if dev_col_idx:
                ws.cell(row=row_num, column=dev_col_idx, value="")

            _atomic_save(wb, excel_file)
            wb = None

        # Also remove from scanned_log CSV
        with lock:
            # Find QR of this reg_no from the attendee map
            qr_map = _get_qr_to_attendee_map()
            qr_to_remove = None
            for qr_val, att in qr_map.items():
                if att.get("reg_no", "").lower() == reg_no.lower():
                    qr_to_remove = qr_val
                    break

            if qr_to_remove:
                mask = scanned_log["QR Data"] == qr_to_remove
                if mask.any():
                    globals()["scanned_log"] = scanned_log[~mask].reset_index(drop=True)
                    scanned_log.to_csv(get_log_file(), index=False)

        _log_audit("Check-In Revoked", f"Attendee: {name_val} ({reg_no}), Reason: {reason}", "Admin Dashboard")
        socketio.emit("registry_updated", {})
        _emit_stats()

        return jsonify(message=f"✅ Check-in for {name_val} ({reg_no}) has been revoked.")
    except Exception as e:
        print(f"[revoke] Error: {e}")
        return jsonify(message=f"Error revoking check-in: {str(e)}"), 500
    finally:
        _wb = locals().get('wb')
        if _wb:
            try:
                _wb.close()
            except Exception:
                pass


# ── Bulk Notify Route ──────────────────────────────────────────────────────────
@app.route("/bulk_notify", methods=["POST"])
def bulk_notify():
    """Trigger bulk notifications for a subgroup or all attendees."""
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
    payload = request.json or {}
    subgroup = payload.get("subgroup", "").strip()
    channel = payload.get("channel", "email").strip().lower()
    group_column = payload.get("group_column", "").strip()

    if channel not in ("email", "whatsapp", "sms", "all"):
        return jsonify(message="Invalid channel. Use: email, whatsapp, sms, or all."), 400

    excel_file = get_excel_file()
    if not os.path.exists(excel_file):
        return jsonify(message="Spreadsheet not found."), 404

    try:
        with lock:
            df_xl = pd.read_excel(excel_file, engine="openpyxl")
        df_xl.columns = df_xl.columns.astype(str).str.strip()
        col_map = {c.lower(): c for c in df_xl.columns}

        reg_col = col_map.get("registration number")
        if not reg_col:
            return jsonify(message="No 'Registration Number' column found."), 500

        # Determine which rows to notify
        if subgroup and group_column:
            matched_col = None
            for c in df_xl.columns:
                if c.lower() == group_column.lower():
                    matched_col = c
                    break
            if matched_col:
                df_filtered = df_xl[df_xl[matched_col].astype(str).str.strip().str.lower() == subgroup.lower()]
            else:
                df_filtered = df_xl
        else:
            df_filtered = df_xl

        reg_nos = df_filtered[reg_col].dropna().astype(str).str.strip().tolist()
        reg_nos = [r for r in reg_nos if r and r.lower() != "nan"]

        if not reg_nos:
            return jsonify(message="No matching attendees found for bulk notification."), 404

        cfg = get_event_config()
        host_url = request.host_url.rstrip("/")

        def _do_bulk_notify():
            sent = 0
            failed = 0
            for reg_no in reg_nos:
                try:
                    if channel in ("email", "all"):
                        _send_single_email_helper(reg_no, cfg)
                        time.sleep(0.2)
                    if channel in ("whatsapp", "all"):
                        _send_single_whatsapp_helper(reg_no, cfg, host_url)
                        time.sleep(0.2)
                    if channel in ("sms", "all"):
                        _send_single_sms_helper(reg_no, cfg, host_url)
                        time.sleep(0.2)
                    sent += 1
                except Exception as ex:
                    print(f"[bulk_notify] Error for {reg_no}: {ex}")
                    failed += 1
            socketio.emit("registry_updated", {})
            print(f"[bulk_notify] Completed: {sent} sent, {failed} failed")

        threading.Thread(target=_do_bulk_notify, daemon=True).start()
        target_desc = f"subgroup '{subgroup}'" if subgroup else "all attendees"
        return jsonify(message=f"✅ Bulk {channel} notifications queued for {len(reg_nos)} attendees in {target_desc}.")
    except Exception as e:
        return jsonify(message=f"Error starting bulk notify: {str(e)}"), 500


# ── Custom Groups Management Routes ───────────────────────────────────────────
def _get_groups_file() -> str:
    return os.path.join(get_active_event_path(), "custom_groups.json")


def _load_groups() -> list:
    gf = _get_groups_file()
    if os.path.exists(gf):
        try:
            with open(gf, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return []


def _save_groups(groups: list) -> None:
    gf = _get_groups_file()
    try:
        with open(gf, "w", encoding="utf-8") as f:
            json.dump(groups, f, indent=4)
    except Exception as ex:
        print(f"[groups] Error saving groups: {ex}")


@app.route("/groups", methods=["GET"])
def get_groups():
    """Return all custom groups for the active event."""
    return jsonify(groups=_load_groups())


@app.route("/groups/save", methods=["POST"])
def save_group():
    """Create or update a custom named group with a list of reg_nos."""
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
    payload = request.json or {}
    group_name = payload.get("name", "").strip()
    reg_nos = payload.get("reg_nos", [])
    description = payload.get("description", "").strip()
    id_card_theme = payload.get("id_card_theme", "").strip()

    if not group_name:
        return jsonify(message="Group name is required."), 400
    if not isinstance(reg_nos, list):
        return jsonify(message="reg_nos must be a list."), 400

    groups = _load_groups()
    existing = next((g for g in groups if g["name"].lower() == group_name.lower()), None)
    
    # Extract template overrides from payload
    id_card_designer_settings = payload.get("id_card_designer_settings")
    email_subject = payload.get("email_subject")
    email_template = payload.get("email_template")

    if existing:
        existing["reg_nos"] = list(set(reg_nos))
        existing["description"] = description
        existing["id_card_theme"] = id_card_theme
        if id_card_designer_settings is not None:
            existing["id_card_designer_settings"] = id_card_designer_settings
        if email_subject is not None:
            existing["email_subject"] = email_subject
        if email_template is not None:
            existing["email_template"] = email_template
        existing["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    else:
        new_grp = {
            "id": f"grp_{random.randint(100000, 999999)}",
            "name": group_name,
            "reg_nos": list(set(reg_nos)),
            "description": description,
            "id_card_theme": id_card_theme,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        if id_card_designer_settings is not None:
            new_grp["id_card_designer_settings"] = id_card_designer_settings
        if email_subject is not None:
            new_grp["email_subject"] = email_subject
        if email_template is not None:
            new_grp["email_template"] = email_template
        groups.append(new_grp)
    _save_groups(groups)
    _log_audit("Custom Group Saved", f"Group: {group_name}, Members: {len(reg_nos)}", "Admin Dashboard")
    socketio.emit("groups_updated", {})
    return jsonify(message=f"✅ Group '{group_name}' saved with {len(reg_nos)} members.")


@app.route("/groups/delete", methods=["POST"])
def delete_group():
    """Delete a custom group by name."""
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
    payload = request.json or {}
    group_name = payload.get("name", "").strip()
    if not group_name:
        return jsonify(message="Group name is required."), 400

    groups = _load_groups()
    original_len = len(groups)
    groups = [g for g in groups if g["name"].lower() != group_name.lower()]
    if len(groups) == original_len:
        return jsonify(message="Group not found."), 404

    _save_groups(groups)
    _log_audit("Custom Group Deleted", f"Group: {group_name}", "Admin Dashboard")
    socketio.emit("groups_updated", {})
    return jsonify(message=f"✅ Group '{group_name}' deleted.")


@app.route("/groups/send_id_cards", methods=["POST"])
def send_group_id_cards():
    """Send ID cards to all members of a custom group."""
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
    payload = request.json or {}
    group_name = payload.get("name", "").strip()
    channel = payload.get("channel", "email").strip().lower()

    groups = _load_groups()
    group = next((g for g in groups if g["name"].lower() == group_name.lower()), None)
    if not group:
        return jsonify(message="Group not found."), 404

    reg_nos = group.get("reg_nos", [])
    if not reg_nos:
        return jsonify(message="Group has no members."), 400

    cfg = get_event_config()
    group_cfg = cfg.copy()
    if "id_card_designer_settings" in group:
        group_cfg["id_card_designer_settings"] = group["id_card_designer_settings"]
    if "email_subject" in group:
        group_cfg["email_subject"] = group["email_subject"]
    if "email_template" in group:
        group_cfg["email_template"] = group["email_template"]

    host_url = request.host_url.rstrip("/")

    def _send_id_cards():
        for reg_no in reg_nos:
            try:
                if channel in ("email", "all"):
                    _send_single_email_helper(reg_no, group_cfg)
                    time.sleep(0.2)
                elif channel == "whatsapp":
                    _send_single_whatsapp_helper(reg_no, group_cfg, host_url)
                    time.sleep(0.2)
            except Exception as ex:
                print(f"[group_id_cards] Error for {reg_no}: {ex}")

    threading.Thread(target=_send_id_cards, daemon=True).start()
    return jsonify(message=f"✅ ID card notifications queued for {len(reg_nos)} members of '{group_name}'.")


@app.route("/registry")
def get_registry():
    try:
        excel_file = get_excel_file()
        if not os.path.exists(excel_file):
            return jsonify([])
            
        with lock:
            df_xl = pd.read_excel(excel_file, engine="openpyxl")

        df_xl.columns = df_xl.columns.astype(str).str.strip()
        col_map = {c.lower(): c for c in df_xl.columns}

        required = {"name", "registration number", "qr"}
        missing  = required - col_map.keys()
        if missing:
            return jsonify(message=f"Excel is missing columns: {missing}"), 500

        scan_col_orig = col_map.get(SCAN_COL_NAME.lower())
        phone_col_orig = col_map.get("phone number")
        email_sent_col = col_map.get("email sent status")
        wa_sent_col    = col_map.get("whatsapp sent status")
        sms_sent_col   = col_map.get("sms sent status")

        records = []
        for index, row in df_xl.iterrows():
            name_v   = _clean_val(row.get(col_map["name"]))
            reg_v    = _clean_val(row.get(col_map["registration number"]))
            qr_v     = _clean_val(row.get(col_map["qr"]))
            status_v = _clean_val(row.get(scan_col_orig), default="Not Scanned") if scan_col_orig else "Not Scanned"
            
            email_v  = _clean_val(row.get(col_map["email address"])) if "email address" in col_map else ""
            phone_v  = _clean_val(row.get(phone_col_orig)) if phone_col_orig else ""
            email_s  = _clean_val(row.get(email_sent_col), default="Not Sent") if email_sent_col else "Not Sent"
            wa_s     = _clean_val(row.get(wa_sent_col),    default="Not Sent") if wa_sent_col    else "Not Sent"
            sms_s    = _clean_val(row.get(sms_sent_col),   default="Not Sent") if sms_sent_col   else "Not Sent"

            if not any([name_v, email_v, reg_v, qr_v]):
                continue

            # Dynamic custom columns — exclude all system-managed columns
            custom_fields = {}
            system_cols = [
                "name", "email address", "registration number", "phone number",
                "unique id", "qr", "barcode", SCAN_COL_NAME.lower(),
                "email sent status", "whatsapp sent status", "sms sent status",
                "qr code image", "barcode image"
            ]
            for c_low, c_orig in col_map.items():
                if c_low not in system_cols:
                    custom_fields[c_orig] = _clean_val(row.get(c_orig))

            records.append({
                "reg_index": int(index + 1),
                "name":      name_v,
                "email":     email_v,
                "reg_no":    reg_v,
                "qr":        qr_v,
                "status":    status_v,
                "phone":     phone_v,
                "email_sent": email_s,
                "wa_sent":    wa_s,
                "sms_sent":   sms_s,
                "custom_fields": custom_fields
            })

        return jsonify(records)
    except Exception as e:
        return jsonify(message="Error loading registry", error=str(e)), 500


@app.route("/active_event_columns")
def active_event_columns():
    try:
        excel_file = get_excel_file()
        if not os.path.exists(excel_file):
            return jsonify(columns=[])
        
        with lock:
            df_xl = pd.read_excel(excel_file, engine="openpyxl")
        
        df_xl.columns = df_xl.columns.astype(str).str.strip()
        system_cols = [
            "name", "email address", "registration number", "phone number",
            "unique id", "qr", "barcode", SCAN_COL_NAME.lower(),
            "email sent status", "whatsapp sent status",
            "qr code image", "barcode image"
        ]
        
        custom_cols = [col for col in df_xl.columns if col.lower() not in system_cols]
        return jsonify(columns=custom_cols)
    except Exception as e:
        return jsonify(message=str(e)), 500


@app.route("/subgroups")
def get_subgroups():
    column = request.args.get("column", "").strip()
    if not column:
        return jsonify(subgroups=[])
    excel_file = get_excel_file()
    if not os.path.exists(excel_file):
        return jsonify(subgroups=[])
    try:
        with lock:
            df_xl = pd.read_excel(excel_file, engine="openpyxl")
        df_xl.columns = df_xl.columns.astype(str).str.strip()
        matched_col = None
        for col in df_xl.columns:
            if col.lower() == column.lower():
                matched_col = col
                break
        if not matched_col:
            return jsonify(subgroups=[])
        unique_vals = df_xl[matched_col].dropna().astype(str).str.strip().unique()
        valid_subgroups = [v for v in unique_vals if v and v.lower() not in ["nan", "none"]]
        return jsonify(subgroups=sorted(valid_subgroups))
    except Exception as e:
        return jsonify(message=f"Error getting subgroups: {str(e)}"), 500


@app.route("/save_subgroup_template", methods=["POST"])
def save_subgroup_template():
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
    payload = request.json or {}
    column = payload.get("column", "").strip()
    subgroup = payload.get("subgroup", "").strip()
    
    if not column or not subgroup:
        return jsonify(message="Column and Subgroup parameters are required."), 400
        
    cfg = get_event_config()
    if "subgroup_templates" not in cfg:
        cfg["subgroup_templates"] = {}
    
    key = f"{column}:{subgroup}"
    cfg["subgroup_templates"][key] = {
        "email_subject": payload.get("email_subject", "").strip(),
        "email_template": payload.get("email_template", ""),
        "whatsapp_template": payload.get("whatsapp_template", ""),
        "sms_template": payload.get("sms_template", "")
    }
    save_event_config(cfg)
    return jsonify(message=f"Templates for subgroup '{subgroup}' saved successfully.")


@app.route("/get_subgroup_template")
def get_subgroup_template():
    column = request.args.get("column", "").strip()
    subgroup = request.args.get("subgroup", "").strip()
    if not column or not subgroup:
        return jsonify(email_subject="", email_template="", whatsapp_template="", sms_template="")
    cfg = get_event_config()
    sub_templates = cfg.get("subgroup_templates", {})
    key = f"{column}:{subgroup}"
    tpls = sub_templates.get(key, {})
    return jsonify(
        email_subject=tpls.get("email_subject", ""),
        email_template=tpls.get("email_template", ""),
        whatsapp_template=tpls.get("whatsapp_template", ""),
        sms_template=tpls.get("sms_template", "")
    )


@app.route("/rename_subgroup", methods=["POST"])
def rename_subgroup():
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
    payload = request.json or {}
    column = payload.get("column", "").strip()
    old_value = payload.get("old_value", "").strip()
    new_value = payload.get("new_value", "").strip()
    
    if not column or not old_value or not new_value:
        return jsonify(message="Column, old_value, and new_value parameters are required."), 400
        
    excel_file = get_excel_file()
    if not os.path.exists(excel_file):
        return jsonify(message="Spreadsheet not found."), 404
        
    try:
        with lock:
            wb = load_workbook(excel_file)
            ws = wb.active
            hdrs, _ = _get_or_create_headers(ws)
            
            matched_col_idx = None
            for col_name, idx in hdrs.items():
                if col_name.lower() == column.lower():
                    matched_col_idx = idx
                    break
            
            if not matched_col_idx:
                wb.close()
                return jsonify(message=f"Column '{column}' not found in spreadsheet."), 404
                
            updated_count = 0
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=matched_col_idx)
                if cell.value is not None and str(cell.value).strip().lower() == old_value.lower():
                    cell.value = new_value
                    updated_count += 1
                    
            if updated_count > 0:
                _atomic_save(wb, excel_file)
            else:
                wb.close()
                
        # Also rename in config subgroup_templates keys if any exist!
        cfg = get_event_config()
        sub_templates = cfg.get("subgroup_templates", {})
        old_key = f"{column}:{old_value}"
        new_key = f"{column}:{new_value}"
        if old_key in sub_templates:
            sub_templates[new_key] = sub_templates.pop(old_key)
            save_event_config(cfg)
            
        return jsonify(message=f"Successfully renamed '{old_value}' to '{new_value}' for {updated_count} attendees.")
    except Exception as e:
        return jsonify(message=f"Error renaming subgroup: {str(e)}"), 500


@app.route("/active_event_form_fields")
def active_event_form_fields():
    try:
        excel_file = get_excel_file()
        if not os.path.exists(excel_file):
            return jsonify(fields=[])
            
        with lock:
            wb = load_workbook(excel_file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1] if cell.value]
            wb.close()
            
        system_cols = {
            "unique id", "qr", "barcode", SCAN_COL_NAME.lower(),
            "email sent status", "whatsapp sent status",
            "qr code image", "barcode image", "scan timestamps", "scan devices"
        }
        
        fields = []
        for h in headers:
            h_clean = str(h).strip()
            h_low = h_clean.lower()
            if h_low not in system_cols:
                is_required = h_low in ["name", "registration number"]
                fields.append({
                    "name": h_clean,
                    "key": h_low,
                    "required": is_required
                })
        return jsonify(fields=fields)
    except Exception as e:
        return jsonify(message=str(e)), 500


@app.route("/clean_duplicates", methods=["POST"])
def clean_duplicates():
    excel_file = get_excel_file()
    if not os.path.exists(excel_file):
        return jsonify(message="No registrations spreadsheet found."), 404
        
    wb = None
    try:
        with lock:
            wb = load_workbook(excel_file)
            ws = wb.active
            hdrs, _ = _get_or_create_headers(ws)
            
            reg_col = hdrs.get("registration number")
            if not reg_col:
                wb.close()
                return jsonify(message="Registration Number column is missing."), 400
                
            seen_regs = set()
            rows_to_delete = []
            
            # Identify duplicates from bottom to top
            for r in range(2, ws.max_row + 1):
                reg_val = ws.cell(row=r, column=reg_col).value
                reg_str = str(reg_val).strip().lower() if reg_val else ""
                
                if not reg_str:
                    continue
                    
                if reg_str in seen_regs:
                    rows_to_delete.append(r)
                else:
                    seen_regs.add(reg_str)
            
            removed_cnt = len(rows_to_delete)
            for r in sorted(rows_to_delete, reverse=True):
                # Clean up images for this row and shift lower ones up
                images_to_keep = []
                for img in ws._images:
                    row_num = None
                    if hasattr(img, 'anchor'):
                        if hasattr(img.anchor, '_from') and hasattr(img.anchor._from, 'row'):
                            row_num = img.anchor._from.row + 1
                        elif isinstance(img.anchor, str):
                            row_num = int(''.join(filter(str.isdigit, img.anchor))) if any(c.isdigit() for c in img.anchor) else None
                    
                    if row_num == r:
                        # Skip this image (deletes it)
                        continue
                    elif row_num is not None and row_num > r:
                        # Shift up
                        if hasattr(img.anchor, '_from') and hasattr(img.anchor._from, 'row'):
                            img.anchor._from.row -= 1
                        elif isinstance(img.anchor, str):
                            col_letters = ''.join(filter(str.isalpha, img.anchor))
                            img.anchor = f"{col_letters}{row_num - 1}"
                    images_to_keep.append(img)
                ws._images = images_to_keep
                
                # Delete the row
                ws.delete_rows(r)
                
            if removed_cnt > 0:
                _atomic_save(wb, excel_file)
            else:
                wb.close()
                
            wb = None
            
        if removed_cnt > 0:
            threading.Thread(target=_rebuild_highlighted, daemon=True).start()
            socketio.emit("registry_updated", {})
            _emit_stats()
            
        return jsonify(message=f"Cleaned {removed_cnt} duplicate rows successfully.", cleaned=removed_cnt), 200
        
    except Exception as e:
        return jsonify(message=f"Error cleaning duplicates: {str(e)}"), 500
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass


@app.route("/regenerate_assets", methods=["POST"])
def regenerate_assets():
    excel_file = get_excel_file()
    if not os.path.exists(excel_file):
        return jsonify(message="No registrations spreadsheet found."), 404
        
    wb = None
    try:
        with lock:
            wb = load_workbook(excel_file)
            ws = wb.active
            hdrs, _ = _get_or_create_headers(ws)
            
            # Clear all existing images from the sheet to regenerate them cleanly
            ws._images = []
            
            # Now run self-healing/generation for every row
            for r in range(2, ws.max_row + 1):
                name_val = ws.cell(row=r, column=hdrs["name"]).value
                reg_val = ws.cell(row=r, column=hdrs["registration number"]).value
                if not name_val or not reg_val:
                    continue
                reg_str = str(reg_val).strip()
                
                # Re-generate QR
                qr_path = os.path.join(get_qr_dir(), f"{reg_str}.png")
                _generate_qr_for_guest(reg_str, reg_str)
                _embed_qr_image(ws, r, qr_path, hdrs["qr code image"])
                
                # Re-generate Barcode
                bc_path = os.path.join(get_barcode_dir(), f"{reg_str}.png")
                _generate_barcode_for_guest(reg_str)
                _embed_barcode_image(ws, r, bc_path, hdrs["barcode image"])
                
                # Re-generate ID Card if enabled
                cfg = get_event_config()
                if cfg.get("enable_id_card_generation"):
                    phone_val = str(ws.cell(row=r, column=hdrs["phone number"]).value or "").strip() if "phone number" in hdrs else ""
                    email_val = str(ws.cell(row=r, column=hdrs["email address"]).value or "").strip() if "email address" in hdrs else ""
                    uid_val = str(ws.cell(row=r, column=hdrs["unique id"]).value or "").strip() if "unique id" in hdrs else ""
                    
                    level = ""
                    group_col = cfg.get("group_column", "")
                    if group_col and group_col.lower() in hdrs:
                        level = str(ws.cell(row=r, column=hdrs[group_col.lower()]).value or "").strip()
                    if not level:
                        for c_low, c_idx in hdrs.items():
                            if "level" in c_low or "subgroup" in c_low:
                                level = str(ws.cell(row=r, column=c_idx).value or "").strip()
                                break
                    try:
                        _generate_id_card(
                            name=str(name_val).strip(),
                            reg_no=reg_str,
                            phone=phone_val,
                            email=email_val,
                            uid=uid_val,
                            qr_path=qr_path,
                            event_name=cfg.get("event_name_template", "the Event"),
                            level=level
                        )
                    except Exception as e:
                        print(f"Error regenerating ID card: {e}")
                
            _atomic_save(wb, excel_file)
            wb = None
            
        threading.Thread(target=_rebuild_highlighted, daemon=True).start()
        socketio.emit("registry_updated", {})
        
        return jsonify(message="Successfully regenerated and re-embedded all QR codes and barcodes."), 200
        
    except Exception as e:
        return jsonify(message=f"Error regenerating assets: {str(e)}"), 500
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass


@app.route("/events", methods=["GET"])
def list_events():
    tree = []
    if os.path.exists(EVENTS_DIR):
        for main_evt in os.listdir(EVENTS_DIR):
            main_path = os.path.join(EVENTS_DIR, main_evt)
            if os.path.isdir(main_path):
                sub_evts = []
                for sub_evt in os.listdir(main_path):
                    sub_path = os.path.join(main_path, sub_evt)
                    if os.path.isdir(sub_path) and sub_evt not in ["qrcodes", "barcodes"]:
                        sub_evts.append(sub_evt)
                tree.append({
                    "name": main_evt,
                    "sub_events": sub_evts
                })
    return jsonify(events=tree, active=active_event)


@app.route("/select_event", methods=["POST"])
def select_event():
    global active_event
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify(message="Event name is required."), 400
    
    clean_name = name.replace("\\", "/").strip("/")
    target_path = os.path.abspath(os.path.join(EVENTS_DIR, clean_name))
    if not target_path.startswith(os.path.abspath(EVENTS_DIR)):
        return jsonify(message="Invalid path."), 400
        
    active_event = clean_name
    load_active_event()
    
    socketio.emit("registry_updated", {})
    _emit_stats()
    
    return jsonify(message=f"Switched active event to '{active_event}'", active=active_event)


@app.route("/create_event", methods=["POST"])
def create_event():
    data = request.json or {}
    name = data.get("name", "").strip()
    parent = data.get("parent", "").strip()
    
    if not name:
        return jsonify(message="Event name is required."), 400
        
    clean_name = re.sub(r'[\\/*?:"<>|]', "", name).strip()
    if not clean_name:
        return jsonify(message="Invalid event name."), 400
        
    if parent:
        parent_clean = parent.replace("\\", "/").strip("/")
        relative_path = f"{parent_clean}/{clean_name}"
    else:
        relative_path = clean_name
        
    target_path = os.path.join(EVENTS_DIR, relative_path)
    if os.path.exists(target_path):
        return jsonify(message="Event folder already exists."), 409
        
    os.makedirs(target_path, exist_ok=True)
    
    # Initialize registrations.xlsx inside it
    excel_path = os.path.join(target_path, "registrations.xlsx")
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "Name", "Email Address", "Registration Number", "Phone Number",
        "Unique ID", "QR", "Barcode", SCAN_COL_NAME, "Email Sent Status", "WhatsApp Sent Status",
        "Scan Timestamps", "Scan Devices", "QR Code Image", "Barcode Image"
    ]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    wb.save(excel_path)
    wb.close()
    
    # Initialize scanned_log.csv
    log_path = os.path.join(target_path, "scanned_log.csv")
    df = pd.DataFrame(columns=["QR Data", "Scan Count", "Timestamps", "Devices"])
    df.to_csv(log_path, index=False)
    
    # Create subfolders for qrcodes and barcodes
    os.makedirs(os.path.join(target_path, "qrcodes"), exist_ok=True)
    os.makedirs(os.path.join(target_path, "barcodes"), exist_ok=True)
    
    return jsonify(message=f"Event '{relative_path}' created successfully.", name=relative_path)


@app.route("/get_config", methods=["GET"])
def get_config():
    return jsonify(get_event_config())


@app.route("/save_config", methods=["POST"])
def save_config():
    data = request.json or {}
    cfg = get_event_config()
    for k, v in data.items():
        cfg[k] = v
    save_event_config(cfg)
    return jsonify(message="Configuration saved successfully.")


@app.route("/help_content", methods=["GET"])
def get_help_content():
    help_path = os.path.join(BASE_DIR, "HELP.md")
    if os.path.exists(help_path):
        try:
            with open(help_path, "r", encoding="utf-8") as f:
                content = f.read()
            return jsonify(content=content)
        except Exception as e:
            return jsonify(content=f"Error reading HELP.md: {str(e)}"), 500
    return jsonify(content="HELP.md not found."), 404


_tunnel_process = None
_tunnel_url = None
_tunnel_active = False
_tunnel_lock = threading.Lock()

def _run_tunnel():
    global _tunnel_process, _tunnel_url, _tunnel_active
    import subprocess
    import re
    
    print("[tunnel] Starting localhost.run SSH tunnel loop...")
    # Add ServerAlive keepalives to prevent disconnect drops and bypass interactive prompt
    cmd = ["ssh", "-o", "StrictHostKeyChecking=no", "-o", "UserKnownHostsFile=NUL" if os.name == "nt" else "/dev/null", "-o", "ServerAliveInterval=30", "-o", "ServerAliveCountMax=3", "-R", "80:127.0.0.1:5001", "nokey@localhost.run"]
    
    url_pattern = re.compile(r"https?://[a-zA-Z0-9.-]+\.lhr\.(?:life|link|run|tunnel)")
    lhrtunnel_pattern = re.compile(r"https?://[a-zA-Z0-9.-]+\.lhrtunnel\.link")
    
    _tunnel_active = True
    socketio.emit("tunnel_status_update", {"active": True, "url": "Connecting..."})
    
    while _tunnel_active:
        try:
            startupinfo = None
            if os.name == 'nt':
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                startupinfo.wShowWindow = subprocess.SW_HIDE

            with _tunnel_lock:
                if not _tunnel_active:
                    break
                _tunnel_process = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1,
                    startupinfo=startupinfo
                )
            
            while _tunnel_active and _tunnel_process.poll() is None:
                line = _tunnel_process.stdout.readline()
                if not line:
                    break
                print(f"[tunnel-ssh] {line.strip()}")
                
                match = url_pattern.search(line) or lhrtunnel_pattern.search(line)
                if match:
                    _tunnel_url = match.group(0)
                    print(f"[tunnel] Extracted tunnel URL: {_tunnel_url}")
                    socketio.emit("tunnel_status_update", {"active": True, "url": _tunnel_url})
            
            if _tunnel_active:
                print("[tunnel] SSH tunnel dropped. Reconnecting in 3 seconds...")
                socketio.emit("tunnel_status_update", {"active": True, "url": "Reconnecting..."})
                time.sleep(3)
        except Exception as e:
            print(f"[tunnel] Loop Error: {e}")
            if _tunnel_active:
                time.sleep(3)
                
    print("[tunnel] Tunnel SSH loop fully stopped.")
    socketio.emit("tunnel_status_update", {"active": False, "url": None})


@app.route("/start_tunnel", methods=["POST"])
def start_tunnel():
    global _tunnel_active
    with _tunnel_lock:
        if _tunnel_active:
            return jsonify(message="Tunnel is already active.", url=_tunnel_url), 200
            
    threading.Thread(target=_run_tunnel, daemon=True).start()
    return jsonify(message="Tunnel starting..."), 202


@app.route("/stop_tunnel", methods=["POST"])
def stop_tunnel():
    global _tunnel_process, _tunnel_active, _tunnel_url
    with _tunnel_lock:
        if _tunnel_process:
            try:
                _tunnel_process.terminate()
                _tunnel_process.kill()
            except Exception:
                pass
        _tunnel_active = False
        _tunnel_url = None
        _tunnel_process = None
    socketio.emit("tunnel_status_update", {"active": False, "url": None})
    return jsonify(message="Tunnel stopped.")


@app.route("/tunnel_status", methods=["GET"])
def tunnel_status():
    return jsonify(active=_tunnel_active, url=_tunnel_url)


@app.route("/tunnel_qr", methods=["GET"])
def tunnel_qr():
    data = request.args.get("data", "").strip()
    if not data:
        return "Missing data parameter", 400
    try:
        qr = qrcode.QRCode(version=1, box_size=5, border=1)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return send_file(buf, mimetype="image/png")
    except Exception as e:
        print(f"[tunnel_qr] Error generating local QR: {e}")
        return "Internal server error", 500


def _get_or_create_headers(ws) -> tuple[dict[str, int], bool]:
    hdrs = _get_headers(ws)
    modified = False
    
    # Define aliases mapping: target_lowercase -> list of lowercase aliases
    aliases = {
        "name": ["name", "fullname", "full name", "attendee name", "visitor name", "guest name"],
        "registration number": ["registration number", "reg number", "reg no", "registration no", "reg_no", "registration_number", "id", "ticket number", "ticket no", "regid"],
        "email address": ["email address", "email", "mail", "email_address"],
        "phone number": ["phone number", "phone", "mobile", "whatsapp", "contact", "phone_number", "contact number", "contact no"],
        "qr": ["qr", "qr code", "qr_code", "qr value", "qr_value"],
        "barcode": ["barcode", "barcode value", "barcode_value"]
    }
    
    # For each target, if it is not in hdrs, see if any alias is in hdrs
    for target, alias_list in aliases.items():
        if target not in hdrs:
            for alias in alias_list:
                if alias in hdrs:
                    col_idx = hdrs[alias]
                    standard_names = {
                        "name": "Name",
                        "registration number": "Registration Number",
                        "email address": "Email Address",
                        "phone number": "Phone Number",
                        "qr": "QR",
                        "barcode": "Barcode"
                    }
                    ws.cell(row=1, column=col_idx, value=standard_names[target])
                    hdrs[target] = col_idx
                    del hdrs[alias]
                    modified = True
                    break
                    
    # Now define standard expected columns and their nice names
    expected = {
        "name": "Name",
        "registration number": "Registration Number",
        "unique id": "Unique ID",
        "qr": "QR",
        "barcode": "Barcode",
        SCAN_COL_NAME.lower(): SCAN_COL_NAME,
        "scan timestamps": "Scan Timestamps",
        "scan devices": "Scan Devices",
        "qr code image": "QR Code Image",
        "barcode image": "Barcode Image"
    }
    
    if "email address" in hdrs:
        expected["email sent status"] = "Email Sent Status"
    if "phone number" in hdrs:
        expected["whatsapp sent status"] = "WhatsApp Sent Status"
        expected["sms sent status"] = "SMS Sent Status"
        
    for key, val in expected.items():
        if key not in hdrs:
            col = ws.max_column + 1
            ws.cell(row=1, column=col, value=val)
            hdrs[key] = col
            modified = True
            
    return hdrs, modified


def _has_image_at_cell(ws, row: int, col: int) -> bool:
    """Check if there is already an image anchored to the given row and column (1-indexed)."""
    for img in ws._images:
        row_num = None
        col_num = None
        if hasattr(img, 'anchor'):
            if hasattr(img.anchor, '_from') and hasattr(img.anchor._from, 'row'):
                row_num = img.anchor._from.row + 1
                col_num = img.anchor._from.col + 1
            elif isinstance(img.anchor, str):
                try:
                    # Parse coord like 'F2' -> row=2, col=6
                    cell = ws[img.anchor]
                    row_num = cell.row
                    row_col = cell.column
                except Exception:
                    pass
        if row_num == row and col_num == col:
            return True
    return False


def _initialize_missing_metadata(ws) -> bool:
    hdrs, modified = _get_or_create_headers(ws)
    
    # Check if there are any rows
    if ws.max_row < 2:
        return modified
        
    existing_uids = set()
    for r in range(2, ws.max_row + 1):
        uid_val = ws.cell(row=r, column=hdrs["unique id"]).value
        if uid_val:
            existing_uids.add(str(uid_val).strip().upper())
            
    row_modified = False
    for r in range(2, ws.max_row + 1):
        name_val = ws.cell(row=r, column=hdrs["name"]).value
        reg_val = ws.cell(row=r, column=hdrs["registration number"]).value
        if not name_val or not reg_val:
            continue
            
        reg_str = str(reg_val).strip()
        
        # 1) Unique ID
        uid_cell = ws.cell(row=r, column=hdrs["unique id"])
        if not uid_cell.value:
            uid = _generate_unique_id(existing_uids)
            uid_cell.value = uid
            existing_uids.add(uid)
            row_modified = True
            
        # 2) QR value
        qr_cell = ws.cell(row=r, column=hdrs["qr"])
        if not qr_cell.value:
            qr_cell.value = reg_str
            row_modified = True
            
        # 3) Barcode value
        bc_cell = ws.cell(row=r, column=hdrs["barcode"])
        if not bc_cell.value:
            bc_cell.value = reg_str
            row_modified = True
            
        # 4) Scanned Status
        status_cell = ws.cell(row=r, column=hdrs[SCAN_COL_NAME.lower()])
        if status_cell.value is None:
            status_cell.value = ""
            row_modified = True
            
        # 5) Email Sent Status
        if "email sent status" in hdrs:
            esc_cell = ws.cell(row=r, column=hdrs["email sent status"])
            if esc_cell.value is None:
                esc_cell.value = "Not Sent"
                row_modified = True
                
        # 6) WhatsApp Sent Status
        if "whatsapp sent status" in hdrs:
            wsc_cell = ws.cell(row=r, column=hdrs["whatsapp sent status"])
            if wsc_cell.value is None:
                wsc_cell.value = "Not Sent"
                row_modified = True

        # 7) SMS Sent Status
        if "sms sent status" in hdrs:
            sms_cell = ws.cell(row=r, column=hdrs["sms sent status"])
            if sms_cell.value is None:
                sms_cell.value = "Not Sent"
                row_modified = True
                
        # 7) QR Image
        qr_path = os.path.join(get_qr_dir(), f"{reg_str}.png")
        if not os.path.exists(qr_path) or not _has_image_at_cell(ws, r, hdrs["qr code image"]):
            _generate_qr_for_guest(reg_str, reg_str)
            _embed_qr_image(ws, r, qr_path, hdrs["qr code image"])
            row_modified = True
            
        # 8) Barcode Image
        bc_path = os.path.join(get_barcode_dir(), f"{reg_str}.png")
        if not os.path.exists(bc_path) or not _has_image_at_cell(ws, r, hdrs["barcode image"]):
            _generate_barcode_for_guest(reg_str)
            _embed_barcode_image(ws, r, bc_path, hdrs["barcode image"])
            row_modified = True
            
    return modified or row_modified


def _get_qr_payload(reg_no: str, cfg: dict = None) -> str:
    """Return the QR code text payload. Signs it cryptographically if enabled in config."""
    if cfg is None:
        cfg = get_event_config()
    
    if cfg.get("cryptographic_qr_verification", False):
        import hmac
        import hashlib
        key = cfg.get("event_signing_key", "").encode("utf-8")
        if key:
            sig = hmac.new(key, reg_no.encode("utf-8"), hashlib.sha256).hexdigest()[:12]
            return f"{reg_no}:{sig}"
    return reg_no


def _verify_qr_payload(qr_data: str, cfg: dict = None) -> tuple[bool, str]:
    """Verifies a scanned QR payload.
    If cryptographic verification is enabled:
      - Returns (True, reg_no) if signature is valid.
      - Returns (False, "") if signature is invalid or missing.
    If disabled:
      - Returns (True, qr_data) (since verification is disabled, any lookup is ok).
    """
    if cfg is None:
        cfg = get_event_config()
        
    qr_data_norm = qr_data.strip()
    if not cfg.get("cryptographic_qr_verification", False):
        return True, qr_data_norm
        
    if ":" in qr_data_norm:
        parts = qr_data_norm.split(":")
        if len(parts) == 2:
            reg_no, sig = parts
            import hmac
            import hashlib
            key = cfg.get("event_signing_key", "").encode("utf-8")
            expected_sig = hmac.new(key, reg_no.encode("utf-8"), hashlib.sha256).hexdigest()[:12]
            if hmac.compare_digest(sig, expected_sig):
                return True, reg_no
                
    return False, ""


def _generate_qr_for_guest(content_str: str, reg_no: str) -> str:
    qr_dir = get_qr_dir()
    qr_path = os.path.join(qr_dir, f"{reg_no.strip()}.png")
    qr_img = qrcode.make(content_str.strip())
    qr_img.save(qr_path)
    return qr_path


def _generate_barcode_for_guest(reg_no: str) -> str:
    import barcode
    from barcode.writer import ImageWriter
    code_class = barcode.get_barcode_class('code128')
    writer = ImageWriter()
    options = {'write_text': True, 'font_size': 9, 'text_distance': 3.0, 'module_height': 12.0}
    
    cleaned_reg = str(reg_no).strip().upper()
    barcode_dir = get_barcode_dir()
    barcode_path_no_ext = os.path.join(barcode_dir, cleaned_reg)
    b = code_class(cleaned_reg, writer=writer)
    b.save(barcode_path_no_ext, options=options)
    return barcode_path_no_ext + ".png"


def _get_attendee_level(row_dict: dict) -> str:
    """Helper to find the attendee level/ticket class/category/role in row data."""
    if not row_dict:
        return ""
    keys_to_check = ["ticket class", "ticket type", "category", "role", "level", "type", "ticket_class", "ticket_type", "ticket"]
    for k, v in row_dict.items():
        if str(k).strip().lower() in keys_to_check:
            val_str = str(v).strip()
            if val_str and val_str != "nan" and val_str != "None":
                return val_str
    return ""


def _generate_id_card(name: str, reg_no: str, phone: str, email: str, uid: str, qr_path: str, event_name: str, level: str = "", cfg: dict = None) -> tuple[str, str]:
    """
    Generates a beautifully styled vertical ID card (600x900 pixels) with a dark gradient,
    abstract design accents, event info, attendee details, and the embedded QR code.
    Supports dynamic themes based on event name, and color accents based on attendee level.
    Saves it as both PNG and PDF, returning (pdf_path, png_path).
    """
    from PIL import Image, ImageDraw, ImageFont
    import re
    
    # 1. Base Setup
    width, height = 600, 900
    card = Image.new("RGBA", (width, height), (255, 255, 255, 0))
    draw = ImageDraw.Draw(card)
    
    # Define Themes
    theme_configs = {
        "cyber_neon": {
            "bg_start": (11, 15, 25),      # Deep navy
            "bg_end": (26, 35, 51),        # Navy gray
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (200, 220, 240, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (0, 245, 160), # Neon Cyan
            "shapes": [
                {"type": "arc", "box": [-200, 600, 300, 1100], "color": (255, 107, 53, 30), "width": 45}, # Bottom-left orange glow
                {"type": "arc", "box": [300, -200, 800, 300], "color": (0, 245, 160, 30), "width": 40}  # Top-right cyan glow
            ]
        },
        "sunset_glow": {
            "bg_start": (40, 10, 35),      # Deep purple-plum
            "bg_end": (80, 25, 30),        # Wine red
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (255, 210, 210, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (255, 107, 53), # Neon Orange/Sunset
            "shapes": [
                {"type": "arc", "box": [-100, -100, 400, 400], "color": (255, 190, 11, 30), "width": 50}, # Top-left yellow sun
                {"type": "circle", "box": [400, 700, 580, 880], "color": (251, 86, 196, 25)}           # Bottom-right pink circle
            ]
        },
        "emerald_luxe": {
            "bg_start": (5, 30, 20),       # Deep forest green
            "bg_end": (15, 60, 40),        # Emerald green
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (210, 240, 220, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (218, 165, 32), # Satin Gold
            "shapes": [
                {"type": "arc", "box": [300, 600, 800, 1100], "color": (46, 204, 113, 20), "width": 40},
                {"type": "line", "coords": [(0, 100), (600, 300)], "color": (218, 165, 32, 15), "width": 4}
            ]
        },
        "royal_amethyst": {
            "bg_start": (20, 10, 45),      # Dark indigo
            "bg_end": (55, 20, 90),        # Royal purple
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (230, 210, 255, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (224, 86, 253), # Bright magenta
            "shapes": [
                {"type": "arc", "box": [-150, 650, 250, 1050], "color": (104, 109, 224, 25), "width": 40},
                {"type": "arc", "box": [350, -150, 750, 250], "color": (224, 86, 253, 25), "width": 30}
            ]
        },
        "classic_corporate": {
            "bg_start": (20, 25, 35),      # Steel dark gray
            "bg_end": (45, 52, 68),        # Slate blue-gray
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (220, 225, 235, 230),
            "border": (255, 255, 255, 15),
            "default_accent": (52, 152, 219), # Corporate blue
            "shapes": [
                {"type": "rectangle", "box": [0, 0, 600, 10], "color": (52, 152, 219, 100)},
                {"type": "arc", "box": [350, 650, 750, 1050], "color": (127, 140, 141, 20), "width": 35}
            ]
        },
        "midnight_gold": {
            "bg_start": (5, 5, 5),         # Solid rich black
            "bg_end": (20, 20, 20),        # Very dark gray
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (210, 180, 140, 230), # Light bronze/tan
            "border": (218, 165, 32, 40),  # Gold border
            "default_accent": (218, 165, 32), # Gold
            "shapes": [
                {"type": "circle", "box": [-100, 700, 150, 950], "color": (218, 165, 32, 10)},
                {"type": "circle", "box": [450, -50, 650, 150], "color": (218, 165, 32, 10)}
            ]
        },
        "aurora_borealis": {
            "bg_start": (5, 25, 25),       # Deep polar green
            "bg_end": (10, 50, 45),        # Deep aurora teal
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (200, 240, 230, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (46, 204, 113), # Aurora green
            "shapes": [
                {"type": "arc", "box": [-200, 500, 400, 1100], "color": (52, 152, 219, 25), "width": 40},
                {"type": "circle", "box": [350, -100, 700, 250], "color": (46, 204, 113, 20)}
            ]
        },
        "frozen_glacier": {
            "bg_start": (15, 30, 45),      # Glacier navy
            "bg_end": (40, 75, 105),       # Ice blue
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (220, 235, 250, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (135, 206, 250), # Glacier light blue
            "shapes": [
                {"type": "circle", "box": [-50, -50, 250, 250], "color": (255, 255, 255, 15)},
                {"type": "arc", "box": [200, 600, 700, 1100], "color": (135, 206, 250, 25), "width": 30}
            ]
        },
        "cherry_blossom": {
            "bg_start": (50, 15, 25),      # Deep berry pink
            "bg_end": (90, 40, 55),        # Cherry pink
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (255, 220, 225, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (255, 182, 193), # Blossom pink
            "shapes": [
                {"type": "circle", "box": [400, 700, 580, 880], "color": (255, 182, 193, 25)},
                {"type": "circle", "box": [-100, 50, 150, 300], "color": (231, 76, 60, 20)}
            ]
        },
        "desert_sand": {
            "bg_start": (45, 25, 15),      # Copper clay
            "bg_end": (85, 50, 30),        # Terracotta sand
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (250, 230, 210, 230),
            "border": (255, 255, 255, 15),
            "default_accent": (244, 164, 96), # Sandy gold
            "shapes": [
                {"type": "arc", "box": [-200, -100, 300, 400], "color": (218, 165, 32, 25), "width": 40},
                {"type": "circle", "box": [300, 650, 700, 1050], "color": (244, 164, 96, 20)}
            ]
        },
        "ocean_breeze": {
            "bg_start": (10, 20, 40),      # Abyssal blue
            "bg_end": (20, 55, 85),        # Marine turquoise
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (210, 245, 245, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (72, 209, 204), # Ocean aquamarine
            "shapes": [
                {"type": "arc", "box": [300, -150, 750, 300], "color": (72, 209, 204, 25), "width": 45},
                {"type": "circle", "box": [-150, 600, 250, 1000], "color": (52, 152, 219, 20)}
            ]
        },
        "volcanic_ash": {
            "bg_start": (20, 15, 15),      # Volcanic black
            "bg_end": (40, 30, 30),        # Obsidian ash
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (230, 210, 210, 230),
            "border": (255, 255, 255, 15),
            "default_accent": (231, 76, 60), # Fiery lava red
            "shapes": [
                {"type": "arc", "box": [-100, 650, 300, 1050], "color": (231, 76, 60, 25), "width": 40},
                {"type": "circle", "box": [400, -100, 700, 200], "color": (211, 84, 0, 20)}
            ]
        },
        "lavender_mist": {
            "bg_start": (30, 20, 45),      # Deep amethyst purple
            "bg_end": (60, 45, 85),        # Violet mist
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (240, 230, 250, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (230, 230, 250), # Soft lavender
            "shapes": [
                {"type": "circle", "box": [-50, 650, 250, 950], "color": (142, 68, 173, 20)},
                {"type": "arc", "box": [350, -100, 700, 250], "color": (230, 230, 250, 25), "width": 30}
            ]
        },
        "monochrome_sleek": {
            "bg_start": (15, 15, 15),      # Jet black
            "bg_end": (35, 35, 35),        # Brushed metal slate
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (220, 220, 220, 230),
            "border": (255, 255, 255, 25),
            "default_accent": (192, 192, 192), # Sleek silver
            "shapes": [
                {"type": "rectangle", "box": [0, 0, 600, 8], "color": (255, 255, 255, 40)},
                {"type": "arc", "box": [350, 650, 750, 1050], "color": (127, 140, 141, 15), "width": 30}
            ]
        },
        "vintage_sepia": {
            "bg_start": (35, 25, 15),      # Dark chocolate
            "bg_end": (65, 50, 35),        # Warm sepia
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (245, 222, 179, 230),
            "border": (255, 255, 255, 15),
            "default_accent": (205, 133, 63), # Aged parchment brown
            "shapes": [
                {"type": "circle", "box": [-100, -100, 300, 300], "color": (139, 69, 19, 15)},
                {"type": "arc", "box": [250, 600, 650, 1000], "color": (245, 222, 179, 20), "width": 35}
            ]
        },
        "electric_violet": {
            "bg_start": (25, 5, 40),       # Cosmic violet
            "bg_end": (55, 15, 80),        # Electric indigo
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (235, 215, 255, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (130, 88, 250), # Neon violet
            "shapes": [
                {"type": "arc", "box": [-150, 600, 250, 1000], "color": (0, 245, 160, 20), "width": 40},
                {"type": "circle", "box": [350, -50, 650, 250], "color": (130, 88, 250, 25)}
            ]
        },
        "forest_moss": {
            "bg_start": (15, 25, 10),      # Moss black
            "bg_end": (40, 55, 30),        # Deep forest pine
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (220, 240, 210, 230),
            "border": (255, 255, 255, 15),
            "default_accent": (173, 223, 173), # Mossy green
            "shapes": [
                {"type": "circle", "box": [450, 700, 650, 900], "color": (218, 165, 32, 15)},
                {"type": "arc", "box": [-150, -150, 350, 350], "color": (46, 204, 113, 20), "width": 40}
            ]
        },
        "golden_hour": {
            "bg_start": (55, 20, 10),      # Crimson sunset
            "bg_end": (90, 50, 15),        # Honey gold
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (255, 230, 200, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (255, 140, 0), # Sunny orange
            "shapes": [
                {"type": "circle", "box": [-150, 650, 250, 1050], "color": (255, 69, 0, 20)},
                {"type": "arc", "box": [300, -100, 750, 350], "color": (255, 215, 0, 25), "width": 35}
            ]
        },
        "berry_smoothie": {
            "bg_start": (50, 10, 35),      # Mulberry
            "bg_end": (90, 20, 60),        # Raspberry
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (255, 215, 235, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (255, 20, 147), # Deep berry pink
            "shapes": [
                {"type": "circle", "box": [-50, -50, 250, 250], "color": (75, 0, 130, 20)},
                {"type": "arc", "box": [250, 600, 650, 1000], "color": (255, 20, 147, 25), "width": 40}
            ]
        },
        "space_nebula": {
            "bg_start": (10, 5, 25),       # Deep space black
            "bg_end": (25, 10, 45),        # Cosmic indigo
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (215, 225, 245, 230),
            "border": (255, 255, 255, 15),
            "default_accent": (186, 85, 211), # Nebula purple
            "shapes": [
                {"type": "circle", "box": [350, 650, 650, 950], "color": (0, 245, 160, 20)},
                {"type": "circle", "box": [-100, -50, 250, 300], "color": (186, 85, 211, 20)}
            ]
        },
        "mint_fresh": {
            "bg_start": (10, 35, 30),      # Dark mint
            "bg_end": (25, 70, 60),        # Vibrant teal green
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (220, 250, 240, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (152, 251, 152), # Mint green
            "shapes": [
                {"type": "arc", "box": [300, 600, 750, 1050], "color": (26, 188, 156, 20), "width": 45},
                {"type": "circle", "box": [-150, -100, 250, 300], "color": (152, 251, 152, 25)}
            ]
        },
        "citrus_twist": {
            "bg_start": (55, 35, 10),      # Citrus bronze
            "bg_end": (90, 75, 15),        # Lime orange yellow
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (255, 250, 210, 230),
            "border": (255, 255, 255, 15),
            "default_accent": (255, 165, 0), # Sunny orange
            "shapes": [
                {"type": "circle", "box": [-50, 650, 250, 950], "color": (218, 165, 32, 20)},
                {"type": "arc", "box": [350, -100, 700, 250], "color": (173, 255, 47, 20), "width": 35}
            ]
        },
        "rose_gold": {
            "bg_start": (45, 25, 30),      # Dusty burgundy
            "bg_end": (75, 45, 50),        # Rose champagne gold
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (255, 230, 230, 230),
            "border": (218, 165, 32, 30),  # Rose gold border
            "default_accent": (235, 160, 160), # Soft rose gold
            "shapes": [
                {"type": "circle", "box": [-100, -100, 300, 300], "color": (218, 165, 32, 15)},
                {"type": "arc", "box": [250, 600, 650, 1000], "color": (235, 160, 160, 25), "width": 30}
            ]
        },
        "carbon_fiber": {
            "bg_start": (10, 10, 12),      # Dark graphite
            "bg_end": (22, 22, 26),        # Racing black grid
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (210, 215, 225, 230),
            "border": (255, 255, 255, 15),
            "default_accent": (231, 76, 60), # Racing red
            "shapes": [
                {"type": "rectangle", "box": [0, 0, 12, 900], "color": (231, 76, 60, 150)},
                {"type": "arc", "box": [300, 600, 750, 1050], "color": (127, 140, 141, 15), "width": 35}
            ]
        },
        "royal_sapphire": {
            "bg_start": (5, 15, 40),       # Imperial navy
            "bg_end": (15, 35, 80),        # Sapphire blue
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (215, 230, 255, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (30, 144, 255), # Royal blue
            "shapes": [
                {"type": "circle", "box": [-150, 600, 250, 1000], "color": (30, 144, 255, 25)},
                {"type": "arc", "box": [350, -150, 750, 250], "color": (255, 255, 255, 15), "width": 45}
            ]
        },
        "candy_pop": {
            "bg_start": (45, 10, 45),      # Orchid berry
            "bg_end": (80, 20, 80),        # Hot pink magenta
            "text_primary": (255, 255, 255, 255),
            "text_secondary": (255, 220, 255, 230),
            "border": (255, 255, 255, 20),
            "default_accent": (255, 105, 180), # Bubblegum hot pink
            "shapes": [
                {"type": "arc", "box": [-200, 600, 300, 1100], "color": (0, 245, 160, 30), "width": 40},
                {"type": "circle", "box": [350, -50, 650, 250], "color": (255, 105, 180, 25)}
            ]
        }
    }
    
    
    # ── Custom Designer Layout Drawing ──────────────────────────────────────
    designer_settings = cfg.get("id_card_designer_settings") if cfg else None
    
    # Check if the custom designer layout is active and valid
    if designer_settings and designer_settings.get("mode") in ("custom", "custom_html"):
        if designer_settings.get("mode") == "custom_html":
            try:
                html_content = designer_settings.get("html", "")
                css_content = designer_settings.get("css", "")
                
                # Replace placeholders
                resolved_html = html_content
                # Split name into Surname and Given Names for templates that separate them
                name_parts = name.strip().split()
                if len(name_parts) > 1:
                    surname = name_parts[-1]
                    given_names = " ".join(name_parts[:-1])
                else:
                    surname = name
                    given_names = ""
                resolved_html = resolved_html.replace("{Surname}", surname)
                resolved_html = resolved_html.replace("{Given Names}", given_names)
                resolved_html = resolved_html.replace("{Name}", name)
                resolved_html = resolved_html.replace("{Registration Number}", reg_no)
                resolved_html = resolved_html.replace("{Email Address}", email)
                resolved_html = resolved_html.replace("{Phone Number}", phone)
                resolved_html = resolved_html.replace("{Unique ID}", uid)
                resolved_html = resolved_html.replace("{Pass Type}", level)
                resolved_html = resolved_html.replace("{Event}", event_name)
                
                # Resolve QR code base64
                qr_base64 = ""
                if qr_path and os.path.exists(qr_path):
                    import base64
                    with open(qr_path, "rb") as f_qr:
                        qr_base64 = base64.b64encode(f_qr.read()).decode("utf-8")
                
                if qr_base64:
                    resolved_html = resolved_html.replace("{QR_CODE_URL}", f"data:image/png;base64,{qr_base64}")
                else:
                    resolved_html = resolved_html.replace("{QR_CODE_URL}", "")
                
                # Use a high-quality professional portrait placeholder for PHOTO_URL
                resolved_html = resolved_html.replace("{PHOTO_URL}", "https://images.unsplash.com/photo-1507003211169-0a1dd7228f2d?auto=format&fit=crop&q=80&w=200")
                
                # Strip body layout styles (min-height, height, display, justify-content, align-items, margin)
                import re
                cleaned_css = css_content or ""
                def _strip_body(m):
                    body_rules = m.group(1)
                    body_rules = re.sub(r'(?:min-)?height\s*:[^;]+;?', '', body_rules, flags=re.I)
                    body_rules = re.sub(r'display\s*:[^;]+;?', '', body_rules, flags=re.I)
                    body_rules = re.sub(r'justify-content\s*:[^;]+;?', '', body_rules, flags=re.I)
                    body_rules = re.sub(r'align-items\s*:[^;]+;?', '', body_rules, flags=re.I)
                    body_rules = re.sub(r'margin\s*:[^;]+;?', '', body_rules, flags=re.I)
                    return f"body {{{body_rules}}}"
                cleaned_css = re.sub(r'body\s*{(.*?)}', _strip_body, cleaned_css, flags=re.DOTALL | re.I)
                
                # Build complete HTML page
                full_document = f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="utf-8">
                    <script src="https://cdn.tailwindcss.com"></script>
                    <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200" />
                    <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;800&family=Outfit:wght@300;400;600;700;800&family=Roboto+Mono:wght@400;700&display=swap" />
                    <style>
                        body {{
                            margin: 0;
                            padding: 0;
                            background: transparent;
                            display: flex;
                            align-items: center;
                            justify-content: center;
                            width: 340px;
                            height: 540px;
                            overflow: hidden;
                        }}
                        {cleaned_css}
                    </style>
                </head>
                <body>
                    {resolved_html}
                </body>
                </html>
                """
                
                id_dir = os.path.join(get_qr_dir(), "id_cards")
                os.makedirs(id_dir, exist_ok=True)
                png_path = os.path.join(id_dir, f"id_{reg_no}.png")
                pdf_path = os.path.join(id_dir, f"id_{reg_no}.pdf")
                
                from html2image import Html2Image
                hti = _get_hti(output_path=id_dir)
                if hti is None:
                    # Fallback: create fresh instance if singleton failed
                    profile_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chrome_profile")
                    os.makedirs(profile_dir, exist_ok=True)
                    hti = Html2Image(
                        output_path=id_dir,
                        custom_flags=["--no-sandbox", "--disable-dev-shm-usage",
                                      "--disable-gpu", "--headless=new",
                                      "--hide-scrollbars", "--disable-extensions",
                                      f"--user-data-dir={profile_dir}"],
                        disable_logging=True
                    )
                else:
                    hti.output_path = id_dir
                hti.screenshot(html_str=full_document, save_as=f"id_{reg_no}.png", size=(340, 540))
                
                if os.path.exists(png_path):
                    from PIL import Image
                    img = Image.open(png_path).convert("RGB")
                    img.save(pdf_path, "PDF", resolution=100.0)
                    return pdf_path, png_path
            except Exception as html_err:
                print(f"HTML to Image rendering failed, falling back to PIL: {html_err}")
                
        # Draw background gradient from custom layout colors
        bg_start = list(designer_settings.get("bg_start", [11, 15, 25]))
        bg_end = list(designer_settings.get("bg_end", [26, 35, 51]))
        
        # Ensure we have RGB tuples
        if len(bg_start) == 3: bg_start = (bg_start[0], bg_start[1], bg_start[2], 255)
        elif len(bg_start) == 4: bg_start = tuple(bg_start)
        if len(bg_end) == 3: bg_end = (bg_end[0], bg_end[1], bg_end[2], 255)
        elif len(bg_end) == 4: bg_end = tuple(bg_end)
        
        for y_coord in range(height):
            ratio = y_coord / height
            r = int(bg_start[0] * (1 - ratio) + bg_end[0] * ratio)
            g = int(bg_start[1] * (1 - ratio) + bg_end[1] * ratio)
            b = int(bg_start[2] * (1 - ratio) + bg_end[2] * ratio)
            draw.line([(0, y_coord), (width, y_coord)], fill=(r, g, b, 255))
            
        # Draw custom shapes
        shapes = designer_settings.get("shapes", [])
        for s in shapes:
            stype = s.get("type")
            color = tuple(s.get("color", [255, 255, 255, 20]))
            box = s.get("box")
            coords = s.get("coords")
            
            if stype == "arc" and box:
                draw.arc(box, start=0, end=360, fill=color, width=s.get("width", 1))
            elif stype == "circle" and box:
                draw.ellipse(box, fill=color)
            elif stype == "line" and coords:
                line_coords = [(pt[0], pt[1]) for pt in coords]
                draw.line(line_coords, fill=color, width=s.get("width", 1))
            elif stype == "rectangle" and box:
                draw.rectangle(box, fill=color)
                
        # Draw custom fields
        fields = designer_settings.get("fields", [])
        for f in fields:
            text_tpl = f.get("text", "")
            x = f.get("x", 300)
            y = f.get("y", 210)
            font_size = f.get("font_size", 16)
            color_val = f.get("color", [255, 255, 255])
            align = f.get("align", "center")
            
            # Resolve text value
            resolved_text = text_tpl
            resolved_text = resolved_text.replace("{Name}", name)
            resolved_text = resolved_text.replace("{Registration Number}", reg_no)
            resolved_text = resolved_text.replace("{Email Address}", email)
            resolved_text = resolved_text.replace("{Phone Number}", phone)
            resolved_text = resolved_text.replace("{Unique ID}", uid)
            resolved_text = resolved_text.replace("{Pass Type}", level)
            resolved_text = resolved_text.replace("{Event}", event_name)
            
            # Draw text
            try:
                if f.get("font_name") == "cour.ttf":
                    font = ImageFont.truetype("cour.ttf", font_size)
                else:
                    font = ImageFont.truetype("arial.ttf", font_size)
            except IOError:
                font = ImageFont.load_default()
                
            anchor = "mm"
            if align == "left":
                anchor = "lm"
            elif align == "right":
                anchor = "rm"
                
            color = tuple(color_val)
            if len(color) == 3:
                color = (color[0], color[1], color[2], 255)
                
            draw.text((x, y), resolved_text, fill=color, font=font, anchor=anchor)
            
        # Draw QR Code
        qr_cfg = designer_settings.get("qr", {"x": 180, "y": 550, "size": 240})
        qr_x = qr_cfg.get("x", 180)
        qr_y = qr_cfg.get("y", 550)
        qr_size = qr_cfg.get("size", 240)
        
        if qr_path and os.path.exists(qr_path):
            qr_img = Image.open(qr_path).convert("RGBA")
            qr_img = qr_img.resize((qr_size, qr_size), Image.Resampling.LANCZOS)
            
            # Draw white background box behind QR code
            draw.rounded_rectangle([qr_x - 10, qr_y - 10, qr_x + qr_size + 10, qr_y + qr_size + 10], radius=15, fill=(255, 255, 255, 255))
            card.paste(qr_img, (qr_x, qr_y), qr_img)
            
        # Draw Outer Border
        border_color = tuple(designer_settings.get("border_color", [255, 255, 255, 20]))
        border_width = designer_settings.get("border_width", 2)
        if len(border_color) == 3: border_color = (border_color[0], border_color[1], border_color[2], 255)
        draw.rectangle([15, 15, width - 15, height - 15], outline=border_color, width=border_width)
        
        # Save custom layout card files
        id_dir = os.path.join(get_qr_dir(), "id_cards")
        os.makedirs(id_dir, exist_ok=True)
        
        png_path = os.path.join(id_dir, f"id_{reg_no}.png")
        pdf_path = os.path.join(id_dir, f"id_{reg_no}.pdf")
        
        card_rgb = card.convert("RGB")
        card_rgb.save(png_path, "PNG")
        card_rgb.save(pdf_path, "PDF", resolution=100.0)
        
        return pdf_path, png_path

# 2. Theme Selection Logic based on event_name or configuration override
    if cfg is None:
        cfg = get_event_config()
    configured_theme = cfg.get("id_card_theme", "auto")
    
    if configured_theme != "auto" and configured_theme in theme_configs:
        theme = configured_theme
    else:
        ev_norm = str(event_name).strip().lower()
        theme = "cyber_neon" # Default
        
        if "corporate" in ev_norm or "business" in ev_norm or "summit" in ev_norm or "conference" in ev_norm:
            theme = "classic_corporate"
        elif "gala" in ev_norm or "award" in ev_norm or "luxe" in ev_norm or "vip" in ev_norm:
            theme = "midnight_gold"
        elif "art" in ev_norm or "sunset" in ev_norm or "music" in ev_norm or "creative" in ev_norm:
            theme = "sunset_glow"
        elif "green" in ev_norm or "eco" in ev_norm or "sustain" in ev_norm or "nature" in ev_norm:
            theme = "emerald_luxe"
        elif "science" in ev_norm or "health" in ev_norm or "edu" in ev_norm or "university" in ev_norm or "academic" in ev_norm:
            theme = "royal_amethyst"
        elif "aurora" in ev_norm or "polar" in ev_norm or "northern" in ev_norm:
            theme = "aurora_borealis"
        elif "ice" in ev_norm or "snow" in ev_norm or "frost" in ev_norm or "glacier" in ev_norm:
            theme = "frozen_glacier"
        elif "cherry" in ev_norm or "blossom" in ev_norm or "sakura" in ev_norm or "spring" in ev_norm:
            theme = "cherry_blossom"
        elif "desert" in ev_norm or "sand" in ev_norm or "sahara" in ev_norm or "dune" in ev_norm:
            theme = "desert_sand"
        elif "ocean" in ev_norm or "sea" in ev_norm or "aqua" in ev_norm or "marine" in ev_norm or "beach" in ev_norm:
            theme = "ocean_breeze"
        elif "volcano" in ev_norm or "lava" in ev_norm or "fire" in ev_norm or "ash" in ev_norm:
            theme = "volcanic_ash"
        elif "lavender" in ev_norm or "mist" in ev_norm or "flower" in ev_norm or "petal" in ev_norm:
            theme = "lavender_mist"
        elif "minimal" in ev_norm or "sleek" in ev_norm or "silver" in ev_norm or "monochrome" in ev_norm:
            theme = "monochrome_sleek"
        elif "retro" in ev_norm or "sepia" in ev_norm or "vintage" in ev_norm or "classic" in ev_norm:
            theme = "vintage_sepia"
        elif "electric" in ev_norm or "lightning" in ev_norm or "volt" in ev_norm:
            theme = "electric_violet"
        elif "moss" in ev_norm or "wood" in ev_norm or "forest" in ev_norm or "timber" in ev_norm:
            theme = "forest_moss"
        elif "golden" in ev_norm or "sunset" in ev_norm or "hour" in ev_norm or "amber" in ev_norm:
            theme = "golden_hour"
        elif "berry" in ev_norm or "smoothie" in ev_norm or "fruit" in ev_norm:
            theme = "berry_smoothie"
        elif "space" in ev_norm or "nebula" in ev_norm or "galaxy" in ev_norm or "star" in ev_norm or "cosmic" in ev_norm:
            theme = "space_nebula"
        elif "mint" in ev_norm or "fresh" in ev_norm or "cool" in ev_norm:
            theme = "mint_fresh"
        elif "citrus" in ev_norm or "lemon" in ev_norm or "lime" in ev_norm or "orange" in ev_norm or "juice" in ev_norm:
            theme = "citrus_twist"
        elif "rose" in ev_norm or "pink" in ev_norm or "champagne" in ev_norm or "gold" in ev_norm:
            theme = "rose_gold"
        elif "carbon" in ev_norm or "fiber" in ev_norm or "racing" in ev_norm or "grid" in ev_norm:
            theme = "carbon_fiber"
        elif "sapphire" in ev_norm or "royal" in ev_norm or "blue" in ev_norm:
            theme = "royal_sapphire"
        elif "candy" in ev_norm or "sweet" in ev_norm or "pop" in ev_norm:
            theme = "candy_pop"
        else:
            # Fallback to deterministic selection using string hash
            theme_hash = sum(ord(char) for char in ev_norm)
            themes = list(theme_configs.keys())
            theme = themes[theme_hash % len(themes)]
            
    cfg_theme = theme_configs[theme]
    
    # 3. Attendee Level / Role Color Coding
    lvl_norm = str(level).strip().lower()
    role_label = str(level).strip().upper() if level else "ATTENDEE"
    
    # Color mapping for different roles/levels (RGB)
    role_colors = {
        "vip": (255, 215, 0),          # Gold
        "organizer": (231, 76, 60),     # Red
        "speaker": (155, 89, 182),      # Purple
        "sponsor": (52, 152, 219),      # Blue
        "volunteer": (46, 204, 113),    # Green
        "attendee": cfg_theme["default_accent"]
    }
    
    accent_rgb = cfg_theme["default_accent"]
    for r_key, r_color in role_colors.items():
        if r_key in lvl_norm:
            accent_rgb = r_color
            break
            
    # Draw Vertical Linear Gradient Background
    bg_start = cfg_theme["bg_start"]
    bg_end = cfg_theme["bg_end"]
    for y in range(height):
        ratio = y / height
        r = int(bg_start[0] * (1 - ratio) + bg_end[0] * ratio)
        g = int(bg_start[1] * (1 - ratio) + bg_end[1] * ratio)
        b = int(bg_start[2] * (1 - ratio) + bg_end[2] * ratio)
        draw.line([(0, y), (width, y)], fill=(r, g, b, 255))
        
    # Draw Theme-specific Shapes
    default_accent = cfg_theme["default_accent"]
    for s in cfg_theme["shapes"]:
        stype = s["type"]
        orig_color = s["color"]
        
        # Replace default accent colors with role-specific accent colors
        if len(orig_color) == 4:
            cr, cg, cb, ca = orig_color
            if (cr, cg, cb) == default_accent or (cr, cg, cb) == (0, 245, 160):
                color = (accent_rgb[0], accent_rgb[1], accent_rgb[2], ca)
            else:
                color = orig_color
        else:
            color = orig_color
            
        if stype == "arc":
            draw.arc(s["box"], start=0, end=360, fill=color, width=s.get("width", 1))
        elif stype == "circle":
            draw.ellipse(s["box"], fill=color)
        elif stype == "line":
            draw.line(s["coords"], fill=color, width=s.get("width", 1))
        elif stype == "rectangle":
            draw.rectangle(s["box"], fill=color)
            
    # Draw Borders
    draw.rectangle([15, 15, width - 15, height - 15], outline=(255, 255, 255, 15), width=2)
    if theme == "midnight_gold":
        draw.rectangle([25, 25, width - 25, height - 25], outline=cfg_theme["border"], width=1)
    else:
        draw.rectangle([25, 25, width - 25, height - 25], outline=(accent_rgb[0], accent_rgb[1], accent_rgb[2], 45), width=1)
        
    # 4. Load Fonts
    try:
        font_large = ImageFont.truetype("arial.ttf", 36)
        font_title = ImageFont.truetype("arial.ttf", 24)
        font_sub = ImageFont.truetype("arial.ttf", 16)
        font_mono = ImageFont.truetype("cour.ttf", 18)
    except IOError:
        font_large = ImageFont.load_default()
        font_title = ImageFont.load_default()
        font_sub = ImageFont.load_default()
        font_mono = ImageFont.load_default()
        
    # 5. Header / Event Details
    badge_text = f"{role_label} PASS"
    draw.rounded_rectangle([width//2 - 110, 45, width//2 + 110, 75], radius=15, 
                           fill=(accent_rgb[0], accent_rgb[1], accent_rgb[2], 25), 
                           outline=(accent_rgb[0], accent_rgb[1], accent_rgb[2], 100), width=1)
    draw.text((width//2, 60), badge_text, fill=(accent_rgb[0], accent_rgb[1], accent_rgb[2], 255), font=font_sub, anchor="mm")
    
    text_primary = cfg_theme["text_primary"]
    text_secondary = cfg_theme["text_secondary"]
    
    header_override = cfg.get("id_card_header", "").strip()
    header_text = header_override if header_override else event_name
    draw.text((width//2, 110), str(header_text).upper(), fill=text_primary, font=font_title, anchor="mm")
    draw.line([(100, 150), (width - 100, 150)], fill=(text_primary[0], text_primary[1], text_primary[2], 30), width=1)
    
    # 6. Attendee Details
    draw.text((width//2, 210), str(name), fill=text_primary, font=font_large, anchor="mm")
    
    details_y = 290
    spacing = 35
    
    details = []
    if cfg.get("id_card_show_reg", True):
        details.append((cfg.get("id_card_label_reg", "REGISTRATION:"), str(reg_no)))
    if cfg.get("id_card_show_email", True) and email:
        details.append((cfg.get("id_card_label_email", "EMAIL:"), str(email)))
    if cfg.get("id_card_show_phone", True) and phone:
        details.append((cfg.get("id_card_label_phone", "PHONE:"), str(phone)))
    if uid and cfg.get("id_card_show_uid", True):
        details.append((cfg.get("id_card_label_uid", "UNIQUE ID:"), str(uid)))
    if level and cfg.get("id_card_show_pass", True):
        details.append((cfg.get("id_card_label_pass", "PASS TYPE:"), str(level).upper()))
        
    for label, val in details:
        if val and val != "None" and val != "":
            draw.text((80, details_y), label, fill=(accent_rgb[0], accent_rgb[1], accent_rgb[2], 180), font=font_sub)
            draw.text((230, details_y), val, fill=text_secondary, font=font_sub)
            details_y += spacing
            
    # 7. Embed QR Code Image
    if qr_path and os.path.exists(qr_path):
        qr_img = Image.open(qr_path).convert("RGBA")
        qr_img = qr_img.resize((240, 240), Image.Resampling.LANCZOS)
        
        qr_x1 = width//2 - 130
        qr_y1 = height - 350
        qr_x2 = width//2 + 130
        qr_y2 = height - 90
        
        draw.rounded_rectangle([qr_x1, qr_y1, qr_x2, qr_y2], radius=15, fill=(255, 255, 255, 255))
        card.paste(qr_img, (width//2 - 120, height - 340), qr_img)
        
    footer_text = cfg.get("id_card_footer", "").strip()
    if footer_text:
        draw.text((width//2, height - 50), str(footer_text).upper(), fill=(text_secondary[0], text_secondary[1], text_secondary[2], 180), font=font_sub, anchor="mm")
        
    # 8. Save Files
    id_dir = os.path.join(get_qr_dir(), "id_cards")
    os.makedirs(id_dir, exist_ok=True)
    
    png_path = os.path.join(id_dir, f"id_{reg_no}.png")
    pdf_path = os.path.join(id_dir, f"id_{reg_no}.pdf")
    
    card_rgb = card.convert("RGB")
    card_rgb.save(png_path, "PNG")
    card_rgb.save(pdf_path, "PDF", resolution=100.0)
    
    return pdf_path, png_path


def _embed_qr_image(ws, r: int, qr_path: str, col_idx: int) -> None:
    if not os.path.exists(qr_path):
        return
    xl_img = XLImage(qr_path)
    xl_img.width = 100
    xl_img.height = 100
    cell_addr = f"{get_column_letter(col_idx)}{r}"
    ws.add_image(xl_img, cell_addr)
    ws.row_dimensions[r].height = 80
    ws.cell(row=r, column=col_idx, value="Embedded")


def _embed_barcode_image(ws, r: int, barcode_path: str, col_idx: int) -> None:
    if not os.path.exists(barcode_path):
        return
    xl_img = XLImage(barcode_path)
    xl_img.width = 150
    xl_img.height = 60
    cell_addr = f"{get_column_letter(col_idx)}{r}"
    ws.add_image(xl_img, cell_addr)
    ws.row_dimensions[r].height = 80
    ws.cell(row=r, column=col_idx, value="Embedded")


def _generate_unique_id(existing_ids: set[str], length: int = 8) -> str:
    charset = string.ascii_uppercase + string.digits
    while True:
        uid = "".join(random.choices(charset, k=length))
        if uid not in existing_ids:
            return uid


@app.route("/qrcodes/<path:filename>")
def serve_qrcode(filename):
    clean_filename = os.path.basename(filename)
    if ".." in filename or clean_filename != filename:
        return jsonify(message="Invalid path reference."), 400
    return send_from_directory(get_qr_dir(), clean_filename)


@app.route("/barcodes/<path:filename>")
def serve_barcode(filename):
    clean_filename = os.path.basename(filename)
    if ".." in filename or clean_filename != filename:
        return jsonify(message="Invalid path reference."), 400
    return send_from_directory(get_barcode_dir(), clean_filename)


@app.route("/add_attendee", methods=["POST"])
def add_attendee():
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
        
    payload = request.json or {}
    name = payload.get("name", "").strip()
    email = payload.get("email", "").strip()
    reg_no = payload.get("reg_no", "").strip()
    phone_raw = payload.get("phone", "").strip()
    phone = _normalize_phone(phone_raw)
    
    excel_file = get_excel_file()
    wb = None
    try:
        with lock:
            wb = load_workbook(excel_file)
            ws = wb.active
            hdrs, _ = _get_or_create_headers(ws)
            
            # Build list of existing registration numbers to check for uniqueness
            existing_regs = set()
            for r in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=hdrs["registration number"]).value
                if cell_val:
                    existing_regs.add(str(cell_val).strip().lower())
                    
            # Run integrity checks
            val_errors = _validate_attendee_integrity({
                "name": name,
                "email": email,
                "reg_no": reg_no,
                "phone": phone_raw
            }, existing_regs=existing_regs)
            
            if val_errors:
                wb.close()
                return jsonify(message="Validation failed.", errors=val_errors), 400
            
            existing_uids = set()
            for r in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=hdrs["unique id"]).value
                if cell_val:
                    existing_uids.add(str(cell_val).strip().upper())
                    
            uid = _generate_unique_id(existing_uids)
            cfg = get_event_config()
            qr_str = _get_qr_payload(reg_no, cfg)
            qr_path = _generate_qr_for_guest(qr_str, reg_no)
            barcode_path = _generate_barcode_for_guest(reg_no)
            
            new_r = ws.max_row + 1
            ws.cell(row=new_r, column=hdrs["name"], value=name)
            if "email address" in hdrs:
                ws.cell(row=new_r, column=hdrs["email address"], value=email)
            ws.cell(row=new_r, column=hdrs["registration number"], value=reg_no)
            if "phone number" in hdrs:
                ws.cell(row=new_r, column=hdrs["phone number"], value=phone)
            ws.cell(row=new_r, column=hdrs["unique id"], value=uid)
            ws.cell(row=new_r, column=hdrs["qr"], value=qr_str)
            ws.cell(row=new_r, column=hdrs["barcode"], value=reg_no)
            ws.cell(row=new_r, column=hdrs[SCAN_COL_NAME.lower()], value="")
            if "email sent status" in hdrs:
                ws.cell(row=new_r, column=hdrs["email sent status"], value="Not Sent")
            if "whatsapp sent status" in hdrs:
                ws.cell(row=new_r, column=hdrs["whatsapp sent status"], value="Not Sent")
            if "sms sent status" in hdrs:
                ws.cell(row=new_r, column=hdrs["sms sent status"], value="Not Sent")
            
            # Custom fields dynamic storage
            custom_payload = payload.get("custom_fields", {})
            for key, val in custom_payload.items():
                key_clean = key.lower().strip()
                if key_clean in hdrs:
                    ws.cell(row=new_r, column=hdrs[key_clean], value=str(val).strip())
            
            _embed_qr_image(ws, new_r, qr_path, hdrs["qr code image"])
            _embed_barcode_image(ws, new_r, barcode_path, hdrs["barcode image"])
            
            # Generate ID Card if enabled
            cfg = get_event_config()
            if cfg.get("enable_id_card_generation"):
                level = ""
                group_col = cfg.get("group_column", "")
                if group_col:
                    for r_k, r_v in custom_payload.items():
                        if r_k.strip().lower() == group_col.strip().lower():
                            level = str(r_v or "").strip()
                            break
                if not level:
                    for r_k, r_v in custom_payload.items():
                        if "level" in r_k.lower() or "subgroup" in r_k.lower():
                            level = str(r_v or "").strip()
                            break
                try:
                    _generate_id_card(
                        name=name,
                        reg_no=reg_no,
                        phone=phone,
                        email=email,
                        uid=uid,
                        qr_path=qr_path,
                        event_name=cfg.get("event_name_template", "the Event"),
                        level=level
                    )
                except Exception as e:
                    print(f"Error generating ID card on add: {e}")
            
            _atomic_save(wb, excel_file)
            wb = None
            
        threading.Thread(target=_rebuild_highlighted, daemon=True).start()
        
        # Queue registration notifications if any are enabled
        if cfg.get("reg_notify_email") or cfg.get("reg_notify_whatsapp") or cfg.get("reg_notify_sms"):
            registration_notifications_queue.put((reg_no, cfg, request.host_url))
            
        socketio.emit("registry_updated", {})
        _emit_stats()
        
        return jsonify(message=f"Successfully registered {name} ({reg_no})"), 200
        
    except PermissionError:
        return jsonify(message="Excel file is locked. Please close it in Excel first."), 409
    except Exception as e:
        return jsonify(message=f"Error adding attendee: {str(e)}"), 500
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass


def _perform_bulk_import(rows, mode, auto_resolve) -> tuple[dict, int]:
    excel_file = get_excel_file()
    log_file = get_log_file()
    wb = None
    try:
        with lock:
            if mode == "reset":
                # Create a fresh Excel file with headers
                wb = load_workbook(excel_file)
                ws = wb.active
                ws.delete_rows(1, ws.max_row + 1)
                
                # We start with the core headers
                headers = ["Name", "Registration Number", "Unique ID", "QR", "Barcode", SCAN_COL_NAME]
                
                has_email = False
                has_phone = False
                if rows:
                    for r_item in rows:
                        for key in r_item.keys():
                            key_clean = key.lower().strip()
                            if "email" in key_clean:
                                has_email = True
                            elif "phone" in key_clean or "whatsapp" in key_clean or "contact" in key_clean:
                                has_phone = True
                
                if has_email:
                    headers.extend(["Email Address", "Email Sent Status"])
                if has_phone:
                    headers.extend(["Phone Number", "WhatsApp Sent Status", "SMS Sent Status"])
                    
                headers.extend(["Scan Timestamps", "Scan Devices", "QR Code Image", "Barcode Image"])
                
                # Dynamic custom columns from rows
                custom_cols = []
                system_names_low = {
                    "name", "registration number", "unique id", "qr", "barcode",
                    "scan timestamps", "scan devices", "qr code image", "barcode image",
                    "email address", "email sent status", "phone number", "whatsapp sent status",
                    "email", "phone", "whatsapp", "contact", SCAN_COL_NAME.lower()
                }
                if rows:
                    for key in rows[0].keys():
                        if key.lower().strip() not in system_names_low:
                            custom_cols.append(key.strip())
                headers.extend(custom_cols)
                
                for c_idx, h_val in enumerate(headers, 1):
                    ws.cell(row=1, column=c_idx, value=h_val)
                    
                hdrs = {h.lower(): i for i, h in enumerate(headers, 1)}
                
                global scanned_log
                scanned_log = pd.DataFrame(columns=["QR Data", "Scan Count", "Timestamps", "Devices"])
                scanned_log.to_csv(log_file, index=False)
                
                with _devices_lock:
                    for d_val in connected_devices.values():
                        d_val["scans"] = 0
                        d_val["last_activity"] = "Reset Registry"
                _broadcast_devices()
                _emit_stats()
            else:
                wb = load_workbook(excel_file)
                ws = wb.active
                hdrs, _ = _get_or_create_headers(ws)
                
                # Append any new custom columns
                if rows:
                    for key in rows[0].keys():
                        key_clean = key.lower().strip()
                        if key_clean not in ["name", "email", "reg_no", "phone"] and key_clean not in hdrs:
                            system_names_low = {
                                "name", "registration number", "unique id", "qr", "barcode",
                                "scan timestamps", "scan devices", "qr code image", "barcode image",
                                "email address", "email sent status", "phone number", "whatsapp sent status",
                                "email", "phone", "whatsapp", "contact", SCAN_COL_NAME.lower()
                            }
                            if key_clean not in system_names_low:
                                col = ws.max_column + 1
                                ws.cell(row=1, column=col, value=key.strip())
                                hdrs[key_clean] = col
                
            # Read existing reg numbers & unique IDs to avoid duplicates
            existing_regs = set()
            existing_emails = set()
            existing_uids = set()
            for r in range(2, ws.max_row + 1):
                reg_val = ws.cell(row=r, column=hdrs["registration number"]).value
                email_val = ws.cell(row=r, column=hdrs["email address"]).value if "email address" in hdrs else None
                uid_val = ws.cell(row=r, column=hdrs["unique id"]).value
                if reg_val:
                    existing_regs.add(str(reg_val).strip().lower())
                if email_val:
                    existing_emails.add(str(email_val).strip().lower())
                if uid_val:
                    existing_uids.add(str(uid_val).strip().upper())
                    
            added_cnt = 0
            skipped_cnt = 0
            new_regs = []
            
            for row in rows:
                name = ""
                email = ""
                reg_no = ""
                phone = ""
                
                for k, v in row.items():
                    k_low = k.lower().strip()
                    val_str = _clean_val(v)
                    if "name" in k_low:
                        name = val_str
                    elif "email" in k_low:
                        email = val_str
                    elif "reg" in k_low or "number" in k_low or "id" in k_low:
                        if "phone" not in k_low and "whatsapp" not in k_low:
                            reg_no = val_str
                    elif "phone" in k_low or "mobile" in k_low or "whatsapp" in k_low or "contact" in k_low:
                        phone = _normalize_phone(v)
                
                # Validation check
                if not name or not reg_no:
                    skipped_cnt += 1
                    continue
                    
                # Skip duplicate registration numbers in append mode, 
                # unless auto-resolve is enabled
                if reg_no.lower() in existing_regs:
                    if auto_resolve:
                        # Generate a non-colliding reg number
                        base_reg = reg_no
                        counter = 1
                        while f"{base_reg}-{counter}".lower() in existing_regs:
                            counter += 1
                        reg_no = f"{base_reg}-{counter}"
                    else:
                        skipped_cnt += 1
                        continue
                        
                if email and email.lower() in existing_emails:
                    if not auto_resolve:
                        skipped_cnt += 1
                        continue
                        
                uid = _generate_unique_id(existing_uids)
                existing_regs.add(reg_no.lower())
                if email:
                    existing_emails.add(email.lower())
                existing_uids.add(uid)
                
                cfg = get_event_config()
                qr_str = _get_qr_payload(reg_no, cfg)
                qr_path = _generate_qr_for_guest(qr_str, reg_no)
                barcode_path = _generate_barcode_for_guest(reg_no)
                
                new_r = ws.max_row + 1
                ws.cell(row=new_r, column=hdrs["name"], value=name)
                ws.cell(row=new_r, column=hdrs["registration number"], value=reg_no)
                if "email address" in hdrs:
                    ws.cell(row=new_r, column=hdrs["email address"], value=email)
                if "phone number" in hdrs:
                    ws.cell(row=new_r, column=hdrs["phone number"], value=phone)
                ws.cell(row=new_r, column=hdrs["unique id"], value=uid)
                ws.cell(row=new_r, column=hdrs["qr"], value=qr_str)
                ws.cell(row=new_r, column=hdrs["barcode"], value=reg_no)
                ws.cell(row=new_r, column=hdrs[SCAN_COL_NAME.lower()], value="")
                if "email sent status" in hdrs:
                    ws.cell(row=new_r, column=hdrs["email sent status"], value="Not Sent")
                if "whatsapp sent status" in hdrs:
                    ws.cell(row=new_r, column=hdrs["whatsapp sent status"], value="Not Sent")
                if "sms sent status" in hdrs:
                    ws.cell(row=new_r, column=hdrs["sms sent status"], value="Not Sent")
                
                # Fill custom columns dynamically
                for key, val in row.items():
                    key_clean = key.lower().strip()
                    excluded = ["name", "email", "reg_no", "phone", "unique id", "qr", "barcode",
                                SCAN_COL_NAME.lower(), "email sent status", "whatsapp sent status", "sms sent status"]
                    if key_clean in hdrs and key_clean not in excluded:
                        ws.cell(row=new_r, column=hdrs[key_clean], value=str(val).strip())
                
                _embed_qr_image(ws, new_r, qr_path, hdrs["qr code image"])
                _embed_barcode_image(ws, new_r, barcode_path, hdrs["barcode image"])
                
                # Generate ID Card if enabled
                cfg = get_event_config()
                if cfg.get("enable_id_card_generation"):
                    level = ""
                    group_col = cfg.get("group_column", "")
                    if group_col:
                        for r_k, r_v in row.items():
                            if r_k.strip().lower() == group_col.strip().lower():
                                level = str(r_v or "").strip()
                                break
                    if not level:
                        for r_k, r_v in row.items():
                            if "level" in r_k.lower() or "subgroup" in r_k.lower():
                                level = str(r_v or "").strip()
                                break
                    try:
                        _generate_id_card(
                            name=name,
                            reg_no=reg_no,
                            phone=phone,
                            email=email,
                            uid=uid,
                            qr_path=qr_path,
                            event_name=cfg.get("event_name_template", "the Event"),
                            level=level
                        )
                    except Exception as e:
                        print(f"Error generating ID card on import: {e}")
                        
                added_cnt += 1
                new_regs.append(reg_no)
                
            _atomic_save(wb, excel_file)
            wb = None
            
        threading.Thread(target=_rebuild_highlighted, daemon=True).start()
        
        # Queue registration notifications if any are enabled
        cfg = get_event_config()
        if cfg.get("reg_notify_email") or cfg.get("reg_notify_whatsapp") or cfg.get("reg_notify_sms"):
            try:
                host_url = request.host_url
                for r_no in new_regs:
                    registration_notifications_queue.put((r_no, cfg, host_url))
            except Exception as ex:
                print(f"[import] Error queuing auto notifications: {ex}")
                
        socketio.emit("registry_updated", {})
        _emit_stats()
        
        return {
            "message": f"Import completed successfully. Added {added_cnt} attendees, skipped {skipped_cnt}.",
            "added": added_cnt,
            "skipped": skipped_cnt
        }, 200
        
    except PermissionError:
        return {"message": "Excel file is locked. Please close it in Excel first."}, 409
    except Exception as e:
        return {"message": f"Error importing attendees: {str(e)}"}, 500
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass


@app.route("/import_csv", methods=["POST"])
def import_csv():
    if "file" not in request.files:
        return jsonify(message="No file uploaded."), 400
    file = request.files["file"]
    mode = request.form.get("mode", "append")
    
    if not file.filename.endswith(".csv"):
        return jsonify(message="Uploaded file must be a CSV."), 400
        
    try:
        stream = io.StringIO(file.stream.read().decode("utf-8", errors="replace"), newline=None)
        reader = csv.DictReader(stream)
        reader.fieldnames = [f.strip() for f in reader.fieldnames]
        
        required_cols = {"Name", "Email Address", "Registration Number"}
        missing_cols = required_cols - set(reader.fieldnames)
        if missing_cols:
            return jsonify(message=f"CSV is missing columns: {missing_cols}"), 400
            
        rows_to_process = []
        for row in reader:
            rows_to_process.append({
                "name": row.get("Name", "").strip(),
                "email": row.get("Email Address", "").strip(),
                "reg_no": row.get("Registration Number", "").strip(),
                "phone": row.get("Phone Number", "").strip() if "Phone Number" in row else ""
            })
            
    except Exception as e:
        return jsonify(message=f"Error parsing CSV: {str(e)}"), 400
        
    res, status_code = _perform_bulk_import(rows_to_process, mode, False)
    return jsonify(res), status_code


@app.route("/preview_import", methods=["POST"])
def preview_import():
    if "file" not in request.files:
        return jsonify(message="No file uploaded."), 400
        
    file = request.files["file"]
    filename = file.filename.lower()
    
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        return jsonify(message="Uploaded file must be a CSV or Excel (.xlsx) file."), 400
        
    try:
        if filename.endswith(".csv"):
            content = file.stream.read()
            stream = io.StringIO(content.decode("utf-8", errors="replace"), newline=None)
            df = pd.read_csv(stream)
        else:
            df = pd.read_excel(file, engine="openpyxl")
            
        df.columns = df.columns.astype(str).str.strip()
    except Exception as e:
        return jsonify(message=f"Failed to read file: {str(e)}"), 400
        
    mapping = {}
    for col in df.columns:
        c_low = col.lower().strip()
        if "name" in c_low:
            mapping["name"] = col
        elif "email" in c_low:
            mapping["email"] = col
        elif "reg" in c_low or "number" in c_low or "id" in c_low:
            if "phone" not in c_low and "whatsapp" not in c_low:
                mapping["reg_no"] = col
        elif "phone" in c_low or "whatsapp" in c_low or "contact" in c_low:
            mapping["phone"] = col
            
    if "name" not in mapping or "reg_no" not in mapping:
        return jsonify(message="File must contain columns for Name and Registration Number."), 400
        
    db_regs = set()
    db_emails = set()
    try:
        with lock:
            df_db = pd.read_excel(get_excel_file(), engine="openpyxl")
        df_db.columns = df_db.columns.str.strip()
        db_col_map = {c.lower(): c for c in df_db.columns}
        
        reg_c = db_col_map.get("registration number")
        email_c = db_col_map.get("email address")
        
        if reg_c:
            for val in df_db[reg_c]:
                if pd.notna(val):
                    db_regs.add(str(val).strip().lower())
        if email_c:
            for val in df_db[email_c]:
                if pd.notna(val):
                    db_emails.add(str(val).strip().lower())
    except Exception as e:
        print(f"[preview] Error loading db: {e}")
        
    file_regs = set()
    file_emails = set()
    
    preview_records = []
    
    for idx, row in df.iterrows():
        name = _clean_val(row.get(mapping["name"]))
        email = _clean_val(row.get(mapping["email"])) if "email" in mapping else ""
        reg_no = _clean_val(row.get(mapping["reg_no"]))
        phone = _clean_val(row.get(mapping.get("phone"))) if "phone" in mapping else ""
        
        if not name and not reg_no:
            continue
            
        # Run integrity validation (ignore duplicate reg number in db_regs since we check duplicate status below)
        val_errors = _validate_attendee_integrity({
            "name": name,
            "email": email,
            "reg_no": reg_no,
            "phone": phone
        }, existing_regs=None)
        
        status = "ok"
        issue = ""
        
        if val_errors:
            status = "invalid"
            issue = "; ".join(val_errors.values())
        
        if not name or not reg_no:
            status = "missing"
            issue = "Missing critical fields"
        else:
            reg_low = reg_no.lower()
            email_low = email.lower() if email else ""
            
            if reg_low in db_regs:
                status = "dup_db_reg"
                issue = f"Reg no '{reg_no}' already in database"
            elif email_low and email_low in db_emails:
                status = "dup_db_email"
                issue = f"Email '{email}' already in database"
            elif reg_low in file_regs:
                status = "dup_file_reg"
                issue = f"Duplicate Reg no '{reg_no}' inside file"
            elif email_low and email_low in file_emails:
                status = "dup_file_email"
                issue = f"Duplicate Email '{email}' inside file"
                
            file_regs.add(reg_low)
            if email_low:
                file_emails.add(email_low)
            
        preview_records.append({
            "name": name,
            "email": email,
            "reg_no": reg_no,
            "phone": phone,
            "status": status,
            "issue": issue,
            "errors": val_errors
        })
        
    return jsonify(preview_records)


@app.route("/confirm_import", methods=["POST"])
def confirm_import():
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
        
    payload = request.json or {}
    rows = payload.get("rows", [])
    mode = payload.get("mode", "append")
    auto_resolve = payload.get("auto_resolve", False)
    
    if not rows:
        return jsonify(message="No data rows to import."), 400
        
    res, status_code = _perform_bulk_import(rows, mode, auto_resolve)
    return jsonify(res), status_code


# ── Template Formatter Helper ──────────────────────────────────────────────────
def format_template(template: str, row_dict: dict, extra: dict = None) -> str:
    result = template
    # Merge row_dict and extra
    data = {}
    for k, v in row_dict.items():
        data[str(k).strip()] = str(v or "").strip()
    if extra:
        for k, v in extra.items():
            data[str(k).strip()] = str(v or "").strip()
            
    for k, v in data.items():
        placeholder = "{" + str(k) + "}"
        result = result.replace(placeholder, str(v))
    return result


def _update_sent_status(reg_no: str, channel: str, status: str):
    excel_file = get_excel_file()
    wb = None
    try:
        with lock:
            wb = load_workbook(excel_file)
            ws = wb.active
            # Build header map inside the function (fixes NameError from undefined 'hdrs')
            hdrs = _get_headers(ws)
            if channel == "email":
                col_name = "email sent status"
            elif channel == "whatsapp":
                col_name = "whatsapp sent status"
            elif channel == "sms":
                col_name = "sms sent status"
            else:
                col_name = None
                
            reg_col_idx = hdrs.get("registration number")
            col_idx = hdrs.get(col_name) if col_name else None
            if col_idx and reg_col_idx:
                for r in range(2, ws.max_row + 1):
                    reg_val = ws.cell(row=r, column=reg_col_idx).value
                    if reg_val and str(reg_val).strip().lower() == reg_no.lower():
                        ws.cell(row=r, column=col_idx, value=status)
                        break
            _atomic_save(wb, excel_file)
            wb = None
    
        socketio.emit("registry_updated", {})
        threading.Thread(target=_rebuild_highlighted, daemon=True).start()
    except Exception as e:
        print(f"[status-update] Error: {e}")
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass


# ── Twilio WhatsApp Campaign Tracker ──────────────────────────────────────────
_whatsapp_sending_lock = threading.Lock()
_whatsapp_sending_active = False
_whatsapp_progress = {
    "sent": 0,
    "skipped": 0,
    "failed": 0,
    "total": 0,
    "current_phone": "",
    "current_status": "Idle",
    "logs": []
}

def _run_whatsapp_campaign(wa_provider, target_column, target_subgroup, host_url, event_name):
    global _whatsapp_sending_active, _whatsapp_progress
    import re
    import urllib.parse
    
    excel_file = get_excel_file()
    cfg = get_event_config()
    
    attendees = []
    try:
        with lock:
            df_xl = pd.read_excel(excel_file, engine="openpyxl")
            
        df_xl.columns = df_xl.columns.str.strip()
        col_map = {c.lower(): c for c in df_xl.columns}
        phone_col = col_map.get("phone number")
        
        for _, row in df_xl.iterrows():
            name_v = _clean_val(row.get(col_map.get("name")))
            phone_v = _clean_val(row.get(phone_col)) if phone_col else ""
            reg_v = _clean_val(row.get(col_map.get("registration number")))
            
            if name_v and phone_v and reg_v:
                row_data = {c_orig: _clean_val(row.get(c_orig)) for c_low, c_orig in col_map.items()}
                
                # Check subgroup match
                if target_column and target_subgroup and target_subgroup.lower() != "all":
                    col_val = row_data.get(target_column)
                    if not col_val or str(col_val).strip().lower() != target_subgroup.lower().strip():
                        continue
                        
                attendees.append({
                    "name": name_v,
                    "phone": phone_v,
                    "reg_no": reg_v,
                    "row_dict": row_data
                })
    except Exception as e:
        with _whatsapp_sending_lock:
            _whatsapp_progress["current_status"] = "Failed to load attendees"
            _whatsapp_progress["logs"].append(f"❌ Error loading attendees: {str(e)}")
            _whatsapp_sending_active = False
        socketio.emit("whatsapp_progress_update", _whatsapp_progress)
        return
        
    total_count = len(attendees)
    with _whatsapp_sending_lock:
        _whatsapp_progress["total"] = total_count
        _whatsapp_progress["sent"] = 0
        _whatsapp_progress["skipped"] = 0
        _whatsapp_progress["failed"] = 0
        _whatsapp_progress["current_status"] = "Starting WhatsApp campaign..."
        _whatsapp_progress["logs"] = [f"🚀 Started WhatsApp campaign ({wa_provider}). Total recipients: {total_count}"]
    socketio.emit("whatsapp_progress_update", _whatsapp_progress)
    
    for idx, att in enumerate(attendees):
        name = att["name"]
        phone = att["phone"]
        reg = att["reg_no"]
        row_dict = att["row_dict"]
        
        with _whatsapp_sending_lock:
            _whatsapp_progress["current_phone"] = phone
            _whatsapp_progress["current_status"] = f"Processing {phone} ({idx+1}/{total_count})"
        socketio.emit("whatsapp_progress_update", _whatsapp_progress)
        
        success, msg, extra_data = _send_single_whatsapp(phone, name, reg, row_dict, cfg, host_url)
        
        if success:
            if extra_data.get("method") == "manual":
                wa_link = f"https://wa.me/{extra_data['phone'].replace('+', '')}?text={urllib.parse.quote(msg)}"
                log_msg = f"🔗 Manual link for {name} ({phone}): <a href='{wa_link}' target='_blank' style='color:var(--accent); text-decoration:underline;'>Open WhatsApp Chat</a>"
                _update_sent_status(reg, "whatsapp", "Pending Click")
                with _whatsapp_sending_lock:
                    _whatsapp_progress["sent"] += 1
                    _whatsapp_progress["logs"].append(log_msg)
            else:
                log_msg = f"✅ Sent WhatsApp to {name} ({phone})"
                _update_sent_status(reg, "whatsapp", "Sent")
                with _whatsapp_sending_lock:
                    _whatsapp_progress["sent"] += 1
                    _whatsapp_progress["logs"].append(log_msg)
        else:
            log_msg = f"❌ Failed to send to {name} ({phone}): {msg}"
            _update_sent_status(reg, "whatsapp", "Failed")
            with _whatsapp_sending_lock:
                _whatsapp_progress["failed"] += 1
                _whatsapp_progress["logs"].append(log_msg)
                
        socketio.emit("whatsapp_progress_update", _whatsapp_progress)
        time.sleep(1.2)
        
    with _whatsapp_sending_lock:
        _whatsapp_progress["current_status"] = "Finished"
        _whatsapp_progress["current_phone"] = ""
        _whatsapp_progress["logs"].append(f"🏁 WhatsApp Campaign finished. Sent: {_whatsapp_progress['sent']}, Skipped: {_whatsapp_progress['skipped']}, Failed: {_whatsapp_progress['failed']}.")
        _whatsapp_sending_active = False
    socketio.emit("whatsapp_progress_update", _whatsapp_progress)


@app.route("/send_whatsapp_bulk", methods=["POST"])
def send_whatsapp_bulk():
    global _whatsapp_sending_active
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
        
    payload = request.json or {}
    wa_provider = payload.get("wa_provider", "manual").strip()
    
    cfg = get_event_config()
    cfg["wa_provider"] = wa_provider
    cfg["event_name_template"] = payload.get("event_name", "the Event").strip()
    
    if wa_provider == "twilio":
        cfg["twilio_sid"] = payload.get("twilio_sid", "").strip()
        cfg["twilio_token"] = payload.get("twilio_token", "").strip()
        cfg["twilio_sender"] = payload.get("twilio_sender", "").strip()
        if not cfg["twilio_sid"] or not cfg["twilio_token"] or not cfg["twilio_sender"]:
            return jsonify(message="Twilio SID, Token, and Sender number are required."), 400
    elif wa_provider == "meta":
        cfg["meta_access_token"] = payload.get("meta_access_token", "").strip()
        cfg["meta_phone_number_id"] = payload.get("meta_phone_number_id", "").strip()
        cfg["meta_template_name"] = payload.get("meta_template_name", "").strip()
        cfg["meta_lang_code"] = payload.get("meta_lang_code", "en_US").strip()
        if not cfg["meta_access_token"] or not cfg["meta_phone_number_id"] or not cfg["meta_template_name"]:
            return jsonify(message="Meta Access Token, Phone ID, and Template Name are required."), 400

    target_column = payload.get("target_column", "").strip()
    target_subgroup = payload.get("target_subgroup", "").strip()
    if target_column:
        cfg["group_column"] = target_column
        
    save_event_config(cfg)
    
    with _whatsapp_sending_lock:
        if _whatsapp_sending_active:
            return jsonify(message="A WhatsApp campaign is already in progress."), 409
        _whatsapp_sending_active = True
        
    host_url = request.host_url
    event_name = cfg["event_name_template"]
    
    threading.Thread(
        target=_run_whatsapp_campaign,
        args=(wa_provider, target_column, target_subgroup, host_url, event_name),
        daemon=True
    ).start()
    
    return jsonify(message="WhatsApp campaign started successfully."), 202


@app.route("/whatsapp_status")
def get_whatsapp_status():
    with _whatsapp_sending_lock:
        return jsonify(
            active=_whatsapp_sending_active,
            sent=_whatsapp_progress["sent"],
            skipped=_whatsapp_progress["skipped"],
            failed=_whatsapp_progress["failed"],
            total=_whatsapp_progress["total"],
            current_phone=_whatsapp_progress["current_phone"],
            current_status=_whatsapp_progress["current_status"],
            logs=_whatsapp_progress["logs"]
        )


# ── Email Campaign Tracker ────────────────────────────────────────────────────
_email_sending_lock = threading.Lock()
_email_sending_active = False
_email_progress = {
    "sent": 0,
    "skipped": 0,
    "failed": 0,
    "total": 0,
    "current_email": "",
    "current_status": "Idle",
    "logs": []
}

def _run_email_campaign(sender_email, app_password, subject, event_name, target_column=None, target_subgroup=None):
    global _email_sending_active, _email_progress
    import yagmail
    
    excel_file = get_excel_file()
    cfg = get_event_config()
    email_template = cfg.get("email_template", "")
    
    attendees = []
    try:
        with lock:
            df_xl = pd.read_excel(excel_file, engine="openpyxl")
            
        df_xl.columns = df_xl.columns.str.strip()
        col_map = {c.lower(): c for c in df_xl.columns}
        
        for _, row in df_xl.iterrows():
            name_v = _clean_val(row.get(col_map.get("name")))
            email_v = _clean_val(row.get(col_map.get("email address")))
            reg_v = _clean_val(row.get(col_map.get("registration number")))
            
            if name_v and email_v and reg_v:
                row_data = {c_orig: _clean_val(row.get(c_orig)) for c_low, c_orig in col_map.items()}
                
                # Check subgroup match
                if target_column and target_subgroup and target_subgroup.lower() != "all":
                    col_val = row_data.get(target_column)
                    if not col_val or str(col_val).strip().lower() != target_subgroup.lower().strip():
                        continue
                        
                attendees.append({
                    "name": name_v,
                    "email": email_v,
                    "reg_no": reg_v,
                    "row_dict": row_data
                })
    except Exception as e:
        with _email_sending_lock:
            _email_progress["current_status"] = "Failed to load attendees"
            _email_progress["logs"].append(f"❌ Error loading attendees: {str(e)}")
            _email_sending_active = False
        socketio.emit("email_progress_update", _email_progress)
        return
        
    total_count = len(attendees)
    with _email_sending_lock:
        _email_progress["total"] = total_count
        _email_progress["sent"] = 0
        _email_progress["skipped"] = 0
        _email_progress["failed"] = 0
        _email_progress["current_status"] = "Starting campaign..."
        _email_progress["logs"] = [f"🚀 Started email campaign. Total recipients: {total_count}"]
    socketio.emit("email_progress_update", _email_progress)
    
    yag = None
    try:
        yag = yagmail.SMTP(user=sender_email, password=app_password)
    except Exception as e:
        with _email_sending_lock:
            _email_progress["current_status"] = "SMTP Authentication failed"
            _email_progress["logs"].append(f"❌ SMTP Connection Error: {str(e)}")
            _email_sending_active = False
        socketio.emit("email_progress_update", _email_progress)
        return
        
    qr_dir = get_qr_dir()
    
    for idx, att in enumerate(attendees):
        name = att["name"]
        recipient = att["email"]
        reg = att["reg_no"]
        row_dict = att["row_dict"]
        qr_file = os.path.join(qr_dir, f"{reg}.png")
        
        with _email_sending_lock:
            _email_progress["current_email"] = recipient
            _email_progress["current_status"] = f"Sending to {recipient} ({idx+1}/{total_count})"
        socketio.emit("email_progress_update", _email_progress)
        
        if not os.path.exists(qr_file):
            msg = f"⚠️ Skipped {name} ({recipient}) — QR image missing"
            _update_sent_status(reg, "email", "Skipped")
            with _email_sending_lock:
                _email_progress["skipped"] += 1
                _email_progress["logs"].append(msg)
            socketio.emit("email_progress_update", _email_progress)
            continue
            
        from_cfg_email_subject = cfg.get("email_subject", "Your Event QR Code")
        from_cfg_email_template = cfg.get("email_template", "")
        
        resolved_subject = _get_attendee_template(cfg, row_dict, "email_subject", from_cfg_email_subject)
        resolved_body_template = _get_attendee_template(cfg, row_dict, "email_template", from_cfg_email_template)
        
        extra = {"Event": event_name}
        body = format_template(resolved_body_template, row_dict, extra)
        subject_formatted = format_template(resolved_subject, row_dict, extra)
        
        attachments_to_send = _resolve_email_attachments(reg, row_dict, cfg)
        success = False
        for attempt in range(1, 4):
            try:
                yag.send(
                    to=recipient,
                    subject=subject_formatted,
                    contents=body,
                    attachments=attachments_to_send,
                )
                success = True
                break
            except Exception as e:
                wait = 2 ** attempt
                with _email_sending_lock:
                    _email_progress["logs"].append(f"   Attempt {attempt} failed for {recipient}: {str(e)}. Retrying in {wait}s...")
                socketio.emit("email_progress_update", _email_progress)
                time.sleep(wait)
                
        if success:
            msg = f"✅ Sent to {name} ({recipient})"
            _update_sent_status(reg, "email", "Sent")
            with _email_sending_lock:
                _email_progress["sent"] += 1
                _email_progress["logs"].append(msg)
        else:
            msg = f"❌ Failed to send to {name} ({recipient})"
            _update_sent_status(reg, "email", "Failed")
            with _email_sending_lock:
                _email_progress["failed"] += 1
                _email_progress["logs"].append(msg)
                
        socketio.emit("email_progress_update", _email_progress)
        time.sleep(1.0)
        
    try:
        yag.close()
    except Exception:
        pass
        
    with _email_sending_lock:
        _email_progress["current_status"] = "Finished"
        _email_progress["current_email"] = ""
        _email_progress["logs"].append(f"🏁 Campaign finished. Sent: {_email_progress['sent']}, Skipped: {_email_progress['skipped']}, Failed: {_email_progress['failed']}.")
        _email_sending_active = False
    socketio.emit("email_progress_update", _email_progress)


@app.route("/send_emails", methods=["POST"])
def send_emails():
    global _email_sending_active
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415
        
    payload = request.json or {}
    sender = payload.get("sender_email", "").strip()
    password = payload.get("app_password", "").strip()
    subject = payload.get("subject", "Your Event QR Code").strip()
    event = payload.get("event_name", "the Event").strip()
    
    if not sender or not password:
        return jsonify(message="Sender email and App password are required."), 400
        
    with _email_sending_lock:
        if _email_sending_active:
            return jsonify(message="An email campaign is already in progress."), 409
        _email_sending_active = True
        
    # Save configuration parameters
    cfg = get_event_config()
    cfg["email_sender"] = sender
    cfg["email_password"] = password
    cfg["email_subject"] = subject
    cfg["event_name_template"] = event
    
    target_column = payload.get("target_column", "").strip()
    target_subgroup = payload.get("target_subgroup", "").strip()
    if target_column:
        cfg["group_column"] = target_column
        
    save_event_config(cfg)
    
    threading.Thread(
        target=_run_email_campaign,
        args=(sender, password, subject, event, target_column, target_subgroup),
        daemon=True
    ).start()
    
    return jsonify(message="Email campaign started successfully."), 202


@app.route("/email_status")
def get_email_status():
    with _email_sending_lock:
        return jsonify(
            active=_email_sending_active,
            sent=_email_progress["sent"],
            skipped=_email_progress["skipped"],
            failed=_email_progress["failed"],
            total=_email_progress["total"],
            current_email=_email_progress["current_email"],
            current_status=_email_progress["current_status"],
            logs=_email_progress["logs"]
        )


# ── SMS Bulk Campaign ─────────────────────────────────────────────────────────
_sms_sending_lock = threading.Lock()
_sms_sending_active = False
_sms_progress = {
    "sent": 0,
    "skipped": 0,
    "failed": 0,
    "total": 0,
    "current_phone": "",
    "current_status": "Idle",
    "logs": []
}


def _run_sms_campaign(sms_provider, event_name, host_url, target_column=None, target_subgroup=None):
    global _sms_sending_active, _sms_progress

    excel_file = get_excel_file()
    cfg = get_event_config()

    attendees = []
    try:
        with lock:
            df_xl = pd.read_excel(excel_file, engine="openpyxl")

        df_xl.columns = df_xl.columns.str.strip()
        col_map = {c.lower(): c for c in df_xl.columns}
        phone_col = col_map.get("phone number")

        for _, row in df_xl.iterrows():
            name_v = _clean_val(row.get(col_map.get("name")))
            phone_v = _clean_val(row.get(phone_col)) if phone_col else ""
            reg_v = _clean_val(row.get(col_map.get("registration number")))

            if name_v and phone_v and reg_v:
                row_data = {c_orig: _clean_val(row.get(c_orig)) for c_low, c_orig in col_map.items()}

                # Subgroup filter
                if target_column and target_subgroup and target_subgroup.lower() != "all":
                    col_val = row_data.get(target_column)
                    if not col_val or str(col_val).strip().lower() != target_subgroup.lower().strip():
                        continue

                attendees.append({
                    "name": name_v,
                    "phone": phone_v,
                    "reg_no": reg_v,
                    "row_dict": row_data
                })
    except Exception as e:
        with _sms_sending_lock:
            _sms_progress["current_status"] = "Failed to load attendees"
            _sms_progress["logs"].append(f"\u274c Error loading attendees: {str(e)}")
            _sms_sending_active = False
        socketio.emit("sms_progress_update", _sms_progress)
        return

    total_count = len(attendees)
    with _sms_sending_lock:
        _sms_progress["total"] = total_count
        _sms_progress["sent"] = 0
        _sms_progress["skipped"] = 0
        _sms_progress["failed"] = 0
        _sms_progress["current_status"] = "Starting SMS campaign..."
        _sms_progress["logs"] = [f"\U0001f680 Started SMS campaign ({sms_provider}). Total recipients: {total_count}"]
    socketio.emit("sms_progress_update", _sms_progress)

    for idx, att in enumerate(attendees):
        name = att["name"]
        phone = att["phone"]
        reg = att["reg_no"]
        row_dict = att["row_dict"]

        with _sms_sending_lock:
            _sms_progress["current_phone"] = phone
            _sms_progress["current_status"] = f"Processing {phone} ({idx + 1}/{total_count})"
        socketio.emit("sms_progress_update", _sms_progress)

        # Resolve per-subgroup SMS template
        from_cfg_sms_template = cfg.get("sms_template", "")
        sms_template = _get_attendee_template(cfg, row_dict, "sms_template", from_cfg_sms_template)
        qr_url = f"{host_url}qrcodes/{reg}.png"
        extra = {"Event": event_name, "QR_URL": qr_url}
        body_text = format_template(sms_template, row_dict, extra)

        success, msg = _send_single_sms(phone, body_text, sms_provider, cfg)

        if success:
            log_msg = f"\u2705 SMS sent to {name} ({phone})"
            _update_sent_status(reg, "sms", "Sent")
            with _sms_sending_lock:
                _sms_progress["sent"] += 1
                _sms_progress["logs"].append(log_msg)
        else:
            log_msg = f"\u274c Failed to send to {name} ({phone}): {msg}"
            _update_sent_status(reg, "sms", "Failed")
            with _sms_sending_lock:
                _sms_progress["failed"] += 1
                _sms_progress["logs"].append(log_msg)

        socketio.emit("sms_progress_update", _sms_progress)
        time.sleep(1.2)  # Polite rate-limit between sends

    with _sms_sending_lock:
        _sms_progress["current_status"] = "Finished"
        _sms_progress["current_phone"] = ""
        _sms_progress["logs"].append(
            f"\U0001f3c1 SMS Campaign finished. Sent: {_sms_progress['sent']}, "
            f"Skipped: {_sms_progress['skipped']}, Failed: {_sms_progress['failed']}."
        )
        _sms_sending_active = False
    socketio.emit("sms_progress_update", _sms_progress)


@app.route("/send_sms_bulk", methods=["POST"])
def send_sms_bulk():
    global _sms_sending_active
    if not request.is_json:
        return jsonify(message="Request must be JSON."), 415

    payload = request.json or {}
    sms_provider = payload.get("sms_provider", "android").strip()

    cfg = get_event_config()
    cfg["sms_provider"] = sms_provider
    cfg["event_name_template"] = payload.get("event_name", "the Event").strip()

    if sms_provider == "android":
        android_ip = payload.get("android_gateway_ip", "").strip()
        android_port = payload.get("android_gateway_port", "8080").strip()
        if not android_ip:
            return jsonify(message="Android Gateway IP address is required."), 400
        cfg["android_gateway_url"] = f"{android_ip}:{android_port}"
    elif sms_provider == "twilio":
        cfg["twilio_sms_sid"] = payload.get("twilio_sms_sid", "").strip()
        cfg["twilio_sms_token"] = payload.get("twilio_sms_token", "").strip()
        cfg["twilio_sms_sender"] = payload.get("twilio_sms_sender", "").strip()
        if not cfg["twilio_sms_sid"] or not cfg["twilio_sms_token"] or not cfg["twilio_sms_sender"]:
            return jsonify(message="Twilio SID, Token, and Sender number are required."), 400
    else:
        return jsonify(message=f"Unknown SMS provider: {sms_provider}"), 400

    target_column = payload.get("target_column", "").strip()
    target_subgroup = payload.get("target_subgroup", "").strip()
    if target_column:
        cfg["group_column"] = target_column

    save_event_config(cfg)

    with _sms_sending_lock:
        if _sms_sending_active:
            return jsonify(message="An SMS campaign is already in progress."), 409
        _sms_sending_active = True

    host_url = request.host_url
    event_name = cfg["event_name_template"]

    threading.Thread(
        target=_run_sms_campaign,
        args=(sms_provider, event_name, host_url, target_column, target_subgroup),
        daemon=True
    ).start()

    return jsonify(message="SMS campaign started successfully."), 202


@app.route("/sms_status")
def get_sms_status():
    with _sms_sending_lock:
        return jsonify(
            active=_sms_sending_active,
            sent=_sms_progress["sent"],
            skipped=_sms_progress["skipped"],
            failed=_sms_progress["failed"],
            total=_sms_progress["total"],
            current_phone=_sms_progress["current_phone"],
            current_status=_sms_progress["current_status"],
            logs=_sms_progress["logs"]
        )


@app.route("/download/excel")
def download_excel():
    with highlight_lock:
        high_file = get_highlighted_file()
        if os.path.exists(high_file):
            return send_file(
                high_file,
                as_attachment=True,
                download_name="event_checkin_highlighted.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    return jsonify(message="Highlighted Excel report not generated yet."), 404


@app.route("/download/csv")
def download_csv():
    with lock:
        log_file_path = get_log_file()
        if os.path.exists(log_file_path):
            return send_file(
                log_file_path,
                as_attachment=True,
                download_name="scanned_log.csv",
                mimetype="text/csv"
            )
    return jsonify(message="No scan log found."), 404


@app.route("/export_audit_log")
def export_audit_log():
    with lock:
        audit_file = os.path.join(get_active_event_path(), "audit_log.csv")
        if not os.path.exists(audit_file):
            _log_audit("System", "Initialised audit log", "System")
        
        return send_file(
            audit_file,
            as_attachment=True,
            download_name="audit_log.csv",
            mimetype="text/csv"
        )




@app.route("/download/id_card/<reg_no>")
def download_id_card(reg_no):
    reg_no = reg_no.strip()
    excel_file = get_excel_file()
    if not os.path.exists(excel_file):
        return jsonify(message="Event spreadsheet missing."), 404
        
    try:
        with lock:
            df_xl = pd.read_excel(excel_file, engine="openpyxl")
        df_xl.columns = df_xl.columns.astype(str).str.strip()
        col_map = {c.lower(): c for c in df_xl.columns}
        
        reg_col = col_map.get("registration number")
        if not reg_col:
            return jsonify(message="Registration number column missing in spreadsheet."), 500
            
        row_match = df_xl[df_xl[reg_col].astype(str).str.strip().str.lower() == reg_no.lower()]
        if row_match.empty:
            return jsonify(message=f"Attendee '{reg_no}' not found."), 404
            
        row = row_match.iloc[0]
        row_dict = {c_orig: _clean_val(row.get(c_orig)) for c_low, c_orig in col_map.items()}
        name = _clean_val(row.get(col_map.get("name")))
        phone = _clean_val(row.get(col_map.get("phone number")))
        email = _clean_val(row.get(col_map.get("email address")))
        uid = _clean_val(row.get(col_map.get("unique id")))
        
        level = ""
        cfg = get_event_config()
        group_col = cfg.get("group_column", "")
        if group_col and group_col in row_dict:
            level = row_dict[group_col]
        else:
            for k, v in row_dict.items():
                if "level" in k.lower() or "subgroup" in k.lower():
                    level = v
                    break
        
        # Resolve group/subgroup specific overrides for download
        designer_settings = cfg.get("id_card_designer_settings") or {}
        if reg_no:
            try:
                groups = _load_groups()
                for g in groups:
                    if reg_no in g.get("reg_nos", []):
                        if "id_card_designer_settings" in g:
                            designer_settings = g["id_card_designer_settings"]
                            break
            except Exception:
                pass
                
        group_col_val = cfg.get("group_column", "").strip()
        if group_col_val:
            col_val = None
            for k, v in row_dict.items():
                if str(k).strip().lower() == group_col_val.lower().strip():
                    col_val = str(v).strip()
                    break
            if col_val and col_val.lower() not in ["nan", "none", ""]:
                sub_templates = cfg.get("subgroup_templates", {})
                key = f"{group_col_val}:{col_val}"
                if key in sub_templates:
                    sub_tpl = sub_templates[key]
                    if "id_card_designer_settings" in sub_tpl:
                        designer_settings = sub_tpl["id_card_designer_settings"]
                        
        resolved_cfg = cfg.copy()
        resolved_cfg["id_card_designer_settings"] = designer_settings
        
        qr_path = os.path.join(get_qr_dir(), f"{reg_no}.png")
        if not os.path.exists(qr_path):
            _generate_qr_for_guest(reg_no, reg_no)
            
        pdf_path, png_path = _generate_id_card(
            name=name,
            reg_no=reg_no,
            phone=phone,
            email=email,
            uid=uid,
            qr_path=qr_path,
            event_name=cfg.get("event_name_template", "the Event"),
            level=level,
            cfg=resolved_cfg
        )
        
        return send_file(pdf_path, as_attachment=True, download_name=f"id_card_{reg_no}.pdf")
    except Exception as e:
        return jsonify(message=f"Error generating ID card: {str(e)}"), 500


@app.route("/list_id_templates", methods=["GET"])
def list_id_templates():
    templates_dir = os.path.join(BASE_DIR, "static", "id_templates")
    if not os.path.exists(templates_dir):
        return jsonify([])
    files = [f for f in os.listdir(templates_dir) if f.endswith(".json")]
    templates = []
    for f in files:
        try:
            with open(os.path.join(templates_dir, f), "r", encoding="utf-8") as file:
                data = json.load(file)
                templates.append({
                    "filename": f,
                    "name": data.get("name", f),
                    "html": data.get("html", ""),
                    "css": data.get("css", "")
                })
        except Exception:
            pass
    return jsonify(templates)


@app.route("/preview_id_card", methods=["GET", "POST"])
def preview_id_card():
    theme = "cyber_neon"
    header = ""
    footer = ""
    show_reg = True
    show_email = True
    show_phone = True
    show_uid = True
    show_pass = True
    label_reg = "REGISTRATION:"
    label_email = "EMAIL:"
    label_phone = "PHONE:"
    label_uid = "UNIQUE ID:"
    label_pass = "PASS TYPE:"
    
    if request.method == "POST":
        # Parse visual designer layout settings from JSON payload
        designer_payload = request.json or {}
        mock_cfg = {
            "id_card_designer_settings": designer_payload,
            "enable_id_card_generation": True
        }
    else:
        # GET request: load from query params as before
        theme = request.args.get("theme", "cyber_neon").strip()
        header = request.args.get("header", "").strip()
        footer = request.args.get("footer", "").strip()
        show_reg = request.args.get("show_reg", "true").lower() == "true"
        show_email = request.args.get("show_email", "true").lower() == "true"
        show_phone = request.args.get("show_phone", "true").lower() == "true"
        show_uid = request.args.get("show_uid", "true").lower() == "true"
        show_pass = request.args.get("show_pass", "true").lower() == "true"
        label_reg = request.args.get("label_reg", "REGISTRATION:").strip()
        label_email = request.args.get("label_email", "EMAIL:").strip()
        label_phone = request.args.get("label_phone", "PHONE:").strip()
        label_uid = request.args.get("label_uid", "UNIQUE ID:").strip()
        label_pass = request.args.get("label_pass", "PASS TYPE:").strip()
        mock_cfg = {
            "id_card_theme": theme,
            "id_card_header": header,
            "id_card_footer": footer,
            "id_card_show_reg": show_reg,
            "id_card_show_email": show_email,
            "id_card_show_phone": show_phone,
            "id_card_show_uid": show_uid,
            "id_card_show_pass": show_pass,
            "id_card_label_reg": label_reg,
            "id_card_label_email": label_email,
            "id_card_label_phone": label_phone,
            "id_card_label_uid": label_uid,
            "id_card_label_pass": label_pass,
            "enable_id_card_generation": True
        }
        
    qr_dir = get_qr_dir()
    os.makedirs(qr_dir, exist_ok=True)
    dummy_qr = os.path.join(qr_dir, "dummy_preview.png")
    if not os.path.exists(dummy_qr):
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data("John Doe | DEMO12345")
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img.save(dummy_qr)
        
    try:
        pdf_path, png_path = _generate_id_card(
            name="John Doe",
            reg_no="DEMO12345",
            phone="+91 98765 43210",
            email="johndoe@example.com",
            uid="ABC123XY",
            qr_path=dummy_qr,
            event_name="SAMPLE CONFERENCE",
            level="VIP",
            cfg=mock_cfg
        )
        return send_file(png_path, mimetype="image/png")
    except Exception as e:
        return jsonify(message=f"Error generating preview: {str(e)}"), 500


# ── Subgroup Template Resolver Helper ──────────────────────────────────────────
def _get_attendee_template(cfg, row_dict, template_type, default_val):
    reg_no = ""
    for k, v in row_dict.items():
        if str(k).lower().strip() in ["registration number", "reg_no", "reg_number"]:
            reg_no = str(v).strip()
            break
            
    # Check custom groups first (since they are specific selections)
    if reg_no:
        try:
            groups = _load_groups()
            for g in groups:
                if reg_no in g.get("reg_nos", []):
                    val = g.get(template_type)
                    if val and str(val).strip():
                        return val
        except Exception:
            pass

    # Check subgroups second
    group_col = cfg.get("group_column", "").strip()
    if group_col:
        col_val = None
        for k, v in row_dict.items():
            if str(k).strip().lower() == group_col.lower().strip():
                col_val = str(v).strip()
                break
        if col_val and col_val.lower() not in ["nan", "none", ""]:
            sub_templates = cfg.get("subgroup_templates", {})
            key = f"{group_col}:{col_val}"
            if key in sub_templates:
                tpl = sub_templates[key].get(template_type)
                if tpl and tpl.strip():
                    return tpl
    return default_val



# ── Individual / Bulk SMS sender helper ─────────────────────────────────────────
def _send_single_sms(phone, message, provider, cfg) -> tuple[bool, str]:
    import requests
    clean_phone = _normalize_phone(phone)
    if provider == "twilio":
        sid = cfg.get("twilio_sms_sid")
        token = cfg.get("twilio_sms_token")
        sender = cfg.get("twilio_sms_sender")
        if not sid or not token or not sender:
            return False, "Twilio SMS configuration missing"
        try:
            from twilio.rest import Client
            client = Client(sid, token)
            client.messages.create(body=message, from_=sender, to=clean_phone)
            return True, "Sent via Twilio"
        except Exception as e:
            return False, f"Twilio SMS error: {str(e)}"
    elif provider == "android":
        ip_port = cfg.get("android_gateway_url", "").strip()
        if not ip_port:
            return False, "Android Gateway IP:Port missing"
        ip_port = ip_port.replace("http://", "").replace("https://", "")

        url_msg = f"http://{ip_port}/message"
        payload_msg = {
            "textMessage": {
                "text": message,
                "phoneNumbers": [clean_phone]
            }
        }
        
        url_send = f"http://{ip_port}/send"
        payload_send = {
            "phone": clean_phone,
            "message": message
        }
        
        # 1. Try Capcom6 style
        try:
            r = requests.post(url_msg, json=payload_msg, timeout=5)
            if r.status_code in [200, 201, 202]:
                return True, "Sent via Android (Capcom6)"
        except Exception:
            pass
            
        # 2. Try Standard POST style
        try:
            r = requests.post(url_send, json=payload_send, timeout=5)
            if r.status_code in [200, 201, 202]:
                return True, "Sent via Android (General)"
        except Exception:
            pass
            
        # 3. Try Standard GET query style
        try:
            r = requests.get(url_send, params={"phone": clean_phone, "message": message}, timeout=5)
            if r.status_code in [200, 201, 202]:
                return True, "Sent via Android (GET)"
        except Exception as e:
            return False, f"Android Gateway connection failed: {str(e)}"
            
        return False, "Android Gateway returned failure code"
    return False, f"Unknown SMS provider: {provider}"


# ── Individual / Bulk WhatsApp sender helper ────────────────────────────────────
def _send_single_whatsapp(phone, name, reg_no, row_dict, cfg, host_url) -> tuple[bool, str, dict]:
    provider = cfg.get("wa_provider", "manual")
    event_name = cfg.get("event_name_template", "the Event")
    qr_url = f"{host_url}qrcodes/{reg_no}.png"
    
    from_cfg_whatsapp_template = cfg.get("whatsapp_template", "")
    whatsapp_template = _get_attendee_template(cfg, row_dict, "whatsapp_template", from_cfg_whatsapp_template)
    
    extra = {"Event": event_name, "QR_URL": qr_url}
    body_text = format_template(whatsapp_template, row_dict, extra)
    
    clean_phone = _normalize_phone(phone)
    if provider == "manual":
        return True, body_text, {"method": "manual", "phone": clean_phone}
        
    elif provider == "twilio":
        sid = cfg.get("twilio_sid")
        token = cfg.get("twilio_token")
        sender = cfg.get("twilio_sender")
        if not sid or not token or not sender:
            return False, "Twilio configuration missing", {}
            
        if not sender.startswith("+") and not sender.startswith("whatsapp:"):
            sender = "+" + sender
            
        try:
            from twilio.rest import Client
            client = Client(sid, token)
            from_number = f"whatsapp:{sender}"
            to_number = f"whatsapp:{clean_phone}"
            
            is_local = "localhost" in host_url or "127.0.0.1" in host_url
            if is_local:
                client.messages.create(
                    body=body_text + f"\n\nDownload QR at: {qr_url}",
                    from_=from_number,
                    to=to_number
                )
            else:
                client.messages.create(
                    body=body_text,
                    media_url=[qr_url],
                    from_=from_number,
                    to=to_number
                )
            return True, "Sent via Twilio", {}
        except Exception as e:
            return False, f"Twilio WhatsApp error: {str(e)}", {}
            
    elif provider == "meta":
        token = cfg.get("meta_access_token")
        phone_id = cfg.get("meta_phone_number_id")
        template_name = cfg.get("meta_template_name")
        lang_code = cfg.get("meta_lang_code", "en_US")
        
        if not token or not phone_id or not template_name:
            return False, "Meta Cloud API configuration missing", {}
            
        import requests
        url = f"https://graph.facebook.com/v19.0/{phone_id}/messages"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        meta_phone = clean_phone.replace("+", "")
        
        payload = {
            "messaging_product": "whatsapp",
            "recipient_type": "individual",
            "to": meta_phone,
            "type": "template",
            "template": {
                "name": template_name,
                "language": {
                    "code": lang_code
                },
                "components": [
                    {
                        "type": "body",
                        "parameters": [
                            {"type": "text", "text": name},
                            {"type": "text", "text": event_name},
                            {"type": "text", "text": reg_no},
                            {"type": "text", "text": qr_url}
                        ]
                    }
                ]
            }
        }
        
        try:
            r = requests.post(url, json=payload, headers=headers, timeout=10)
            res_json = r.json()
            if r.status_code in [200, 201, 202] and "messages" in res_json:
                return True, "Sent via Meta API", {}
            else:
                err_msg = res_json.get("error", {}).get("message", "Unknown Meta error")
                return False, f"Meta API error: {err_msg}", {}
        except Exception as e:
            return False, f"Meta request failed: {str(e)}", {}
            
    return False, f"Unknown WhatsApp provider: {provider}", {}


# ── Single Send Messaging Helpers & Endpoints ─────────────────────────────────────
def _resolve_email_attachments(reg_no, row_dict, cfg):
    # 1. Start with global designer settings
    designer_settings = cfg.get("id_card_designer_settings") or {}
    
    # 2. Check custom named groups first (if this attendee belongs to a custom group)
    if reg_no:
        try:
            groups = _load_groups()
            for g in groups:
                if reg_no in g.get("reg_nos", []):
                    if "id_card_designer_settings" in g:
                        designer_settings = g["id_card_designer_settings"]
                        break
        except Exception:
            pass
            
    # 3. Check subgroups second (if group column is set and matches attendee value)
    group_col = cfg.get("group_column", "").strip()
    if group_col:
        col_val = None
        for k, v in row_dict.items():
            if str(k).strip().lower() == group_col.lower().strip():
                col_val = str(v).strip()
                break
        if col_val and col_val.lower() not in ["nan", "none", ""]:
            sub_templates = cfg.get("subgroup_templates", {})
            key = f"{group_col}:{col_val}"
            if key in sub_templates:
                sub_tpl = sub_templates[key]
                if "id_card_designer_settings" in sub_tpl:
                    designer_settings = sub_tpl["id_card_designer_settings"]

    attachment_type = designer_settings.get("attachment_type", "qr")
    
    # Clone and inject the resolved designer settings into cfg
    resolved_cfg = cfg.copy()
    resolved_cfg["id_card_designer_settings"] = designer_settings
    
    qr_dir = get_qr_dir()
    qr_file = os.path.join(qr_dir, f"{reg_no}.png")
    
    # We do case-insensitive search in row_dict keys
    def get_val_ci(keys_to_try):
        for k, v in row_dict.items():
            if str(k).lower().strip() in keys_to_try:
                return _clean_val(v)
        return ""
        
    name = get_val_ci(["name", "full name"])
    phone = get_val_ci(["phone", "phone number", "mobile"])
    email = get_val_ci(["email", "email address"])
    uid = get_val_ci(["unique id", "uid", "id"])
    level = get_val_ci(["pass type", "level", "role", "designation"])
    event_name = cfg.get("event_name_template", "the Event")
    
    # We need to make sure the QR code exists
    if not os.path.exists(qr_file):
        # Generate dummy or dynamic QR code if missing
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(f"{name} | {reg_no}")
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img.save(qr_file)
        
    # Generate ID card files if needed
    pdf_path = None
    png_path = None
    if attachment_type in ("id_pdf", "id_png", "id_both", "all"):
        try:
            pdf_path, png_path = _generate_id_card(
                name=name,
                reg_no=reg_no,
                phone=phone,
                email=email,
                uid=uid,
                qr_path=qr_file,
                event_name=event_name,
                level=level,
                cfg=resolved_cfg
            )
        except Exception as e:
            print(f"Error generating ID card for attachment: {str(e)}")
            
    # Return file list based on attachment_type
    if attachment_type == "id_pdf" and pdf_path and os.path.exists(pdf_path):
        return [pdf_path]
    elif attachment_type == "id_png" and png_path and os.path.exists(png_path):
        return [png_path]
    elif attachment_type == "id_both" and pdf_path and png_path and os.path.exists(pdf_path) and os.path.exists(png_path):
        return [pdf_path, png_path]
    elif attachment_type == "all" and pdf_path and png_path and os.path.exists(pdf_path) and os.path.exists(png_path):
        return [qr_file, pdf_path, png_path]
    else:
        # Default fallback is always just QR code
        return [qr_file]


def _send_single_email_helper(reg_no, cfg) -> tuple[bool, str]:
    excel_file = get_excel_file()
    if not os.path.exists(excel_file):
        return False, "Registration spreadsheet missing."
    sender_email = cfg.get("email_sender")
    app_password = cfg.get("email_password")
    event_name = cfg.get("event_name_template", "the Event")
    if not sender_email or not app_password:
        return False, "Gmail SMTP credentials are not configured in settings."
    try:
        with lock:
            df_xl = pd.read_excel(excel_file, engine="openpyxl")
        df_xl.columns = df_xl.columns.astype(str).str.strip()
        col_map = {c.lower(): c for c in df_xl.columns}
        reg_col = col_map.get("registration number")
        if not reg_col:
            return False, "Registration number column missing in sheet."
        row_match = df_xl[df_xl[reg_col].astype(str).str.strip().str.lower() == reg_no.lower()]
        if row_match.empty:
            return False, f"Attendee with Registration Number '{reg_no}' not found."
        row = row_match.iloc[0]
        row_dict = {c_orig: _clean_val(row.get(c_orig)) for c_low, c_orig in col_map.items()}
        recipient = _clean_val(row.get(col_map.get("email address")))
        if not recipient:
            return False, "Email address missing for this attendee."
            
        import yagmail
        qr_dir = get_qr_dir()
        qr_file = os.path.join(qr_dir, f"{reg_no}.png")
        if not os.path.exists(qr_file):
            # Generate dummy or dynamic QR code if missing
            qr = qrcode.QRCode(version=1, box_size=10, border=4)
            qr.add_data(f"{row_dict.get('name', 'Attendee')} | {reg_no}")
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            img.save(qr_file)
            
        from_cfg_email_subject = cfg.get("email_subject", "Your Event QR Code")
        from_cfg_email_template = cfg.get("email_template", "")
        subject = _get_attendee_template(cfg, row_dict, "email_subject", from_cfg_email_subject)
        email_template = _get_attendee_template(cfg, row_dict, "email_template", from_cfg_email_template)
        extra = {"Event": event_name}
        body = format_template(email_template, row_dict, extra)
        subject_formatted = format_template(subject, row_dict, extra)
        
        yag = yagmail.SMTP(user=sender_email, password=app_password)
        attachments_to_send = _resolve_email_attachments(reg_no, row_dict, cfg)
        yag.send(to=recipient, subject=subject_formatted, contents=body, attachments=attachments_to_send)
        yag.close()
        _update_sent_status(reg_no, "email", "Sent")
        return True, f"Email sent successfully to {recipient}!"
    except Exception as e:
        _update_sent_status(reg_no, "email", "Failed")
        return False, f"Error sending email: {str(e)}"

def _send_single_whatsapp_helper(reg_no, cfg, host_url) -> tuple[bool, str, dict]:
    excel_file = get_excel_file()
    if not os.path.exists(excel_file):
        return False, "Registration spreadsheet missing.", {}
    try:
        with lock:
            df_xl = pd.read_excel(excel_file, engine="openpyxl")
        df_xl.columns = df_xl.columns.astype(str).str.strip()
        col_map = {c.lower(): c for c in df_xl.columns}
        reg_col = col_map.get("registration number")
        if not reg_col:
            return False, "Registration number column missing in sheet.", {}
        row_match = df_xl[df_xl[reg_col].astype(str).str.strip().str.lower() == reg_no.lower()]
        if row_match.empty:
            return False, f"Attendee '{reg_no}' not found.", {}
        row = row_match.iloc[0]
        row_dict = {c_orig: _clean_val(row.get(c_orig)) for c_low, c_orig in col_map.items()}
        name = _clean_val(row.get(col_map.get("name")))
        phone = _clean_val(row.get(col_map.get("phone number")))
        if not phone:
            return False, "Phone number missing for this attendee.", {}
        success, msg, extra_data = _send_single_whatsapp(phone, name, reg_no, row_dict, cfg, host_url)
        if success:
            if extra_data.get("method") != "manual":
                _update_sent_status(reg_no, "whatsapp", "Sent")
        else:
            _update_sent_status(reg_no, "whatsapp", "Failed")
        return success, msg, extra_data
    except Exception as e:
        return False, f"Error sending WhatsApp: {str(e)}", {}

def _send_single_sms_helper(reg_no, cfg, host_url) -> tuple[bool, str]:
    excel_file = get_excel_file()
    if not os.path.exists(excel_file):
        return False, "Registration spreadsheet missing."
    provider = cfg.get("sms_provider", "android")
    event_name = cfg.get("event_name_template", "the Event")
    try:
        with lock:
            df_xl = pd.read_excel(excel_file, engine="openpyxl")
        df_xl.columns = df_xl.columns.astype(str).str.strip()
        col_map = {c.lower(): c for c in df_xl.columns}
        reg_col = col_map.get("registration number")
        if not reg_col:
            return False, "Registration number column missing in sheet."
        row_match = df_xl[df_xl[reg_col].astype(str).str.strip().str.lower() == reg_no.lower()]
        if row_match.empty:
            return False, f"Attendee '{reg_no}' not found."
        row = row_match.iloc[0]
        row_dict = {c_orig: _clean_val(row.get(c_orig)) for c_low, c_orig in col_map.items()}
        phone = _clean_val(row.get(col_map.get("phone number")))
        if not phone:
            return False, "Phone number missing for this attendee."
        from_cfg_sms_template = cfg.get("sms_template", "")
        sms_template = _get_attendee_template(cfg, row_dict, "sms_template", from_cfg_sms_template)
        extra = {"Event": event_name, "QR_URL": f"{host_url}qrcodes/{reg_no}.png"}
        body_text = format_template(sms_template, row_dict, extra)
        success, msg = _send_single_sms(phone, body_text, provider, cfg)
        if success:
            _update_sent_status(reg_no, "sms", "Sent")
        else:
            _update_sent_status(reg_no, "sms", "Failed")
        return success, msg
    except Exception as e:
        return False, f"Error sending SMS: {str(e)}"

@app.route("/send_email_single", methods=["POST"])
def send_email_single():
    if not request.is_json:
        return jsonify(success=False, message="Request must be JSON."), 415
    payload = request.json or {}
    reg_no = payload.get("reg_no", "").strip()
    if not reg_no:
        return jsonify(success=False, message="Registration number is required."), 400
    cfg = get_event_config()
    success, msg = _send_single_email_helper(reg_no, cfg)
    return jsonify(success=success, message=msg), (200 if success else 400)

@app.route("/send_whatsapp_single", methods=["POST"])
def send_whatsapp_single():
    if not request.is_json:
        return jsonify(success=False, message="Request must be JSON."), 415
    payload = request.json or {}
    reg_no = payload.get("reg_no", "").strip()
    if not reg_no:
        return jsonify(success=False, message="Registration number is required."), 400
    cfg = get_event_config()
    success, msg, extra_data = _send_single_whatsapp_helper(reg_no, cfg, request.host_url)
    if success:
        if extra_data.get("method") == "manual":
            return jsonify(success=True, method="manual", phone=extra_data["phone"], message=msg)
        return jsonify(success=True, message=msg)
    return jsonify(success=False, message=msg), 500

@app.route("/send_sms_single", methods=["POST"])
def send_sms_single():
    if not request.is_json:
        return jsonify(success=False, message="Request must be JSON."), 415
    payload = request.json or {}
    reg_no = payload.get("reg_no", "").strip()
    if not reg_no:
        return jsonify(success=False, message="Registration number is required."), 400
    cfg = get_event_config()
    success, msg = _send_single_sms_helper(reg_no, cfg, request.host_url)
    return jsonify(success=success, message=msg), (200 if success else 500)
load_active_event()


# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port  = int(os.environ.get("PORT", 5001))
    
    # Gracefully free the port if occupied
    import os, subprocess, signal
    try:
        if os.name == 'nt':
            cmd = f"netstat -ano | findstr :{port}"
            out = subprocess.check_output(cmd, shell=True).decode('utf-8', errors='ignore')
            pids = set()
            for line in out.strip().split('\n'):
                line = line.strip()
                if not line:
                    continue
                parts = line.split()
                if len(parts) >= 5 and "LISTENING" in parts:
                    pid = parts[-1]
                    try:
                        pids.add(int(pid))
                    except ValueError:
                        pass
            my_pid = os.getpid()
            for pid in pids:
                if pid != my_pid:
                    print(f"[*] Port {port} occupied by PID {pid}. Terminating it for a clean start...")
                    subprocess.run(f"taskkill /F /PID {pid}", shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    time.sleep(1.0)
        else:
            subprocess.run(f"fuser -k -n tcp {port}", shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            time.sleep(1.0)
    except Exception as e:
        print(f"[*] Note: Port check completed: {e}")

    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    lan_ip = get_lan_ip()
    print("")
    print("=== QR Check-In System is starting ===")
    print("")
    print(f"   Local   -> http://localhost:{port}")
    print(f"   Network -> http://{lan_ip}:{port}  (share this with other devices)")
    print("")
    print("   All devices on the same Wi-Fi can open the network URL above.")
    print("")
    socketio.run(app, debug=debug, host="0.0.0.0", port=port)
